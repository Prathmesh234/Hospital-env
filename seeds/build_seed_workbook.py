"""Generate Hospital-env seed xlsx workbooks.

Produces:
- ``seeds/hospital_seed.xlsx`` — single consolidated workbook with every sheet
  in canonical load order. Feed this to ``uv run hospital-env load``.
- ``seeds/by_domain/*.xlsx`` — one workbook per data domain (catalogs,
  organization, patients, coverage, scheduling, encounters, clinical,
  medications, labs_imaging, billing, communications, operations). The loader
  also accepts the ``seeds/by_domain/`` directory directly.

All IDs are deterministic ``uuid5`` values derived from human-readable keys so
that re-running the script produces byte-identical files and so cross-domain
references (e.g. "this prescription belongs to patient Margaret Johnson") are
easy to follow when reading the data.

The dummy data is hand-curated to be clinically and administratively realistic:
- Real ICD-10-CM, CPT, LOINC, RxNorm, SNOMED CT codes
- Realistic lab reference ranges and abnormal flags
- Realistic CARC/RARC denial codes and 837 claim structures
- A 10-patient panel covering primary care, ED, inpatient, and OB scenarios
"""

from __future__ import annotations

import json
from datetime import date, datetime, time, timedelta
from pathlib import Path
from uuid import NAMESPACE_OID, uuid5

from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent
NS = uuid5(NAMESPACE_OID, "hospital-env.springfield-regional-medical-center")


def uid(kind: str, name: str) -> str:
    return str(uuid5(NS, f"{kind}:{name}"))


def iso(d: date | datetime) -> str:
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%dT%H:%M:%SZ")
    return d.isoformat()


def ts(y: int, m: int, d: int, h: int = 0, mi: int = 0, s: int = 0) -> str:
    return iso(datetime(y, m, d, h, mi, s))


# ---------------------------------------------------------------------------
# 1. CATALOG DATA — real code systems
# ---------------------------------------------------------------------------

ICD10_CODES = [
    {"code": "E11.9",  "description": "Type 2 diabetes mellitus without complications",                                  "chapter": "IV. Endocrine, nutritional and metabolic diseases",       "is_billable": True},
    {"code": "E11.21", "description": "Type 2 diabetes mellitus with diabetic nephropathy",                              "chapter": "IV. Endocrine, nutritional and metabolic diseases",       "is_billable": True},
    {"code": "E11.65", "description": "Type 2 diabetes mellitus with hyperglycemia",                                     "chapter": "IV. Endocrine, nutritional and metabolic diseases",       "is_billable": True},
    {"code": "E78.5",  "description": "Hyperlipidemia, unspecified",                                                     "chapter": "IV. Endocrine, nutritional and metabolic diseases",       "is_billable": True},
    {"code": "E03.9",  "description": "Hypothyroidism, unspecified",                                                     "chapter": "IV. Endocrine, nutritional and metabolic diseases",       "is_billable": True},
    {"code": "E55.9",  "description": "Vitamin D deficiency, unspecified",                                               "chapter": "IV. Endocrine, nutritional and metabolic diseases",       "is_billable": True},
    {"code": "I10",    "description": "Essential (primary) hypertension",                                                "chapter": "IX. Diseases of the circulatory system",                  "is_billable": True},
    {"code": "I25.10", "description": "Atherosclerotic heart disease of native coronary artery without angina pectoris", "chapter": "IX. Diseases of the circulatory system",                  "is_billable": True},
    {"code": "I21.4",  "description": "Non-ST elevation (NSTEMI) myocardial infarction",                                 "chapter": "IX. Diseases of the circulatory system",                  "is_billable": True},
    {"code": "I50.32", "description": "Chronic diastolic (congestive) heart failure",                                    "chapter": "IX. Diseases of the circulatory system",                  "is_billable": True},
    {"code": "I48.91", "description": "Unspecified atrial fibrillation",                                                 "chapter": "IX. Diseases of the circulatory system",                  "is_billable": True},
    {"code": "J44.1",  "description": "Chronic obstructive pulmonary disease with (acute) exacerbation",                 "chapter": "X. Diseases of the respiratory system",                    "is_billable": True},
    {"code": "J45.40", "description": "Moderate persistent asthma, uncomplicated",                                       "chapter": "X. Diseases of the respiratory system",                    "is_billable": True},
    {"code": "J45.901","description": "Unspecified asthma with (acute) exacerbation",                                    "chapter": "X. Diseases of the respiratory system",                    "is_billable": True},
    {"code": "J18.9",  "description": "Pneumonia, unspecified organism",                                                 "chapter": "X. Diseases of the respiratory system",                    "is_billable": True},
    {"code": "J20.9",  "description": "Acute bronchitis, unspecified",                                                   "chapter": "X. Diseases of the respiratory system",                    "is_billable": True},
    {"code": "K21.9",  "description": "Gastro-esophageal reflux disease without esophagitis",                            "chapter": "XI. Diseases of the digestive system",                     "is_billable": True},
    {"code": "N18.3",  "description": "Chronic kidney disease, stage 3 (moderate)",                                      "chapter": "XIV. Diseases of the genitourinary system",                "is_billable": True},
    {"code": "N39.0",  "description": "Urinary tract infection, site not specified",                                     "chapter": "XIV. Diseases of the genitourinary system",                "is_billable": True},
    {"code": "F32.9",  "description": "Major depressive disorder, single episode, unspecified",                          "chapter": "V. Mental, behavioral and neurodevelopmental disorders",   "is_billable": True},
    {"code": "F41.1",  "description": "Generalized anxiety disorder",                                                    "chapter": "V. Mental, behavioral and neurodevelopmental disorders",   "is_billable": True},
    {"code": "G30.9",  "description": "Alzheimer's disease, unspecified",                                                "chapter": "VI. Diseases of the nervous system",                       "is_billable": True},
    {"code": "M54.50", "description": "Low back pain, unspecified",                                                      "chapter": "XIII. Diseases of the musculoskeletal system",             "is_billable": True},
    {"code": "S72.001A","description": "Fracture of unspecified part of neck of right femur, initial encounter",         "chapter": "XIX. Injury, poisoning, external causes",                  "is_billable": True},
    {"code": "R07.9",  "description": "Chest pain, unspecified",                                                         "chapter": "XVIII. Symptoms, signs and abnormal findings",             "is_billable": True},
    {"code": "R10.9",  "description": "Unspecified abdominal pain",                                                      "chapter": "XVIII. Symptoms, signs and abnormal findings",             "is_billable": True},
    {"code": "R51.9",  "description": "Headache, unspecified",                                                           "chapter": "XVIII. Symptoms, signs and abnormal findings",             "is_billable": True},
    {"code": "Z00.00", "description": "Encounter for general adult medical examination without abnormal findings",       "chapter": "XXI. Factors influencing health status",                   "is_billable": True},
    {"code": "Z23",    "description": "Encounter for immunization",                                                      "chapter": "XXI. Factors influencing health status",                   "is_billable": True},
    {"code": "Z34.83", "description": "Encounter for supervision of normal pregnancy, third trimester",                  "chapter": "XXI. Factors influencing health status",                   "is_billable": True},
    {"code": "Z79.4",  "description": "Long term (current) use of insulin",                                              "chapter": "XXI. Factors influencing health status",                   "is_billable": True},
]

CPT_CODES = [
    {"code": "99203", "description": "Office or other outpatient visit, new patient, low MDM (30-44 min)",        "category": "E/M",       "default_charge": 215.00},
    {"code": "99213", "description": "Office or other outpatient visit, established patient, low MDM (20-29 min)","category": "E/M",       "default_charge": 145.00},
    {"code": "99214", "description": "Office or other outpatient visit, established patient, moderate MDM (30-39 min)","category": "E/M",   "default_charge": 215.00},
    {"code": "99215", "description": "Office or other outpatient visit, established patient, high MDM (40-54 min)","category": "E/M",      "default_charge": 305.00},
    {"code": "99284", "description": "Emergency department visit, moderate severity",                              "category": "E/M",      "default_charge": 565.00},
    {"code": "99285", "description": "Emergency department visit, high severity",                                  "category": "E/M",      "default_charge": 835.00},
    {"code": "99221", "description": "Initial hospital inpatient or observation care, low complexity (40 min)",    "category": "E/M",      "default_charge": 312.00},
    {"code": "99232", "description": "Subsequent hospital inpatient or observation care, moderate complexity",     "category": "E/M",      "default_charge": 175.00},
    {"code": "99291", "description": "Critical care, first 30-74 minutes",                                         "category": "E/M",      "default_charge": 685.00},
    {"code": "36415", "description": "Routine venipuncture for collection of specimen(s)",                         "category": "medicine", "default_charge": 25.00},
    {"code": "80053", "description": "Comprehensive metabolic panel",                                              "category": "pathology","default_charge": 65.00},
    {"code": "85025", "description": "Complete CBC with automated differential WBC",                               "category": "pathology","default_charge": 45.00},
    {"code": "83036", "description": "Hemoglobin A1c (glycosylated)",                                              "category": "pathology","default_charge": 55.00},
    {"code": "80061", "description": "Lipid panel",                                                                "category": "pathology","default_charge": 75.00},
    {"code": "84443", "description": "Thyroid stimulating hormone (TSH)",                                          "category": "pathology","default_charge": 60.00},
    {"code": "84484", "description": "Troponin, quantitative",                                                     "category": "pathology","default_charge": 95.00},
    {"code": "81003", "description": "Urinalysis, automated, without microscopy",                                  "category": "pathology","default_charge": 18.00},
    {"code": "93000", "description": "Electrocardiogram, routine ECG with at least 12 leads",                      "category": "medicine", "default_charge": 95.00},
    {"code": "93306", "description": "Echocardiography, transthoracic, complete with spectral and color Doppler",  "category": "medicine", "default_charge": 425.00},
    {"code": "71046", "description": "Radiologic examination, chest, 2 views",                                     "category": "radiology","default_charge": 145.00},
    {"code": "71250", "description": "Computed tomography, thorax, without contrast material",                     "category": "radiology","default_charge": 525.00},
    {"code": "74176", "description": "Computed tomography, abdomen and pelvis, without contrast material",         "category": "radiology","default_charge": 685.00},
    {"code": "76700", "description": "Ultrasound, abdominal, real time with image documentation; complete",       "category": "radiology","default_charge": 285.00},
    {"code": "27130", "description": "Total hip arthroplasty, with or without autograft or allograft",             "category": "surgery",  "default_charge": 8500.00},
    {"code": "45378", "description": "Colonoscopy, flexible; diagnostic, including specimen collection",           "category": "surgery",  "default_charge": 1250.00},
    {"code": "96365", "description": "Intravenous infusion, prophylactic, initial, up to 1 hour",                  "category": "medicine", "default_charge": 165.00},
    {"code": "90471", "description": "Immunization administration, percutaneous, one vaccine",                     "category": "medicine", "default_charge": 28.00},
    {"code": "94640", "description": "Pressurized or nonpressurized inhalation treatment (nebulizer)",             "category": "medicine", "default_charge": 65.00},
]

LOINC_CODES = [
    {"code": "718-7",   "display": "Hemoglobin [Mass/volume] in Blood",                  "class": "HEM/BC", "system": "Bld",      "units_default": "g/dL"},
    {"code": "4544-3",  "display": "Hematocrit [Volume Fraction] of Blood by Automated count","class": "HEM/BC","system": "Bld",   "units_default": "%"},
    {"code": "6690-2",  "display": "Leukocytes [#/volume] in Blood by Automated count",  "class": "HEM/BC", "system": "Bld",      "units_default": "10*3/uL"},
    {"code": "777-3",   "display": "Platelets [#/volume] in Blood by Automated count",   "class": "HEM/BC", "system": "Bld",      "units_default": "10*3/uL"},
    {"code": "57021-8", "display": "CBC W Auto Differential panel - Blood",              "class": "PANEL.HEM","system": "Bld",    "units_default": ""},
    {"code": "24323-8", "display": "Comprehensive metabolic 2000 panel - Serum or Plasma","class": "PANEL.CHEM","system": "Ser/Plas","units_default": ""},
    {"code": "2345-7",  "display": "Glucose [Mass/volume] in Serum or Plasma",           "class": "CHEM",   "system": "Ser/Plas", "units_default": "mg/dL"},
    {"code": "2160-0",  "display": "Creatinine [Mass/volume] in Serum or Plasma",        "class": "CHEM",   "system": "Ser/Plas", "units_default": "mg/dL"},
    {"code": "3094-0",  "display": "Urea nitrogen [Mass/volume] in Serum or Plasma",     "class": "CHEM",   "system": "Ser/Plas", "units_default": "mg/dL"},
    {"code": "2951-2",  "display": "Sodium [Moles/volume] in Serum or Plasma",           "class": "CHEM",   "system": "Ser/Plas", "units_default": "mmol/L"},
    {"code": "2823-3",  "display": "Potassium [Moles/volume] in Serum or Plasma",        "class": "CHEM",   "system": "Ser/Plas", "units_default": "mmol/L"},
    {"code": "2075-0",  "display": "Chloride [Moles/volume] in Serum or Plasma",         "class": "CHEM",   "system": "Ser/Plas", "units_default": "mmol/L"},
    {"code": "2028-9",  "display": "Carbon dioxide, total [Moles/volume] in Serum or Plasma","class": "CHEM","system": "Ser/Plas","units_default": "mmol/L"},
    {"code": "1742-6",  "display": "Alanine aminotransferase [Enzymatic activity/volume] in Serum or Plasma","class": "CHEM","system": "Ser/Plas","units_default": "U/L"},
    {"code": "1920-8",  "display": "Aspartate aminotransferase [Enzymatic activity/volume] in Serum or Plasma","class": "CHEM","system": "Ser/Plas","units_default": "U/L"},
    {"code": "4548-4",  "display": "Hemoglobin A1c/Hemoglobin.total in Blood",           "class": "CHEM",   "system": "Bld",      "units_default": "%"},
    {"code": "2089-1",  "display": "LDL cholesterol [Mass/volume] in Serum or Plasma",   "class": "CHEM",   "system": "Ser/Plas", "units_default": "mg/dL"},
    {"code": "2085-9",  "display": "HDL cholesterol [Mass/volume] in Serum or Plasma",   "class": "CHEM",   "system": "Ser/Plas", "units_default": "mg/dL"},
    {"code": "2571-8",  "display": "Triglyceride [Mass/volume] in Serum or Plasma",      "class": "CHEM",   "system": "Ser/Plas", "units_default": "mg/dL"},
    {"code": "2093-3",  "display": "Cholesterol [Mass/volume] in Serum or Plasma",       "class": "CHEM",   "system": "Ser/Plas", "units_default": "mg/dL"},
    {"code": "14979-9", "display": "INR in Platelet poor plasma by Coagulation assay",   "class": "COAG",   "system": "PPP",      "units_default": "{INR}"},
    {"code": "10839-9", "display": "Troponin I.cardiac [Mass/volume] in Serum or Plasma","class": "CHEM",   "system": "Ser/Plas", "units_default": "ng/mL"},
    {"code": "3173-2",  "display": "aPTT in Platelet poor plasma by Coagulation assay",  "class": "COAG",   "system": "PPP",      "units_default": "s"},
    {"code": "33914-3", "display": "Glomerular filtration rate/1.73 sq M.predicted",     "class": "CHEM",   "system": "Ser/Plas", "units_default": "mL/min/{1.73_m2}"},
    {"code": "8480-6",  "display": "Systolic blood pressure",                            "class": "BP",     "system": "Arterial", "units_default": "mm[Hg]"},
    {"code": "8462-4",  "display": "Diastolic blood pressure",                           "class": "BP",     "system": "Arterial", "units_default": "mm[Hg]"},
    {"code": "8867-4",  "display": "Heart rate",                                         "class": "HRTRATE","system": "Heart",    "units_default": "{beats}/min"},
    {"code": "2708-6",  "display": "Oxygen saturation in Arterial blood",                "class": "BLDGAS", "system": "BldA",     "units_default": "%"},
    {"code": "20150-9", "display": "FEV1/FVC predicted [Pure number]",                   "class": "PULM",   "system": "Resp",     "units_default": "L"},
    {"code": "33452-4", "display": "Peak expiratory flow",                               "class": "PULM",   "system": "Resp",     "units_default": "L/min"},
    {"code": "11631-9", "display": "Fetal Heart rate Doppler",                           "class": "OB.US",  "system": "Fetus",    "units_default": "{beats}/min"},
    {"code": "11879-4", "display": "Fundal height Tape measure",                         "class": "OB.US",  "system": "^Patient", "units_default": "cm"},
    {"code": "44261-6", "display": "Patient Health Questionnaire-9 (PHQ-9) total score", "class": "SURVEY", "system": "^Patient", "units_default": "{score}"},
    {"code": "70274-6", "display": "Generalized Anxiety Disorder 7 item (GAD-7) total score","class":"SURVEY","system":"^Patient","units_default": "{score}"},
    {"code": "72172-0", "display": "Mini-Mental State Exam (MMSE) total score",          "class": "SURVEY", "system": "^Patient", "units_default": "{score}"},
]

SNOMED_CODES = [
    {"code": "44054006",  "display": "Type 2 diabetes mellitus (disorder)",                "semantic_tag": "disorder"},
    {"code": "38341003",  "display": "Hypertensive disorder, systemic arterial (disorder)","semantic_tag": "disorder"},
    {"code": "195967001", "display": "Asthma (disorder)",                                   "semantic_tag": "disorder"},
    {"code": "13645005",  "display": "Chronic obstructive lung disease (disorder)",        "semantic_tag": "disorder"},
    {"code": "13644009",  "display": "Hypercholesterolemia (disorder)",                    "semantic_tag": "disorder"},
    {"code": "84114007",  "display": "Heart failure (disorder)",                           "semantic_tag": "disorder"},
    {"code": "26929004",  "display": "Alzheimer's disease (disorder)",                     "semantic_tag": "disorder"},
    {"code": "279039007", "display": "Low back pain (finding)",                            "semantic_tag": "finding"},
    {"code": "29857009",  "display": "Chest pain (finding)",                               "semantic_tag": "finding"},
    {"code": "267036007", "display": "Dyspnea (finding)",                                  "semantic_tag": "finding"},
    {"code": "271737000", "display": "Anemia (disorder)",                                  "semantic_tag": "disorder"},
    {"code": "709044004", "display": "Chronic kidney disease (disorder)",                  "semantic_tag": "disorder"},
    {"code": "73211009",  "display": "Diabetes mellitus (disorder)",                       "semantic_tag": "disorder"},
    {"code": "235595009", "display": "Gastroesophageal reflux disease (disorder)",         "semantic_tag": "disorder"},
    {"code": "91936005",  "display": "Allergy to penicillin (finding)",                    "semantic_tag": "finding"},
]

RXNORM_CONCEPTS = [
    {"rxcui": "860975",  "name": "Metformin hydrochloride 1000 MG Oral Tablet",                   "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "314076",  "name": "Lisinopril 10 MG Oral Tablet",                                  "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "314077",  "name": "Lisinopril 20 MG Oral Tablet",                                  "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "617314",  "name": "Atorvastatin calcium 40 MG Oral Tablet",                        "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "617318",  "name": "Atorvastatin calcium 80 MG Oral Tablet",                        "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "308136",  "name": "Albuterol Sulfate 0.083 % Inhalation Solution",                 "tty": "SCD", "route": "inhalation","is_controlled": False,"dea_schedule": None},
    {"rxcui": "745679",  "name": "Albuterol 90 MCG/ACTUAT Metered Dose Inhaler",                  "tty": "SCD", "route": "inhalation","is_controlled": False,"dea_schedule": None},
    {"rxcui": "856987",  "name": "Tiotropium bromide 18 MCG Inhalation Powder",                   "tty": "SCD", "route": "inhalation","is_controlled": False,"dea_schedule": None},
    {"rxcui": "313988",  "name": "Omeprazole 20 MG Oral Capsule",                                 "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "197604",  "name": "Levothyroxine sodium 50 MCG Oral Tablet",                       "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "855332",  "name": "Aspirin 81 MG Oral Tablet",                                     "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "309362",  "name": "Clopidogrel 75 MG Oral Tablet",                                 "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "866412",  "name": "Metoprolol tartrate 50 MG Oral Tablet",                         "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "310429",  "name": "Furosemide 40 MG Oral Tablet",                                  "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "313850",  "name": "Hydrochlorothiazide 25 MG Oral Tablet",                         "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "308191",  "name": "Amoxicillin 500 MG Oral Capsule",                               "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "313782",  "name": "Acetaminophen 325 MG Oral Tablet",                              "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "310965",  "name": "Ibuprofen 600 MG Oral Tablet",                                  "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "892494",  "name": "Morphine sulfate 4 MG/ML Injectable Solution",                  "tty": "SCD", "route": "iv",       "is_controlled": True,  "dea_schedule": "CII"},
    {"rxcui": "1014676", "name": "Oxycodone hydrochloride 5 MG Oral Tablet",                      "tty": "SCD", "route": "oral",     "is_controlled": True,  "dea_schedule": "CII"},
    {"rxcui": "966177",  "name": "Insulin glargine 100 UNT/ML Injectable Solution",               "tty": "SCD", "route": "subq",     "is_controlled": False, "dea_schedule": None},
    {"rxcui": "847232",  "name": "Sertraline 50 MG Oral Tablet",                                  "tty": "SCD", "route": "oral",     "is_controlled": False, "dea_schedule": None},
]

# ---------------------------------------------------------------------------
# 2. ORGANIZATION — locations, specialties, departments, units, rooms, beds
# ---------------------------------------------------------------------------

LOC_HOSPITAL = uid("location", "springfield-regional-medical-center")
LOC_CLINIC   = uid("location", "oakwood-family-medicine-clinic")
LOC_IMAGING  = uid("location", "valley-imaging-center")

LOCATIONS = [
    {"id": LOC_HOSPITAL, "name": "Springfield Regional Medical Center", "facility_type": "hospital", "address_line1": "1450 Memorial Drive", "city": "Springfield", "state": "CA", "postal_code": "93210", "phone": "+1-559-555-2000", "npi": "1417938562"},
    {"id": LOC_CLINIC,   "name": "Oakwood Family Medicine Clinic",     "facility_type": "clinic",   "address_line1": "382 Oakwood Avenue, Suite 200", "city": "Springfield", "state": "CA", "postal_code": "93211", "phone": "+1-559-555-3100", "npi": "1730247105"},
    {"id": LOC_IMAGING,  "name": "Valley Imaging Center",              "facility_type": "imaging",  "address_line1": "27 Industrial Parkway",        "city": "Springfield", "state": "CA", "postal_code": "93212", "phone": "+1-559-555-4400", "npi": "1932148209"},
]

SPECIALTIES_DATA = [
    ("family-medicine",  "207Q00000X", "Family Medicine",       "primary_care"),
    ("internal-medicine","207R00000X", "Internal Medicine",     "primary_care"),
    ("pediatrics",       "208000000X", "Pediatrics",            "primary_care"),
    ("cardiology",       "207RC0000X", "Cardiology",            "specialty"),
    ("endocrinology",    "207RE0101X", "Endocrinology",         "specialty"),
    ("pulmonology",      "207RP1001X", "Pulmonology",           "specialty"),
    ("nephrology",       "207RN0300X", "Nephrology",            "specialty"),
    ("ob-gyn",           "207V00000X", "Obstetrics & Gynecology","specialty"),
    ("orthopedics",      "207X00000X", "Orthopedic Surgery",    "surgical"),
    ("emergency",        "207P00000X", "Emergency Medicine",    "emergency"),
    ("hospitalist",      "208M00000X", "Hospital Medicine",     "hospitalist"),
    ("radiology",        "2085R0202X", "Diagnostic Radiology",  "diagnostic"),
    ("psychiatry",       "2084P0800X", "Psychiatry",            "specialty"),
    ("neurology",        "2084N0400X", "Neurology",             "specialty"),
    ("gastroenterology", "207RG0100X", "Gastroenterology",      "specialty"),
]

SPECIALTIES = [{"id": uid("specialty", code), "code": code, "name": name, "category": category}
               for code, _npi_taxonomy, name, category in SPECIALTIES_DATA]

SHIFTS = [
    {"id": uid("shift", "day"),   "name": "Day Shift (07-19)",   "start_time": "07:00:00", "end_time": "19:00:00", "crosses_midnight": False},
    {"id": uid("shift", "night"), "name": "Night Shift (19-07)", "start_time": "19:00:00", "end_time": "07:00:00", "crosses_midnight": True},
    {"id": uid("shift", "evening"),"name": "Evening Shift (15-23)", "start_time": "15:00:00", "end_time": "23:00:00", "crosses_midnight": False},
    {"id": uid("shift", "office"),"name": "Office Hours (08-17)","start_time": "08:00:00", "end_time": "17:00:00", "crosses_midnight": False},
]

APPT_TYPES_DATA = [
    ("NEW-PT",      "New Patient Visit",            45, False),
    ("FU-EST",      "Established Patient Follow-up",20, False),
    ("ANNUAL-PE",   "Annual Physical Exam",         45, False),
    ("URGENT",      "Urgent Care Same-Day",         20, False),
    ("CHRONIC",     "Chronic Disease Management",   30, False),
    ("PROC-OFFICE", "In-Office Procedure",          30, False),
    ("CARDIO-FU",   "Cardiology Follow-up",         30, True),
    ("PRENATAL",    "Prenatal Visit",               20, False),
    ("PRE-OP",      "Pre-Operative Clearance",      45, True),
    ("TELEHEALTH",  "Telehealth Visit",             20, False),
]

APPOINTMENT_TYPES = [{"id": uid("appointment_type", code), "code": code, "name": name,
                      "default_duration_minutes": dur, "requires_referral": req}
                     for code, name, dur, req in APPT_TYPES_DATA]

PAYER_AETNA    = uid("payer", "aetna")
PAYER_BCBS     = uid("payer", "bcbs-california")
PAYER_MEDICARE = uid("payer", "medicare-cms")
PAYER_MEDICAL  = uid("payer", "medi-cal")
PAYER_KAISER   = uid("payer", "kaiser-permanente")
PAYER_UHC      = uid("payer", "united-healthcare")
PAYER_SELFPAY  = uid("payer", "self-pay")

PAYERS = [
    {"id": PAYER_AETNA,    "name": "Aetna Health Inc.",                "payer_type": "commercial", "payer_id_external": "60054", "address_line1": "151 Farmington Avenue", "city": "Hartford", "state": "CT", "postal_code": "06156", "phone": "+1-800-872-3862", "claims_phone": "+1-888-632-3862", "claims_fax": "+1-860-262-7705", "electronic_claims_supported": True},
    {"id": PAYER_BCBS,     "name": "Blue Shield of California",        "payer_type": "commercial", "payer_id_external": "BS001", "address_line1": "601 12th Street",       "city": "Oakland",  "state": "CA", "postal_code": "94607", "phone": "+1-800-393-6130", "claims_phone": "+1-800-541-6652", "claims_fax": "+1-510-607-2390", "electronic_claims_supported": True},
    {"id": PAYER_MEDICARE, "name": "Medicare (Noridian JE)",           "payer_type": "medicare",   "payer_id_external": "01182", "address_line1": "PO Box 6781",           "city": "Fargo",    "state": "ND", "postal_code": "58108", "phone": "+1-855-609-9960", "claims_phone": "+1-855-609-9960", "claims_fax": None, "electronic_claims_supported": True},
    {"id": PAYER_MEDICAL,  "name": "Medi-Cal (California Medicaid)",    "payer_type": "medicaid",   "payer_id_external": "MEDCAL", "address_line1": "PO Box 13029",          "city": "Sacramento","state": "CA", "postal_code": "95813", "phone": "+1-800-541-5555", "claims_phone": "+1-800-541-5555", "claims_fax": None, "electronic_claims_supported": True},
    {"id": PAYER_KAISER,   "name": "Kaiser Permanente Northern CA",    "payer_type": "commercial", "payer_id_external": "94135", "address_line1": "One Kaiser Plaza",       "city": "Oakland",  "state": "CA", "postal_code": "94612", "phone": "+1-800-464-4000", "claims_phone": "+1-800-390-3510", "claims_fax": "+1-866-590-3960", "electronic_claims_supported": True},
    {"id": PAYER_UHC,      "name": "UnitedHealthcare of California",   "payer_type": "commercial", "payer_id_external": "87726", "address_line1": "5701 Katella Avenue",    "city": "Cypress",  "state": "CA", "postal_code": "90630", "phone": "+1-866-633-2446", "claims_phone": "+1-877-842-3210", "claims_fax": "+1-248-733-6086", "electronic_claims_supported": True},
    {"id": PAYER_SELFPAY,  "name": "Self-Pay",                          "payer_type": "self_pay",   "payer_id_external": None,    "address_line1": None, "city": None, "state": None, "postal_code": None, "phone": None, "claims_phone": None, "claims_fax": None, "electronic_claims_supported": False},
]

PHARM_INHOUSE = uid("pharmacy", "springfield-regional-inpatient")
PHARM_CVS     = uid("pharmacy", "cvs-springfield-1234")
PHARM_WALGRN  = uid("pharmacy", "walgreens-springfield-5678")
PHARM_MAIL    = uid("pharmacy", "express-scripts-mail")

PHARMACIES = [
    {"id": PHARM_INHOUSE, "name": "Springfield Regional Inpatient Pharmacy", "ncpdp_id": "0512987", "npi": "1265789031", "phone": "+1-559-555-2078", "fax": "+1-559-555-2079", "address_line1": "1450 Memorial Drive, Lower Level", "city": "Springfield", "state": "CA", "postal_code": "93210", "is_mail_order": False},
    {"id": PHARM_CVS,     "name": "CVS Pharmacy #1234",                       "ncpdp_id": "0512345", "npi": "1457289103", "phone": "+1-559-555-1234", "fax": "+1-559-555-1235", "address_line1": "240 Main Street",                  "city": "Springfield", "state": "CA", "postal_code": "93211", "is_mail_order": False},
    {"id": PHARM_WALGRN,  "name": "Walgreens #5678",                          "ncpdp_id": "0556789", "npi": "1839271046", "phone": "+1-559-555-5678", "fax": "+1-559-555-5679", "address_line1": "8810 Tulare Avenue",              "city": "Springfield", "state": "CA", "postal_code": "93212", "is_mail_order": False},
    {"id": PHARM_MAIL,    "name": "Express Scripts Mail Service",             "ncpdp_id": "0888100", "npi": "1689407251", "phone": "+1-800-282-2881", "fax": "+1-800-282-2882", "address_line1": "PO Box 66577",                    "city": "St. Louis",   "state": "MO", "postal_code": "63166", "is_mail_order": True},
]

DEPT_FAMILY   = uid("department", "family-medicine-clinic")
DEPT_IM       = uid("department", "internal-medicine-clinic")
DEPT_CARDIO   = uid("department", "cardiology-clinic")
DEPT_OBGYN    = uid("department", "ob-gyn-clinic")
DEPT_ED       = uid("department", "emergency-department")
DEPT_MEDSURG  = uid("department", "medical-surgical")
DEPT_ICU      = uid("department", "critical-care")
DEPT_ORTHO    = uid("department", "orthopedic-surgery")
DEPT_RAD      = uid("department", "radiology")
DEPT_LAB      = uid("department", "clinical-laboratory")

DEPARTMENTS = [
    {"id": DEPT_FAMILY,  "location_id": LOC_CLINIC,   "name": "Family Medicine",          "code": "FAM",   "department_type": "outpatient",  "phone": "+1-559-555-3120"},
    {"id": DEPT_IM,      "location_id": LOC_CLINIC,   "name": "Internal Medicine",        "code": "IM",    "department_type": "outpatient",  "phone": "+1-559-555-3130"},
    {"id": DEPT_CARDIO,  "location_id": LOC_HOSPITAL, "name": "Cardiology",               "code": "CARD",  "department_type": "outpatient",  "phone": "+1-559-555-2300"},
    {"id": DEPT_OBGYN,   "location_id": LOC_HOSPITAL, "name": "Obstetrics & Gynecology",  "code": "OBGYN", "department_type": "outpatient",  "phone": "+1-559-555-2400"},
    {"id": DEPT_ED,      "location_id": LOC_HOSPITAL, "name": "Emergency Department",     "code": "ED",    "department_type": "emergency",   "phone": "+1-559-555-2911"},
    {"id": DEPT_MEDSURG, "location_id": LOC_HOSPITAL, "name": "Medical-Surgical",         "code": "MS",    "department_type": "inpatient",   "phone": "+1-559-555-2500"},
    {"id": DEPT_ICU,     "location_id": LOC_HOSPITAL, "name": "Critical Care (ICU)",      "code": "ICU",   "department_type": "inpatient",   "phone": "+1-559-555-2600"},
    {"id": DEPT_ORTHO,   "location_id": LOC_HOSPITAL, "name": "Orthopedic Surgery",       "code": "ORTHO", "department_type": "surgical",    "phone": "+1-559-555-2700"},
    {"id": DEPT_RAD,     "location_id": LOC_HOSPITAL, "name": "Diagnostic Radiology",     "code": "RAD",   "department_type": "diagnostic",  "phone": "+1-559-555-2800"},
    {"id": DEPT_LAB,     "location_id": LOC_HOSPITAL, "name": "Clinical Laboratory",      "code": "LAB",   "department_type": "diagnostic",  "phone": "+1-559-555-2850"},
]

UNIT_ED_ACUTE  = uid("unit", "ed-acute")
UNIT_ED_FT     = uid("unit", "ed-fast-track")
UNIT_MS_3W     = uid("unit", "medsurg-3-west")
UNIT_MS_4E     = uid("unit", "medsurg-4-east")
UNIT_MICU      = uid("unit", "micu")
UNIT_SICU      = uid("unit", "sicu")
UNIT_ORTHO_OR  = uid("unit", "ortho-or")

UNITS = [
    {"id": UNIT_ED_ACUTE, "department_id": DEPT_ED,      "name": "ED Acute",          "unit_type": "ed",        "bed_capacity": 24},
    {"id": UNIT_ED_FT,    "department_id": DEPT_ED,      "name": "ED Fast Track",     "unit_type": "ed",        "bed_capacity": 8},
    {"id": UNIT_MS_3W,    "department_id": DEPT_MEDSURG, "name": "3 West Med-Surg",   "unit_type": "med_surg",  "bed_capacity": 30},
    {"id": UNIT_MS_4E,    "department_id": DEPT_MEDSURG, "name": "4 East Med-Surg",   "unit_type": "med_surg",  "bed_capacity": 30},
    {"id": UNIT_MICU,     "department_id": DEPT_ICU,     "name": "Medical ICU",       "unit_type": "icu",       "bed_capacity": 12},
    {"id": UNIT_SICU,     "department_id": DEPT_ICU,     "name": "Surgical ICU",      "unit_type": "icu",       "bed_capacity": 10},
    {"id": UNIT_ORTHO_OR, "department_id": DEPT_ORTHO,   "name": "Orthopedic OR Suite","unit_type": "or",       "bed_capacity": 4},
]

ROOMS = []
BEDS = []
def _add_room_with_beds(unit_id: str, room_no: str, room_type: str, bed_labels: list[str], monitored: bool = False) -> None:
    room_id = uid("room", f"{unit_id[:8]}-{room_no}")
    ROOMS.append({"id": room_id, "unit_id": unit_id, "room_number": room_no, "room_type": room_type})
    for label in bed_labels:
        BEDS.append({"id": uid("bed", f"{room_id[:8]}-{label}"), "room_id": room_id, "bed_label": label, "status": "available", "is_monitored": monitored})

for r in range(1, 9):
    _add_room_with_beds(UNIT_ED_ACUTE, f"E-{r:02d}", "exam", ["A"], monitored=True)
for r in range(1, 5):
    _add_room_with_beds(UNIT_ED_FT, f"FT-{r:02d}", "exam", ["A"], monitored=False)
for r in range(301, 311):
    _add_room_with_beds(UNIT_MS_3W, str(r), "patient", ["A", "B"], monitored=False)
for r in range(401, 411):
    _add_room_with_beds(UNIT_MS_4E, str(r), "patient", ["A", "B"], monitored=False)
for r in range(1, 7):
    _add_room_with_beds(UNIT_MICU, f"ICU-{r}", "icu", ["A"], monitored=True)
for r in range(1, 4):
    _add_room_with_beds(UNIT_ORTHO_OR, f"OR-{r}", "or", ["A"], monitored=True)

# Providers ---------------------------------------------------------------
PROV_MORGAN    = uid("provider", "morgan-blackwell-md")     # family medicine PCP
PROV_RAJESH    = uid("provider", "rajesh-iyer-md")          # internal medicine PCP
PROV_HARPER    = uid("provider", "harper-quinones-md")      # endocrinology
PROV_LISA      = uid("provider", "lisa-chen-md")            # cardiology
PROV_TYRELL    = uid("provider", "tyrell-okafor-md")        # cardiology
PROV_AMANDA    = uid("provider", "amanda-foster-md")        # pulmonology
PROV_KOFI      = uid("provider", "kofi-asante-md")          # nephrology
PROV_SARA      = uid("provider", "sara-meyers-md")          # ob-gyn
PROV_DANIEL    = uid("provider", "daniel-park-md")          # orthopedics
PROV_HASAN     = uid("provider", "hasan-ozdemir-md")        # emergency
PROV_JOHANNA   = uid("provider", "johanna-weiss-md")        # emergency
PROV_PRIYA     = uid("provider", "priya-shah-md")           # hospitalist
PROV_NATE      = uid("provider", "nate-rivera-md")          # hospitalist
PROV_KATHRYN   = uid("provider", "kathryn-mcleod-md")       # radiology
PROV_RYAN_RN   = uid("provider", "ryan-fitzgerald-rn")      # RN ED
PROV_MAYA_RN   = uid("provider", "maya-thompson-rn")        # RN MICU
PROV_BRIAN_PA  = uid("provider", "brian-okafor-pa")         # ED PA-C

PROVIDERS = [
    {"id": PROV_MORGAN,   "npi": "1437201568", "dea_number": "BB1234563", "first_name": "Morgan",   "last_name": "Blackwell",  "credentials": "MD",   "provider_type": "physician", "email": "morgan.blackwell@srmc.org",   "phone": "+1-559-555-3121", "hire_date": "2014-08-15", "termination_date": None, "is_active": True, "primary_department_id": DEPT_FAMILY},
    {"id": PROV_RAJESH,   "npi": "1295830174", "dea_number": "BI8217405", "first_name": "Rajesh",   "last_name": "Iyer",       "credentials": "MD",   "provider_type": "physician", "email": "rajesh.iyer@srmc.org",        "phone": "+1-559-555-3131", "hire_date": "2011-07-01", "termination_date": None, "is_active": True, "primary_department_id": DEPT_IM},
    {"id": PROV_HARPER,   "npi": "1356104728", "dea_number": "BQ4501827", "first_name": "Harper",   "last_name": "Quinones",   "credentials": "MD",   "provider_type": "physician", "email": "harper.quinones@srmc.org",    "phone": "+1-559-555-3144", "hire_date": "2018-09-04", "termination_date": None, "is_active": True, "primary_department_id": DEPT_IM},
    {"id": PROV_LISA,     "npi": "1841730219", "dea_number": "BC7204918", "first_name": "Lisa",     "last_name": "Chen",       "credentials": "MD",   "provider_type": "physician", "email": "lisa.chen@srmc.org",          "phone": "+1-559-555-2301", "hire_date": "2009-06-15", "termination_date": None, "is_active": True, "primary_department_id": DEPT_CARDIO},
    {"id": PROV_TYRELL,   "npi": "1620385471", "dea_number": "BO6172839", "first_name": "Tyrell",   "last_name": "Okafor",     "credentials": "MD",   "provider_type": "physician", "email": "tyrell.okafor@srmc.org",      "phone": "+1-559-555-2302", "hire_date": "2016-08-22", "termination_date": None, "is_active": True, "primary_department_id": DEPT_CARDIO},
    {"id": PROV_AMANDA,   "npi": "1734206518", "dea_number": "BF2937104", "first_name": "Amanda",   "last_name": "Foster",     "credentials": "MD",   "provider_type": "physician", "email": "amanda.foster@srmc.org",      "phone": "+1-559-555-2350", "hire_date": "2013-09-10", "termination_date": None, "is_active": True, "primary_department_id": DEPT_MEDSURG},
    {"id": PROV_KOFI,     "npi": "1820471639", "dea_number": "BA8104728", "first_name": "Kofi",     "last_name": "Asante",     "credentials": "MD",   "provider_type": "physician", "email": "kofi.asante@srmc.org",        "phone": "+1-559-555-2380", "hire_date": "2015-01-12", "termination_date": None, "is_active": True, "primary_department_id": DEPT_MEDSURG},
    {"id": PROV_SARA,     "npi": "1571928034", "dea_number": "BM5028371", "first_name": "Sara",     "last_name": "Meyers",     "credentials": "MD",   "provider_type": "physician", "email": "sara.meyers@srmc.org",        "phone": "+1-559-555-2401", "hire_date": "2010-04-20", "termination_date": None, "is_active": True, "primary_department_id": DEPT_OBGYN},
    {"id": PROV_DANIEL,   "npi": "1290385714", "dea_number": "BP7140392", "first_name": "Daniel",   "last_name": "Park",       "credentials": "MD",   "provider_type": "physician", "email": "daniel.park@srmc.org",        "phone": "+1-559-555-2701", "hire_date": "2012-09-01", "termination_date": None, "is_active": True, "primary_department_id": DEPT_ORTHO},
    {"id": PROV_HASAN,    "npi": "1408372956", "dea_number": "BO3917284", "first_name": "Hasan",    "last_name": "Ozdemir",    "credentials": "MD",   "provider_type": "physician", "email": "hasan.ozdemir@srmc.org",      "phone": "+1-559-555-2902", "hire_date": "2017-11-15", "termination_date": None, "is_active": True, "primary_department_id": DEPT_ED},
    {"id": PROV_JOHANNA,  "npi": "1683019472", "dea_number": "BW6201738", "first_name": "Johanna",  "last_name": "Weiss",      "credentials": "MD",   "provider_type": "physician", "email": "johanna.weiss@srmc.org",      "phone": "+1-559-555-2903", "hire_date": "2015-08-04", "termination_date": None, "is_active": True, "primary_department_id": DEPT_ED},
    {"id": PROV_PRIYA,    "npi": "1972834510", "dea_number": "BS4180925", "first_name": "Priya",    "last_name": "Shah",       "credentials": "MD",   "provider_type": "physician", "email": "priya.shah@srmc.org",         "phone": "+1-559-555-2501", "hire_date": "2014-01-20", "termination_date": None, "is_active": True, "primary_department_id": DEPT_MEDSURG},
    {"id": PROV_NATE,     "npi": "1108423957", "dea_number": "BR9028354", "first_name": "Nate",     "last_name": "Rivera",     "credentials": "MD",   "provider_type": "physician", "email": "nate.rivera@srmc.org",        "phone": "+1-559-555-2502", "hire_date": "2018-07-15", "termination_date": None, "is_active": True, "primary_department_id": DEPT_MEDSURG},
    {"id": PROV_KATHRYN,  "npi": "1739182056", "dea_number": None,        "first_name": "Kathryn",  "last_name": "McLeod",     "credentials": "MD",   "provider_type": "physician", "email": "kathryn.mcleod@srmc.org",     "phone": "+1-559-555-2801", "hire_date": "2008-06-01", "termination_date": None, "is_active": True, "primary_department_id": DEPT_RAD},
    {"id": PROV_RYAN_RN,  "npi": "1850294736", "dea_number": None,        "first_name": "Ryan",     "last_name": "Fitzgerald", "credentials": "RN, BSN","provider_type": "nurse",   "email": "ryan.fitzgerald@srmc.org",    "phone": "+1-559-555-2925", "hire_date": "2019-03-10", "termination_date": None, "is_active": True, "primary_department_id": DEPT_ED},
    {"id": PROV_MAYA_RN,  "npi": "1463829071", "dea_number": None,        "first_name": "Maya",     "last_name": "Thompson",   "credentials": "RN, BSN","provider_type": "nurse",   "email": "maya.thompson@srmc.org",      "phone": "+1-559-555-2625", "hire_date": "2017-05-22", "termination_date": None, "is_active": True, "primary_department_id": DEPT_ICU},
    {"id": PROV_BRIAN_PA, "npi": "1572389104", "dea_number": "BO5301294", "first_name": "Brian",    "last_name": "O'Konnor",   "credentials": "PA-C", "provider_type": "pa",        "email": "brian.okonnor@srmc.org",      "phone": "+1-559-555-2917", "hire_date": "2020-06-15", "termination_date": None, "is_active": True, "primary_department_id": DEPT_ED},
]

PROVIDER_SPECIALTIES = [
    {"provider_id": PROV_MORGAN,  "specialty_id": uid("specialty", "family-medicine"),  "is_primary": True, "board_certified": True, "certification_date": "2015-11-08"},
    {"provider_id": PROV_RAJESH,  "specialty_id": uid("specialty", "internal-medicine"),"is_primary": True, "board_certified": True, "certification_date": "2012-10-15"},
    {"provider_id": PROV_HARPER,  "specialty_id": uid("specialty", "endocrinology"),    "is_primary": True, "board_certified": True, "certification_date": "2019-09-20"},
    {"provider_id": PROV_HARPER,  "specialty_id": uid("specialty", "internal-medicine"),"is_primary": False,"board_certified": True, "certification_date": "2016-10-12"},
    {"provider_id": PROV_LISA,    "specialty_id": uid("specialty", "cardiology"),       "is_primary": True, "board_certified": True, "certification_date": "2011-11-04"},
    {"provider_id": PROV_TYRELL,  "specialty_id": uid("specialty", "cardiology"),       "is_primary": True, "board_certified": True, "certification_date": "2018-11-09"},
    {"provider_id": PROV_AMANDA,  "specialty_id": uid("specialty", "pulmonology"),      "is_primary": True, "board_certified": True, "certification_date": "2015-10-08"},
    {"provider_id": PROV_KOFI,    "specialty_id": uid("specialty", "nephrology"),       "is_primary": True, "board_certified": True, "certification_date": "2017-09-22"},
    {"provider_id": PROV_SARA,    "specialty_id": uid("specialty", "ob-gyn"),           "is_primary": True, "board_certified": True, "certification_date": "2012-11-16"},
    {"provider_id": PROV_DANIEL,  "specialty_id": uid("specialty", "orthopedics"),      "is_primary": True, "board_certified": True, "certification_date": "2014-07-15"},
    {"provider_id": PROV_HASAN,   "specialty_id": uid("specialty", "emergency"),        "is_primary": True, "board_certified": True, "certification_date": "2019-10-04"},
    {"provider_id": PROV_JOHANNA, "specialty_id": uid("specialty", "emergency"),        "is_primary": True, "board_certified": True, "certification_date": "2017-10-06"},
    {"provider_id": PROV_PRIYA,   "specialty_id": uid("specialty", "hospitalist"),      "is_primary": True, "board_certified": True, "certification_date": "2016-10-14"},
    {"provider_id": PROV_NATE,    "specialty_id": uid("specialty", "hospitalist"),      "is_primary": True, "board_certified": True, "certification_date": "2020-10-09"},
    {"provider_id": PROV_KATHRYN, "specialty_id": uid("specialty", "radiology"),        "is_primary": True, "board_certified": True, "certification_date": "2010-09-24"},
]

PROVIDER_LICENSES = [
    {"id": uid("license", f"{p}-CA-MD"), "provider_id": p, "license_type": "medical", "license_number": f"A{i:06d}", "issuing_state": "CA", "issue_date": "2014-08-01", "expiration_date": "2026-08-01", "status": "active"}
    for i, p in enumerate([PROV_MORGAN, PROV_RAJESH, PROV_HARPER, PROV_LISA, PROV_TYRELL, PROV_AMANDA, PROV_KOFI, PROV_SARA, PROV_DANIEL, PROV_HASAN, PROV_JOHANNA, PROV_PRIYA, PROV_NATE, PROV_KATHRYN, PROV_BRIAN_PA], start=84512)
]
PROVIDER_LICENSES += [
    {"id": uid("license", f"{PROV_RYAN_RN}-CA-RN"), "provider_id": PROV_RYAN_RN, "license_type": "nursing", "license_number": "RN748391", "issuing_state": "CA", "issue_date": "2019-02-15", "expiration_date": "2027-02-28", "status": "active"},
    {"id": uid("license", f"{PROV_MAYA_RN}-CA-RN"), "provider_id": PROV_MAYA_RN, "license_type": "nursing", "license_number": "RN623157", "issuing_state": "CA", "issue_date": "2017-04-30", "expiration_date": "2025-04-30", "status": "active"},
    {"id": uid("license", f"{PROV_BRIAN_PA}-CA-PA"),"provider_id": PROV_BRIAN_PA,"license_type": "pa",      "license_number": "PA019384","issuing_state": "CA", "issue_date": "2020-05-15", "expiration_date": "2026-05-15", "status": "active"},
]

# ---------------------------------------------------------------------------
# 3. PATIENTS — 10 patient panel with distinct clinical scenarios
# ---------------------------------------------------------------------------

P_MARGARET = uid("patient", "margaret-johnson")
P_ROBERT   = uid("patient", "robert-chen")
P_LAKSHMI  = uid("patient", "lakshmi-patel")
P_JAMES    = uid("patient", "james-oconnor")
P_AISHA    = uid("patient", "aisha-rodriguez")
P_MARCUS   = uid("patient", "marcus-williams")
P_DOROTHY  = uid("patient", "dorothy-kim")
P_CARLOS   = uid("patient", "carlos-mendoza")
P_EMILY    = uid("patient", "emily-watson")
P_DAVID    = uid("patient", "david-nakamura")

PATIENTS = [
    {"id": P_MARGARET, "mrn": "SRMC-00010001", "first_name": "Margaret", "middle_name": "Louise",   "last_name": "Johnson",  "prefix": "Mrs.","suffix": None, "date_of_birth": "1956-04-12", "sex_at_birth": "female","gender_identity": "female","pronouns": "she/her", "race": "black",  "ethnicity": "non_hispanic","preferred_language": "en", "marital_status": "widowed", "religion": "baptist",      "ssn_last4": "7831", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_MORGAN, "vip_status": None,    "notes": "Long-standing T2DM with progressive CKD. Lives alone, daughter visits weekly."},
    {"id": P_ROBERT,   "mrn": "SRMC-00010002", "first_name": "Robert",   "middle_name": "Wei",      "last_name": "Chen",     "prefix": "Mr.", "suffix": None, "date_of_birth": "1970-09-23", "sex_at_birth": "male",  "gender_identity": "male",  "pronouns": "he/him",  "race": "asian",  "ethnicity": "non_hispanic","preferred_language": "en", "marital_status": "married", "religion": None,           "ssn_last4": "4502", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_RAJESH, "vip_status": None,    "notes": "S/p NSTEMI 2024-03 with PCI to LAD; on DAPT through 2025-03."},
    {"id": P_LAKSHMI,  "mrn": "SRMC-00010003", "first_name": "Lakshmi",  "middle_name": "Devi",     "last_name": "Patel",    "prefix": "Dr.", "suffix": "PhD","date_of_birth": "1982-11-04", "sex_at_birth": "female","gender_identity": "female","pronouns": "she/her", "race": "asian",  "ethnicity": "non_hispanic","preferred_language": "en", "marital_status": "married", "religion": "hindu",        "ssn_last4": "8214", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_MORGAN, "vip_status": None,    "notes": "Moderate persistent asthma, well-controlled on ICS. GERD on PPI."},
    {"id": P_JAMES,    "mrn": "SRMC-00010004", "first_name": "James",    "middle_name": "Patrick",  "last_name": "O'Connor", "prefix": "Mr.", "suffix": "Sr.","date_of_birth": "1951-02-17", "sex_at_birth": "male",  "gender_identity": "male",  "pronouns": "he/him",  "race": "white",  "ethnicity": "non_hispanic","preferred_language": "en", "marital_status": "married", "religion": "catholic",     "ssn_last4": "1029", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_RAJESH, "vip_status": None,    "notes": "Severe COPD (GOLD D) + chronic diastolic CHF. 40 pack-year smoker, quit 2019."},
    {"id": P_AISHA,    "mrn": "SRMC-00010005", "first_name": "Aisha",    "middle_name": "Marie",    "last_name": "Rodriguez","prefix": "Ms.", "suffix": None, "date_of_birth": "1995-07-08", "sex_at_birth": "female","gender_identity": "female","pronouns": "she/her", "race": "white",  "ethnicity": "hispanic",    "preferred_language": "es", "marital_status": "partnered","religion": "catholic",    "ssn_last4": "6394", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_SARA,   "vip_status": None,    "notes": "G2P1 32 weeks gestation, EDD 2025-04-18. No GDM, no preeclampsia."},
    {"id": P_MARCUS,   "mrn": "SRMC-00010006", "first_name": "Marcus",   "middle_name": "Anthony",  "last_name": "Williams", "prefix": "Mr.", "suffix": None, "date_of_birth": "1989-12-19", "sex_at_birth": "male",  "gender_identity": "male",  "pronouns": "he/him",  "race": "black",  "ethnicity": "non_hispanic","preferred_language": "en", "marital_status": "single",  "religion": None,           "ssn_last4": "2058", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_MORGAN, "vip_status": None,    "notes": "Warehouse worker with mechanical chronic LBP. No red flags on prior MRI."},
    {"id": P_DOROTHY,  "mrn": "SRMC-00010007", "first_name": "Dorothy",  "middle_name": "Hye-Jin",  "last_name": "Kim",      "prefix": "Mrs.","suffix": None, "date_of_birth": "1943-06-30", "sex_at_birth": "female","gender_identity": "female","pronouns": "she/her", "race": "asian",  "ethnicity": "non_hispanic","preferred_language": "ko", "marital_status": "widowed", "religion": "methodist",    "ssn_last4": "9847", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_RAJESH, "vip_status": None,    "notes": "Moderate Alzheimer's (MoCA 14/30). Son is HCPOA. Korean interpreter required."},
    {"id": P_CARLOS,   "mrn": "SRMC-00010008", "first_name": "Carlos",   "middle_name": "Eduardo",  "last_name": "Mendoza",  "prefix": "Mr.", "suffix": None, "date_of_birth": "1977-03-26", "sex_at_birth": "male",  "gender_identity": "male",  "pronouns": "he/him",  "race": "white",  "ethnicity": "hispanic",    "preferred_language": "es", "marital_status": "married", "religion": "catholic",     "ssn_last4": "3741", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_RAJESH, "vip_status": None,    "notes": "Anxiety with panic-like chest pain; no known cardiac disease."},
    {"id": P_EMILY,    "mrn": "SRMC-00010009", "first_name": "Emily",    "middle_name": "Grace",    "last_name": "Watson",   "prefix": "Ms.", "suffix": None, "date_of_birth": "2000-08-15", "sex_at_birth": "female","gender_identity": "female","pronouns": "she/her", "race": "white",  "ethnicity": "non_hispanic","preferred_language": "en", "marital_status": "single",  "religion": None,           "ssn_last4": "5728", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_MORGAN, "vip_status": None,    "notes": "College student. Asthma since age 9, several ED visits during wildfire season."},
    {"id": P_DAVID,    "mrn": "SRMC-00010010", "first_name": "David",    "middle_name": "Hiroshi",  "last_name": "Nakamura", "prefix": "Mr.", "suffix": None, "date_of_birth": "1962-10-11", "sex_at_birth": "male",  "gender_identity": "male",  "pronouns": "he/him",  "race": "asian",  "ethnicity": "non_hispanic","preferred_language": "en", "marital_status": "married", "religion": "buddhist",     "ssn_last4": "6105", "is_deceased": False, "deceased_at": None, "primary_provider_id": PROV_RAJESH, "vip_status": None,    "notes": "New T2DM dx Q4 2024 (A1c 9.2). Counseled on lifestyle + initiated metformin."},
]

PATIENT_IDENTIFIERS = [
    {"id": uid("patid", "margaret-ssn"),    "patient_id": P_MARGARET, "identifier_type": "ssn",      "identifier_value": "574-92-7831", "issuing_authority": "SSA",                  "valid_from": "1956-04-12", "valid_to": None},
    {"id": uid("patid", "margaret-mcare"),  "patient_id": P_MARGARET, "identifier_type": "medicare", "identifier_value": "1A23-BC4-DE56","issuing_authority": "CMS",                  "valid_from": "2021-05-01", "valid_to": None},
    {"id": uid("patid", "margaret-dl"),     "patient_id": P_MARGARET, "identifier_type": "drivers_license","identifier_value": "B4218704", "issuing_authority": "CA DMV",         "valid_from": "2020-04-12", "valid_to": "2026-04-12"},
    {"id": uid("patid", "robert-ssn"),      "patient_id": P_ROBERT,   "identifier_type": "ssn",      "identifier_value": "638-22-4502", "issuing_authority": "SSA",                  "valid_from": "1970-09-23", "valid_to": None},
    {"id": uid("patid", "lakshmi-ssn"),     "patient_id": P_LAKSHMI,  "identifier_type": "ssn",      "identifier_value": "604-91-8214", "issuing_authority": "SSA",                  "valid_from": "1982-11-04", "valid_to": None},
    {"id": uid("patid", "james-ssn"),       "patient_id": P_JAMES,    "identifier_type": "ssn",      "identifier_value": "513-44-1029", "issuing_authority": "SSA",                  "valid_from": "1951-02-17", "valid_to": None},
    {"id": uid("patid", "james-mcare"),     "patient_id": P_JAMES,    "identifier_type": "medicare", "identifier_value": "2J34-KL5-MN67","issuing_authority": "CMS",                 "valid_from": "2016-03-01", "valid_to": None},
    {"id": uid("patid", "aisha-ssn"),       "patient_id": P_AISHA,    "identifier_type": "ssn",      "identifier_value": "601-83-6394", "issuing_authority": "SSA",                  "valid_from": "1995-07-08", "valid_to": None},
    {"id": uid("patid", "marcus-ssn"),      "patient_id": P_MARCUS,   "identifier_type": "ssn",      "identifier_value": "578-21-2058", "issuing_authority": "SSA",                  "valid_from": "1989-12-19", "valid_to": None},
    {"id": uid("patid", "dorothy-ssn"),     "patient_id": P_DOROTHY,  "identifier_type": "ssn",      "identifier_value": "549-37-9847", "issuing_authority": "SSA",                  "valid_from": "1965-08-12", "valid_to": None},
    {"id": uid("patid", "dorothy-mcare"),   "patient_id": P_DOROTHY,  "identifier_type": "medicare", "identifier_value": "5K12-PQ7-RS89","issuing_authority": "CMS",                 "valid_from": "2008-07-01", "valid_to": None},
    {"id": uid("patid", "carlos-ssn"),      "patient_id": P_CARLOS,   "identifier_type": "ssn",      "identifier_value": "612-50-3741", "issuing_authority": "SSA",                  "valid_from": "1977-03-26", "valid_to": None},
    {"id": uid("patid", "emily-ssn"),       "patient_id": P_EMILY,    "identifier_type": "ssn",      "identifier_value": "569-04-5728", "issuing_authority": "SSA",                  "valid_from": "2000-08-15", "valid_to": None},
    {"id": uid("patid", "david-ssn"),       "patient_id": P_DAVID,    "identifier_type": "ssn",      "identifier_value": "582-19-6105", "issuing_authority": "SSA",                  "valid_from": "1962-10-11", "valid_to": None},
]

PATIENT_ADDRESSES = [
    {"id": uid("addr", "margaret-home"), "patient_id": P_MARGARET, "address_use": "home", "line1": "812 Maplewood Lane",    "line2": None,        "city": "Springfield", "state": "CA", "postal_code": "93210", "country": "US", "is_primary": True, "valid_from": "1998-06-01", "valid_to": None},
    {"id": uid("addr", "robert-home"),   "patient_id": P_ROBERT,   "address_use": "home", "line1": "144 Crescent Ridge Rd", "line2": None,        "city": "Springfield", "state": "CA", "postal_code": "93211", "country": "US", "is_primary": True, "valid_from": "2008-09-15", "valid_to": None},
    {"id": uid("addr", "lakshmi-home"),  "patient_id": P_LAKSHMI,  "address_use": "home", "line1": "2901 University Drive", "line2": "Apt 14C",   "city": "Springfield", "state": "CA", "postal_code": "93212", "country": "US", "is_primary": True, "valid_from": "2019-08-01", "valid_to": None},
    {"id": uid("addr", "james-home"),    "patient_id": P_JAMES,    "address_use": "home", "line1": "67 Oakhurst Drive",     "line2": None,        "city": "Springfield", "state": "CA", "postal_code": "93210", "country": "US", "is_primary": True, "valid_from": "1985-04-20", "valid_to": None},
    {"id": uid("addr", "aisha-home"),    "patient_id": P_AISHA,    "address_use": "home", "line1": "3389 Olive Avenue",     "line2": "Unit B",    "city": "Springfield", "state": "CA", "postal_code": "93211", "country": "US", "is_primary": True, "valid_from": "2022-11-01", "valid_to": None},
    {"id": uid("addr", "marcus-home"),   "patient_id": P_MARCUS,   "address_use": "home", "line1": "5018 Industrial Way",   "line2": "Apt 312",   "city": "Springfield", "state": "CA", "postal_code": "93212", "country": "US", "is_primary": True, "valid_from": "2021-02-10", "valid_to": None},
    {"id": uid("addr", "dorothy-home"),  "patient_id": P_DOROTHY,  "address_use": "home", "line1": "7240 Cedar Brook Circle","line2": None,       "city": "Fresno",      "state": "CA", "postal_code": "93720", "country": "US", "is_primary": True, "valid_from": "1995-07-12", "valid_to": None},
    {"id": uid("addr", "carlos-home"),   "patient_id": P_CARLOS,   "address_use": "home", "line1": "1108 Vineyard Trail",   "line2": None,        "city": "Springfield", "state": "CA", "postal_code": "93210", "country": "US", "is_primary": True, "valid_from": "2014-06-01", "valid_to": None},
    {"id": uid("addr", "emily-home"),    "patient_id": P_EMILY,    "address_use": "home", "line1": "240 Campus View Drive", "line2": "Room 418B", "city": "Springfield", "state": "CA", "postal_code": "93211", "country": "US", "is_primary": True, "valid_from": "2023-08-22", "valid_to": None},
    {"id": uid("addr", "emily-billing"), "patient_id": P_EMILY,    "address_use": "billing","line1": "984 Hazelwood Court","line2": None,        "city": "Modesto",     "state": "CA", "postal_code": "95350", "country": "US", "is_primary": False,"valid_from": "2000-08-15", "valid_to": None},
    {"id": uid("addr", "david-home"),    "patient_id": P_DAVID,    "address_use": "home", "line1": "5519 Sunset Boulevard", "line2": None,        "city": "Springfield", "state": "CA", "postal_code": "93212", "country": "US", "is_primary": True, "valid_from": "2002-11-15", "valid_to": None},
]

PATIENT_CONTACTS = [
    {"id": uid("contact", "margaret-phone"),  "patient_id": P_MARGARET, "contact_system": "phone", "contact_value": "+1-559-555-1182", "contact_use": "home",   "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "margaret-email"),  "patient_id": P_MARGARET, "contact_system": "email", "contact_value": "mljohnson1956@gmail.com", "contact_use": "personal","is_primary": False, "consent_to_contact": True},
    {"id": uid("contact", "robert-phone"),    "patient_id": P_ROBERT,   "contact_system": "phone", "contact_value": "+1-559-555-2073", "contact_use": "mobile", "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "robert-work"),     "patient_id": P_ROBERT,   "contact_system": "phone", "contact_value": "+1-559-555-2099", "contact_use": "work",   "is_primary": False, "consent_to_contact": False},
    {"id": uid("contact", "robert-email"),    "patient_id": P_ROBERT,   "contact_system": "email", "contact_value": "rchen.eng@outlook.com","contact_use": "personal","is_primary": False, "consent_to_contact": True},
    {"id": uid("contact", "lakshmi-phone"),   "patient_id": P_LAKSHMI,  "contact_system": "phone", "contact_value": "+1-559-555-3158", "contact_use": "mobile", "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "lakshmi-email"),   "patient_id": P_LAKSHMI,  "contact_system": "email", "contact_value": "l.patel@springfieldu.edu","contact_use": "work","is_primary": False, "consent_to_contact": True},
    {"id": uid("contact", "james-phone"),     "patient_id": P_JAMES,    "contact_system": "phone", "contact_value": "+1-559-555-4421", "contact_use": "home",   "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "aisha-phone"),     "patient_id": P_AISHA,    "contact_system": "phone", "contact_value": "+1-559-555-7710", "contact_use": "mobile", "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "aisha-email"),     "patient_id": P_AISHA,    "contact_system": "email", "contact_value": "arodriguez95@protonmail.com","contact_use": "personal","is_primary": False, "consent_to_contact": True},
    {"id": uid("contact", "marcus-phone"),    "patient_id": P_MARCUS,   "contact_system": "phone", "contact_value": "+1-559-555-8809", "contact_use": "mobile", "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "dorothy-phone"),   "patient_id": P_DOROTHY,  "contact_system": "phone", "contact_value": "+1-559-555-9920", "contact_use": "home",   "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "carlos-phone"),    "patient_id": P_CARLOS,   "contact_system": "phone", "contact_value": "+1-559-555-1147", "contact_use": "mobile", "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "emily-phone"),     "patient_id": P_EMILY,    "contact_system": "phone", "contact_value": "+1-209-555-3328", "contact_use": "mobile", "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "emily-email"),     "patient_id": P_EMILY,    "contact_system": "email", "contact_value": "ewatson@students.springfieldu.edu","contact_use": "work","is_primary": False, "consent_to_contact": True},
    {"id": uid("contact", "david-phone"),     "patient_id": P_DAVID,    "contact_system": "phone", "contact_value": "+1-559-555-2278", "contact_use": "mobile", "is_primary": True,  "consent_to_contact": True},
    {"id": uid("contact", "david-email"),     "patient_id": P_DAVID,    "contact_system": "email", "contact_value": "d.nakamura62@yahoo.com","contact_use": "personal","is_primary": False, "consent_to_contact": True},
]

EMERGENCY_CONTACTS = [
    {"id": uid("emerg", "margaret-daughter"),"patient_id": P_MARGARET, "name": "Patricia Johnson-Reeves","relationship": "child",      "phone": "+1-559-555-1183", "email": "pjreeves@gmail.com",  "address_line1": "812 Maplewood Lane",    "city": "Springfield", "state": "CA", "postal_code": "93210", "priority_rank": 1, "has_medical_poa": True},
    {"id": uid("emerg", "robert-wife"),     "patient_id": P_ROBERT,   "name": "Jennifer Chen",          "relationship": "spouse",     "phone": "+1-559-555-2074", "email": "jchen.dpt@yahoo.com",  "address_line1": "144 Crescent Ridge Rd", "city": "Springfield", "state": "CA", "postal_code": "93211", "priority_rank": 1, "has_medical_poa": True},
    {"id": uid("emerg", "lakshmi-spouse"),  "patient_id": P_LAKSHMI,  "name": "Vikram Patel",           "relationship": "spouse",     "phone": "+1-559-555-3159", "email": "v.patel.md@gmail.com", "address_line1": "2901 University Drive, Apt 14C","city": "Springfield","state": "CA","postal_code":"93212","priority_rank":1,"has_medical_poa": True},
    {"id": uid("emerg", "james-wife"),      "patient_id": P_JAMES,    "name": "Margaret Mary O'Connor", "relationship": "spouse",     "phone": "+1-559-555-4422", "email": None,                   "address_line1": "67 Oakhurst Drive",     "city": "Springfield", "state": "CA", "postal_code": "93210", "priority_rank": 1, "has_medical_poa": True},
    {"id": uid("emerg", "james-son"),       "patient_id": P_JAMES,    "name": "Daniel O'Connor Jr.",    "relationship": "child",      "phone": "+1-559-555-4488", "email": "doc.jr@gmail.com",     "address_line1": "1840 Cypress Hills Lane","city": "Springfield","state":"CA","postal_code":"93212","priority_rank":2,"has_medical_poa": False},
    {"id": uid("emerg", "aisha-partner"),   "patient_id": P_AISHA,    "name": "Miguel Rodriguez",       "relationship": "partner",    "phone": "+1-559-555-7711", "email": "m.rodriguez@gmail.com","address_line1": "3389 Olive Avenue, Unit B","city": "Springfield","state":"CA","postal_code":"93211","priority_rank": 1, "has_medical_poa": False},
    {"id": uid("emerg", "marcus-mom"),      "patient_id": P_MARCUS,   "name": "Yolanda Williams",       "relationship": "parent",     "phone": "+1-559-555-8810", "email": None,                   "address_line1": "904 Sycamore St",       "city": "Springfield","state": "CA","postal_code": "93210","priority_rank": 1, "has_medical_poa": False},
    {"id": uid("emerg", "dorothy-son"),     "patient_id": P_DOROTHY,  "name": "Min-Jun Kim",            "relationship": "child",      "phone": "+1-559-555-9921", "email": "mjkim.cpa@gmail.com",  "address_line1": "7240 Cedar Brook Circle","city": "Fresno",   "state": "CA","postal_code":"93720","priority_rank": 1, "has_medical_poa": True},
    {"id": uid("emerg", "dorothy-dau"),     "patient_id": P_DOROTHY,  "name": "Soo-Yeon Kim-Bennett",   "relationship": "child",      "phone": "+1-415-555-3047", "email": "sybennett@gmail.com",  "address_line1": "188 Vermont St",        "city": "San Francisco","state":"CA","postal_code":"94103","priority_rank": 2, "has_medical_poa": False},
    {"id": uid("emerg", "carlos-wife"),     "patient_id": P_CARLOS,   "name": "Sofia Mendoza",          "relationship": "spouse",     "phone": "+1-559-555-1148", "email": "smendoza.rn@srmc.org", "address_line1": "1108 Vineyard Trail",   "city": "Springfield", "state":"CA","postal_code":"93210","priority_rank": 1, "has_medical_poa": True},
    {"id": uid("emerg", "emily-mom"),       "patient_id": P_EMILY,    "name": "Linda Watson",           "relationship": "parent",     "phone": "+1-209-555-3329", "email": "lwatson.modesto@gmail.com","address_line1": "984 Hazelwood Court","city": "Modesto","state":"CA","postal_code": "95350","priority_rank": 1, "has_medical_poa": True},
    {"id": uid("emerg", "david-wife"),      "patient_id": P_DAVID,    "name": "Yuki Nakamura",          "relationship": "spouse",     "phone": "+1-559-555-2279", "email": "yuki.nakamura@gmail.com","address_line1": "5519 Sunset Boulevard","city": "Springfield","state":"CA","postal_code":"93212","priority_rank": 1, "has_medical_poa": True},
]

PATIENT_CONSENTS = [
    {"id": uid("consent", f"{p}-hipaa"),    "patient_id": p, "consent_type": "hipaa",         "granted": True, "granted_at": ts(2020, 1, 15, 9, 30, 0), "expires_at": None, "witness_name": "Front Desk Staff", "document_ref": "doc://consent/hipaa_2020v3.pdf"}
    for p in [P_MARGARET, P_ROBERT, P_LAKSHMI, P_JAMES, P_AISHA, P_MARCUS, P_DOROTHY, P_CARLOS, P_EMILY, P_DAVID]
] + [
    {"id": uid("consent", "margaret-portal"),"patient_id": P_MARGARET, "consent_type": "patient_portal","granted": True,"granted_at": ts(2022, 3, 4, 14, 15, 0), "expires_at": None, "witness_name": None, "document_ref": None},
    {"id": uid("consent", "robert-portal"),  "patient_id": P_ROBERT,   "consent_type": "patient_portal","granted": True,"granted_at": ts(2021, 6, 22, 11, 4, 0), "expires_at": None, "witness_name": None, "document_ref": None},
    {"id": uid("consent", "aisha-research"), "patient_id": P_AISHA,    "consent_type": "research",      "granted": True,"granted_at": ts(2024, 11, 20, 10, 12, 0),"expires_at": ts(2026, 11, 20, 10, 12, 0), "witness_name": "Dr. Sara Meyers","document_ref": "doc://research/CIRB-2024-318_consent.pdf"},
    {"id": uid("consent", "dorothy-poaref"), "patient_id": P_DOROTHY,  "consent_type": "advance_directive","granted":True,"granted_at": ts(2022, 5, 9, 13, 30, 0), "expires_at": None, "witness_name": "Min-Jun Kim", "document_ref": "doc://advdir/dorothy_kim_ad_2022.pdf"},
]

# ---------------------------------------------------------------------------
# 4. COVERAGE — insurance plans, patient coverages, authorizations
# ---------------------------------------------------------------------------

PLAN_MCARE_AB    = uid("plan", "medicare-part-ab")
PLAN_MEDICAL_HMO = uid("plan", "medi-cal-managed-hmo")
PLAN_AETNA_PPO   = uid("plan", "aetna-choice-pos-ii")
PLAN_BCBS_PPO    = uid("plan", "bcbs-trio-ace-ppo")
PLAN_KAISER_HMO  = uid("plan", "kaiser-signature-hmo")
PLAN_UHC_EPO     = uid("plan", "uhc-navigate-balanced-epo")
PLAN_AETNA_HDHP  = uid("plan", "aetna-high-deductible-hdhp")
PLAN_SELFPAY     = uid("plan", "self-pay-default")

INSURANCE_PLANS = [
    {"id": PLAN_MCARE_AB,    "payer_id": PAYER_MEDICARE, "plan_name": "Medicare Part A & B (Original)",      "plan_type": "medicare",  "metal_tier": None,    "group_number": None,         "effective_date": "2003-01-01", "termination_date": None, "requires_referrals": False, "requires_prior_auth_for_imaging": False},
    {"id": PLAN_MEDICAL_HMO, "payer_id": PAYER_MEDICAL,  "plan_name": "Medi-Cal Managed Care (CalViva)",     "plan_type": "medicaid",  "metal_tier": None,    "group_number": "MEDCAL-MC",   "effective_date": "2014-01-01", "termination_date": None, "requires_referrals": True,  "requires_prior_auth_for_imaging": True},
    {"id": PLAN_AETNA_PPO,   "payer_id": PAYER_AETNA,    "plan_name": "Aetna Choice POS II - Open Access",   "plan_type": "ppo",       "metal_tier": "gold",  "group_number": "SRMC-AET-G1", "effective_date": "2024-01-01", "termination_date": "2025-12-31","requires_referrals": False,"requires_prior_auth_for_imaging": True},
    {"id": PLAN_BCBS_PPO,    "payer_id": PAYER_BCBS,     "plan_name": "Blue Shield Trio ACE PPO Gold",        "plan_type": "ppo",       "metal_tier": "gold",  "group_number": "CSEU-BS-7782","effective_date": "2024-01-01","termination_date": "2025-12-31","requires_referrals": False,"requires_prior_auth_for_imaging": True},
    {"id": PLAN_KAISER_HMO,  "payer_id": PAYER_KAISER,   "plan_name": "Kaiser Permanente Signature HMO",      "plan_type": "hmo",       "metal_tier": "silver","group_number": "KP-NORCAL-22","effective_date": "2024-01-01","termination_date": "2025-12-31","requires_referrals": True, "requires_prior_auth_for_imaging": True},
    {"id": PLAN_UHC_EPO,     "payer_id": PAYER_UHC,      "plan_name": "UnitedHealthcare Navigate Balanced EPO","plan_type": "epo",      "metal_tier": "silver","group_number": "UHC-NAV-SI",  "effective_date": "2024-01-01","termination_date": "2025-12-31","requires_referrals": True, "requires_prior_auth_for_imaging": True},
    {"id": PLAN_AETNA_HDHP,  "payer_id": PAYER_AETNA,    "plan_name": "Aetna HealthFund HDHP / HSA",          "plan_type": "hdhp",      "metal_tier": "bronze","group_number": "SRMC-AET-B2", "effective_date": "2024-01-01","termination_date": "2025-12-31","requires_referrals": False,"requires_prior_auth_for_imaging": True},
    {"id": PLAN_SELFPAY,     "payer_id": PAYER_SELFPAY,  "plan_name": "Self-Pay (No Coverage)",               "plan_type": "self_pay",  "metal_tier": None,    "group_number": None,         "effective_date": "2020-01-01", "termination_date": None, "requires_referrals": False, "requires_prior_auth_for_imaging": False},
]

COV_MARGARET_MCARE = uid("cov", "margaret-medicare")
COV_MARGARET_MCAL  = uid("cov", "margaret-medical")
COV_ROBERT_AETNA   = uid("cov", "robert-aetna")
COV_LAKSHMI_BCBS   = uid("cov", "lakshmi-bcbs")
COV_JAMES_MCARE    = uid("cov", "james-medicare")
COV_JAMES_BCBS     = uid("cov", "james-bcbs-supp")
COV_AISHA_MEDICAL  = uid("cov", "aisha-medical")
COV_MARCUS_UHC     = uid("cov", "marcus-uhc")
COV_DOROTHY_MCARE  = uid("cov", "dorothy-medicare")
COV_CARLOS_KAISER  = uid("cov", "carlos-kaiser")
COV_EMILY_AETNA    = uid("cov", "emily-aetna-hdhp")
COV_DAVID_AETNA    = uid("cov", "david-aetna-ppo")

PATIENT_COVERAGES = [
    {"id": COV_MARGARET_MCARE,"patient_id": P_MARGARET, "insurance_plan_id": PLAN_MCARE_AB,    "subscriber_relationship": "self",     "subscriber_name": "Margaret L. Johnson","subscriber_dob": "1956-04-12","member_id": "1A23BC4DE56", "group_number": None,         "coverage_rank": 1, "effective_date": "2021-05-01", "termination_date": None, "copay_pcp": 0,    "copay_specialist": 0,    "copay_er": 0,    "deductible_individual": 240,   "deductible_family": None,  "oop_max_individual": None,    "verified_at": ts(2024,12,1,9,0,0)},
    {"id": COV_MARGARET_MCAL, "patient_id": P_MARGARET, "insurance_plan_id": PLAN_MEDICAL_HMO, "subscriber_relationship": "self",     "subscriber_name": "Margaret L. Johnson","subscriber_dob": "1956-04-12","member_id": "97834561200","group_number": "MEDCAL-MC",   "coverage_rank": 2, "effective_date": "2021-05-01", "termination_date": None, "copay_pcp": 0,    "copay_specialist": 0,    "copay_er": 0,    "deductible_individual": 0,     "deductible_family": None,  "oop_max_individual": 0,       "verified_at": ts(2024,12,1,9,5,0)},
    {"id": COV_ROBERT_AETNA,  "patient_id": P_ROBERT,   "insurance_plan_id": PLAN_AETNA_PPO,   "subscriber_relationship": "self",     "subscriber_name": "Robert W. Chen",     "subscriber_dob": "1970-09-23","member_id": "W283047192", "group_number": "SRMC-AET-G1","coverage_rank": 1, "effective_date": "2024-01-01", "termination_date": "2025-12-31","copay_pcp":25,"copay_specialist":50, "copay_er":250,  "deductible_individual": 1500,  "deductible_family": 3000,  "oop_max_individual": 6500,    "verified_at": ts(2024,12,15,10,30,0)},
    {"id": COV_LAKSHMI_BCBS,  "patient_id": P_LAKSHMI,  "insurance_plan_id": PLAN_BCBS_PPO,    "subscriber_relationship": "self",     "subscriber_name": "Lakshmi D. Patel",   "subscriber_dob": "1982-11-04","member_id": "XSC718305294","group_number": "CSEU-BS-7782","coverage_rank": 1, "effective_date": "2024-01-01","termination_date": "2025-12-31","copay_pcp":30,"copay_specialist":60, "copay_er":300,  "deductible_individual": 1000,  "deductible_family": 2000,  "oop_max_individual": 5000,    "verified_at": ts(2025,1,8,9,15,0)},
    {"id": COV_JAMES_MCARE,   "patient_id": P_JAMES,    "insurance_plan_id": PLAN_MCARE_AB,    "subscriber_relationship": "self",     "subscriber_name": "James P. O'Connor",  "subscriber_dob": "1951-02-17","member_id": "2J34KL5MN67", "group_number": None,         "coverage_rank": 1, "effective_date": "2016-03-01","termination_date": None, "copay_pcp": 0,    "copay_specialist": 0,    "copay_er": 0,    "deductible_individual": 240,   "deductible_family": None,  "oop_max_individual": None,    "verified_at": ts(2024,12,20,14,0,0)},
    {"id": COV_JAMES_BCBS,    "patient_id": P_JAMES,    "insurance_plan_id": PLAN_BCBS_PPO,    "subscriber_relationship": "self",     "subscriber_name": "James P. O'Connor",  "subscriber_dob": "1951-02-17","member_id": "BSCASUPP1029","group_number": "CSEU-BS-MED","coverage_rank": 2, "effective_date": "2016-03-01","termination_date": None, "copay_pcp":15,"copay_specialist":30, "copay_er":150,  "deductible_individual": 200,   "deductible_family": None,  "oop_max_individual": 2000,    "verified_at": ts(2024,12,20,14,5,0)},
    {"id": COV_AISHA_MEDICAL, "patient_id": P_AISHA,    "insurance_plan_id": PLAN_MEDICAL_HMO, "subscriber_relationship": "self",     "subscriber_name": "Aisha M. Rodriguez", "subscriber_dob": "1995-07-08","member_id": "98234581100","group_number": "MEDCAL-MC",   "coverage_rank": 1, "effective_date": "2024-11-01","termination_date": None, "copay_pcp": 0,    "copay_specialist": 0,    "copay_er": 0,    "deductible_individual": 0,     "deductible_family": None,  "oop_max_individual": 0,       "verified_at": ts(2024,11,5,11,0,0)},
    {"id": COV_MARCUS_UHC,    "patient_id": P_MARCUS,   "insurance_plan_id": PLAN_UHC_EPO,     "subscriber_relationship": "self",     "subscriber_name": "Marcus A. Williams", "subscriber_dob": "1989-12-19","member_id": "UHC927183054","group_number": "UHC-NAV-SI","coverage_rank": 1, "effective_date": "2024-01-01","termination_date": "2025-12-31","copay_pcp":35,"copay_specialist":75, "copay_er":400,  "deductible_individual": 2500,  "deductible_family": 5000,  "oop_max_individual": 7900,    "verified_at": ts(2024,10,10,9,0,0)},
    {"id": COV_DOROTHY_MCARE, "patient_id": P_DOROTHY,  "insurance_plan_id": PLAN_MCARE_AB,    "subscriber_relationship": "self",     "subscriber_name": "Dorothy H. Kim",     "subscriber_dob": "1943-06-30","member_id": "5K12PQ7RS89", "group_number": None,         "coverage_rank": 1, "effective_date": "2008-07-01","termination_date": None, "copay_pcp": 0,    "copay_specialist": 0,    "copay_er": 0,    "deductible_individual": 240,   "deductible_family": None,  "oop_max_individual": None,    "verified_at": ts(2024,12,30,13,15,0)},
    {"id": COV_CARLOS_KAISER, "patient_id": P_CARLOS,   "insurance_plan_id": PLAN_KAISER_HMO,  "subscriber_relationship": "self",     "subscriber_name": "Carlos E. Mendoza",  "subscriber_dob": "1977-03-26","member_id": "KP-NOR-7283019","group_number": "KP-NORCAL-22","coverage_rank": 1,"effective_date":"2024-01-01","termination_date":"2025-12-31","copay_pcp":20,"copay_specialist":40,"copay_er":150,"deductible_individual":500,"deductible_family":1000,"oop_max_individual":4500,"verified_at": ts(2024,11,12,10,0,0)},
    {"id": COV_EMILY_AETNA,   "patient_id": P_EMILY,    "insurance_plan_id": PLAN_AETNA_HDHP,  "subscriber_relationship": "child",    "subscriber_name": "Linda Watson",       "subscriber_dob": "1972-04-22","member_id": "W491738205", "group_number": "SRMC-AET-B2","coverage_rank": 1, "effective_date": "2024-01-01","termination_date": "2025-12-31","copay_pcp":0, "copay_specialist":0,  "copay_er":0,    "deductible_individual": 5500,  "deductible_family": 11000, "oop_max_individual": 8050,    "verified_at": ts(2024,9,1,9,0,0)},
    {"id": COV_DAVID_AETNA,   "patient_id": P_DAVID,    "insurance_plan_id": PLAN_AETNA_PPO,   "subscriber_relationship": "self",     "subscriber_name": "David H. Nakamura",  "subscriber_dob": "1962-10-11","member_id": "W618290473", "group_number": "SRMC-AET-G1","coverage_rank": 1, "effective_date": "2024-01-01","termination_date": "2025-12-31","copay_pcp":25,"copay_specialist":50, "copay_er":250,  "deductible_individual": 1500,  "deductible_family": 3000,  "oop_max_individual": 6500,    "verified_at": ts(2025,1,2,11,30,0)},
]

AUTH_DOROTHY_THA  = uid("auth", "dorothy-tha-27130")
AUTH_ROBERT_ECHO  = uid("auth", "robert-echo-93306")
AUTH_AISHA_US     = uid("auth", "aisha-ob-us-76700")
AUTH_JAMES_CTCHEST= uid("auth", "james-ct-71250")
AUTHORIZATIONS = [
    {"id": AUTH_DOROTHY_THA, "patient_id": P_DOROTHY, "coverage_id": COV_DOROTHY_MCARE, "auth_number": "MCARE-PA-204781", "cpt_code": "27130", "requested_units": 1, "approved_units": 1, "status": "approved",  "effective_date": "2024-11-22", "expiration_date": "2025-02-22", "requested_at": ts(2024,11,19,14,30,0), "decided_at": ts(2024,11,21,10,15,0), "denial_reason": None},
    {"id": AUTH_ROBERT_ECHO, "patient_id": P_ROBERT,  "coverage_id": COV_ROBERT_AETNA,  "auth_number": "AET-PA-7193058", "cpt_code": "93306", "requested_units": 1, "approved_units": 1, "status": "approved",  "effective_date": "2024-12-05", "expiration_date": "2025-03-05", "requested_at": ts(2024,12,2,9,12,0),   "decided_at": ts(2024,12,3,13,45,0),  "denial_reason": None},
    {"id": AUTH_AISHA_US,    "patient_id": P_AISHA,   "coverage_id": COV_AISHA_MEDICAL, "auth_number": "MCAL-PA-8302174","cpt_code": "76700", "requested_units": 1, "approved_units": 1, "status": "approved",  "effective_date": "2025-01-10", "expiration_date": "2025-04-10", "requested_at": ts(2025,1,6,10,0,0),    "decided_at": ts(2025,1,8,11,30,0),   "denial_reason": None},
    {"id": AUTH_JAMES_CTCHEST,"patient_id": P_JAMES,  "coverage_id": COV_JAMES_MCARE,   "auth_number": "MCARE-PA-204913","cpt_code": "71250", "requested_units": 1, "approved_units": 0, "status": "denied",   "effective_date": None,          "expiration_date": None,          "requested_at": ts(2025,1,14,8,30,0),   "decided_at": ts(2025,1,15,14,0,0),   "denial_reason": "Medical necessity not established - patient already had CT chest within 12 months without significant clinical change. Recommend CXR first."},
]

# ---------------------------------------------------------------------------
# 5. SCHEDULING — slots, referrals, appointments, status history, reminders, waitlist
# ---------------------------------------------------------------------------

AT_NEW_PT   = uid("appointment_type", "NEW-PT")
AT_FU_EST   = uid("appointment_type", "FU-EST")
AT_ANNUAL   = uid("appointment_type", "ANNUAL-PE")
AT_URGENT   = uid("appointment_type", "URGENT")
AT_CHRONIC  = uid("appointment_type", "CHRONIC")
AT_CARDIO   = uid("appointment_type", "CARDIO-FU")
AT_PRENATAL = uid("appointment_type", "PRENATAL")
AT_PREOP    = uid("appointment_type", "PRE-OP")
AT_TELE     = uid("appointment_type", "TELEHEALTH")

# Generate a small grid of slots for two days in mid-January 2025 for the PCPs
APPOINTMENT_SLOTS = []
_slot_seed = [
    (PROV_MORGAN,  LOC_CLINIC,   "2025-01-13", [(8,0),(8,20),(8,40),(9,0),(9,20),(9,40),(10,0),(10,20),(10,40)]),
    (PROV_MORGAN,  LOC_CLINIC,   "2025-01-14", [(8,0),(8,20),(8,40),(9,0),(9,20),(9,40),(10,0),(10,20),(10,40)]),
    (PROV_RAJESH,  LOC_CLINIC,   "2025-01-13", [(13,0),(13,30),(14,0),(14,30),(15,0),(15,30)]),
    (PROV_HARPER,  LOC_CLINIC,   "2025-01-15", [(9,0),(9,30),(10,0),(10,30),(11,0)]),
    (PROV_LISA,    LOC_HOSPITAL, "2025-01-22", [(8,0),(8,30),(9,0),(9,30),(10,0),(10,30)]),
    (PROV_SARA,    LOC_HOSPITAL, "2025-01-16", [(9,0),(9,20),(9,40),(10,0),(10,20),(10,40)]),
]
for provider, loc, day, times in _slot_seed:
    for h, mi in times:
        sid = uid("slot", f"{provider}-{day}-{h:02d}{mi:02d}")
        start = f"{day}T{h:02d}:{mi:02d}:00Z"
        end_m = mi + 20
        eh = h + (end_m // 60)
        em = end_m % 60
        end = f"{day}T{eh:02d}:{em:02d}:00Z"
        APPOINTMENT_SLOTS.append({"id": sid, "provider_id": provider, "location_id": loc, "slot_start": start, "slot_end": end, "is_available": True, "appointment_type_id": AT_FU_EST})

# Referrals
REF_MARGARET_NEPHRO = uid("ref", "margaret-nephrology")
REF_MARGARET_ENDO   = uid("ref", "margaret-endocrinology")
REF_ROBERT_CARDIO   = uid("ref", "robert-cardio-fu")
REF_DOROTHY_ORTHO   = uid("ref", "dorothy-ortho-hip")
REF_DAVID_ENDO      = uid("ref", "david-endocrinology")
REF_MARCUS_PT       = uid("ref", "marcus-physical-therapy")

REFERRALS = [
    {"id": REF_MARGARET_NEPHRO, "patient_id": P_MARGARET, "referring_provider_id": PROV_MORGAN, "referred_to_provider_id": PROV_KOFI,   "referred_to_specialty_id": uid("specialty", "nephrology"),    "reason": "Stage 3 CKD with worsening eGFR (52→44 over 6 months). Please evaluate.", "priority": "routine","status": "active","created_at_date": "2024-12-08","expires_on": "2025-06-08","authorization_id": None},
    {"id": REF_MARGARET_ENDO,   "patient_id": P_MARGARET, "referring_provider_id": PROV_MORGAN, "referred_to_provider_id": PROV_HARPER, "referred_to_specialty_id": uid("specialty", "endocrinology"), "reason": "T2DM with A1c 7.8% despite metformin/sulfonylurea, considering GLP-1.","priority":"routine","status":"active","created_at_date":"2024-11-15","expires_on":"2025-05-15","authorization_id":None},
    {"id": REF_ROBERT_CARDIO,   "patient_id": P_ROBERT,   "referring_provider_id": PROV_RAJESH, "referred_to_provider_id": PROV_LISA,   "referred_to_specialty_id": uid("specialty", "cardiology"),    "reason": "Post-NSTEMI 9mo f/u. Stable on DAPT. Repeat echo and consider DAPT discontinuation.","priority":"routine","status":"completed","created_at_date":"2024-11-28","expires_on":"2025-05-28","authorization_id":None},
    {"id": REF_DOROTHY_ORTHO,   "patient_id": P_DOROTHY,  "referring_provider_id": PROV_RAJESH, "referred_to_provider_id": PROV_DANIEL, "referred_to_specialty_id": uid("specialty", "orthopedics"),   "reason": "S/p right hip fracture 2024-11-18, ORIF considered vs THA. Surgical evaluation.","priority":"urgent","status":"completed","created_at_date":"2024-11-19","expires_on":"2025-02-19","authorization_id":AUTH_DOROTHY_THA},
    {"id": REF_DAVID_ENDO,      "patient_id": P_DAVID,    "referring_provider_id": PROV_RAJESH, "referred_to_provider_id": PROV_HARPER, "referred_to_specialty_id": uid("specialty", "endocrinology"), "reason": "Newly diagnosed T2DM (A1c 9.2). Initial endocrinology consultation.","priority":"routine","status":"active","created_at_date":"2025-01-09","expires_on":"2025-07-09","authorization_id":None},
    {"id": REF_MARCUS_PT,       "patient_id": P_MARCUS,   "referring_provider_id": PROV_MORGAN, "referred_to_provider_id": None,        "referred_to_specialty_id": None,                                "reason": "Chronic mechanical LBP. 6 visits of PT for core strengthening and McKenzie ext program.","priority":"routine","status":"active","created_at_date":"2025-01-10","expires_on":"2025-04-10","authorization_id":None},
]

# Appointments
APPT_MARGARET_FU  = uid("appt", "margaret-fu-2025-01-13")
APPT_MARGARET_NEPHRO = uid("appt", "margaret-nephrology-2025-02-04")
APPT_ROBERT_CARDIO= uid("appt", "robert-cardio-2024-12-12")
APPT_LAKSHMI_FU   = uid("appt", "lakshmi-asthma-fu-2025-01-13")
APPT_AISHA_OB     = uid("appt", "aisha-prenatal-2025-01-16")
APPT_MARCUS_LBP   = uid("appt", "marcus-lbp-2025-01-08")
APPT_DAVID_NEWPT  = uid("appt", "david-newdm-2024-12-19")
APPT_DOROTHY_PREOP= uid("appt", "dorothy-preop-2024-11-26")
APPT_CARLOS_TELE  = uid("appt", "carlos-tele-anxiety-2025-01-09")
APPT_EMILY_URGENT = uid("appt", "emily-asthma-urgent-2024-10-22")

APPOINTMENTS = [
    {"id": APPT_MARGARET_FU,    "patient_id": P_MARGARET, "provider_id": PROV_MORGAN, "appointment_type_id": AT_CHRONIC,  "location_id": LOC_CLINIC,   "department_id": DEPT_FAMILY, "scheduled_start": ts(2025,1,13,9,0,0),  "scheduled_end": ts(2025,1,13,9,30,0),  "actual_start": ts(2025,1,13,9,5,0),   "actual_end": ts(2025,1,13,9,38,0),  "status": "completed", "reason_for_visit": "Diabetes & hypertension follow-up; medication refills", "referral_id": None,                  "created_by_provider_id": PROV_MORGAN},
    {"id": APPT_MARGARET_NEPHRO,"patient_id": P_MARGARET, "provider_id": PROV_KOFI,   "appointment_type_id": AT_NEW_PT,   "location_id": LOC_HOSPITAL, "department_id": DEPT_MEDSURG,"scheduled_start": ts(2025,2,4,10,0,0),  "scheduled_end": ts(2025,2,4,10,45,0),  "actual_start": None,                   "actual_end": None,                   "status": "scheduled", "reason_for_visit": "New nephrology consultation - CKD stage 3 evaluation",         "referral_id": REF_MARGARET_NEPHRO,   "created_by_provider_id": PROV_MORGAN},
    {"id": APPT_ROBERT_CARDIO,  "patient_id": P_ROBERT,   "provider_id": PROV_LISA,   "appointment_type_id": AT_CARDIO,   "location_id": LOC_HOSPITAL, "department_id": DEPT_CARDIO, "scheduled_start": ts(2024,12,12,8,30,0),"scheduled_end": ts(2024,12,12,9,0,0),  "actual_start": ts(2024,12,12,8,35,0), "actual_end": ts(2024,12,12,9,10,0), "status": "completed", "reason_for_visit": "Post-NSTEMI 9-month follow-up + echocardiogram",                "referral_id": REF_ROBERT_CARDIO,     "created_by_provider_id": PROV_RAJESH},
    {"id": APPT_LAKSHMI_FU,     "patient_id": P_LAKSHMI,  "provider_id": PROV_MORGAN, "appointment_type_id": AT_FU_EST,   "location_id": LOC_CLINIC,   "department_id": DEPT_FAMILY, "scheduled_start": ts(2025,1,13,10,0,0), "scheduled_end": ts(2025,1,13,10,20,0), "actual_start": ts(2025,1,13,10,2,0),  "actual_end": ts(2025,1,13,10,22,0), "status": "completed", "reason_for_visit": "Asthma controller refill, peak flows trending well",            "referral_id": None,                  "created_by_provider_id": PROV_MORGAN},
    {"id": APPT_AISHA_OB,       "patient_id": P_AISHA,    "provider_id": PROV_SARA,   "appointment_type_id": AT_PRENATAL, "location_id": LOC_HOSPITAL, "department_id": DEPT_OBGYN,  "scheduled_start": ts(2025,1,16,9,20,0), "scheduled_end": ts(2025,1,16,9,40,0),  "actual_start": ts(2025,1,16,9,25,0),  "actual_end": ts(2025,1,16,9,52,0),  "status": "completed", "reason_for_visit": "32-week prenatal visit + GBS swab + Tdap",                       "referral_id": None,                  "created_by_provider_id": PROV_SARA},
    {"id": APPT_MARCUS_LBP,     "patient_id": P_MARCUS,   "provider_id": PROV_MORGAN, "appointment_type_id": AT_FU_EST,   "location_id": LOC_CLINIC,   "department_id": DEPT_FAMILY, "scheduled_start": ts(2025,1,8,15,0,0),  "scheduled_end": ts(2025,1,8,15,20,0),  "actual_start": ts(2025,1,8,15,10,0),  "actual_end": ts(2025,1,8,15,32,0),  "status": "completed", "reason_for_visit": "Chronic low back pain flare - 5/10 with bending",                "referral_id": None,                  "created_by_provider_id": PROV_MORGAN},
    {"id": APPT_DAVID_NEWPT,    "patient_id": P_DAVID,    "provider_id": PROV_RAJESH, "appointment_type_id": AT_NEW_PT,   "location_id": LOC_CLINIC,   "department_id": DEPT_IM,     "scheduled_start": ts(2024,12,19,13,0,0),"scheduled_end": ts(2024,12,19,13,45,0),"actual_start": ts(2024,12,19,13,8,0), "actual_end": ts(2024,12,19,13,56,0),"status": "completed", "reason_for_visit": "Polyuria/polydipsia x 6 weeks; new patient evaluation",         "referral_id": None,                  "created_by_provider_id": PROV_RAJESH},
    {"id": APPT_DOROTHY_PREOP,  "patient_id": P_DOROTHY,  "provider_id": PROV_RAJESH, "appointment_type_id": AT_PREOP,    "location_id": LOC_CLINIC,   "department_id": DEPT_IM,     "scheduled_start": ts(2024,11,26,14,0,0),"scheduled_end": ts(2024,11,26,14,45,0),"actual_start": ts(2024,11,26,14,5,0), "actual_end": ts(2024,11,26,14,48,0),"status": "completed", "reason_for_visit": "Pre-op medical clearance for elective R THA",                    "referral_id": REF_DOROTHY_ORTHO,     "created_by_provider_id": PROV_DANIEL},
    {"id": APPT_CARLOS_TELE,    "patient_id": P_CARLOS,   "provider_id": PROV_RAJESH, "appointment_type_id": AT_TELE,     "location_id": LOC_CLINIC,   "department_id": DEPT_IM,     "scheduled_start": ts(2025,1,9,13,30,0), "scheduled_end": ts(2025,1,9,13,50,0),  "actual_start": ts(2025,1,9,13,32,0),  "actual_end": ts(2025,1,9,13,55,0),  "status": "completed", "reason_for_visit": "Telehealth - GAD follow-up after ED visit, SSRI titration",      "referral_id": None,                  "created_by_provider_id": PROV_RAJESH},
    {"id": APPT_EMILY_URGENT,   "patient_id": P_EMILY,    "provider_id": PROV_MORGAN, "appointment_type_id": AT_URGENT,   "location_id": LOC_CLINIC,   "department_id": DEPT_FAMILY, "scheduled_start": ts(2024,10,22,14,40,0),"scheduled_end": ts(2024,10,22,15,0,0), "actual_start": ts(2024,10,22,14,45,0),"actual_end": ts(2024,10,22,15,5,0), "status": "completed", "reason_for_visit": "Increased albuterol use during wildfire smoke",                   "referral_id": None,                  "created_by_provider_id": PROV_MORGAN},
]

APPOINTMENT_STATUS_HISTORY = [
    {"id": uid("apptstat", f"{APPT_MARGARET_FU}-1"), "appointment_id": APPT_MARGARET_FU, "from_status": None,        "to_status": "scheduled","changed_at": ts(2024,12,8,11,30,0), "changed_by_provider_id": PROV_MORGAN, "reason": "Routine 3-month follow-up scheduled by PCP"},
    {"id": uid("apptstat", f"{APPT_MARGARET_FU}-2"), "appointment_id": APPT_MARGARET_FU, "from_status": "scheduled","to_status": "arrived",  "changed_at": ts(2025,1,13,8,50,0),  "changed_by_provider_id": None,        "reason": "Patient checked in at front desk"},
    {"id": uid("apptstat", f"{APPT_MARGARET_FU}-3"), "appointment_id": APPT_MARGARET_FU, "from_status": "arrived",   "to_status": "in_progress","changed_at": ts(2025,1,13,9,5,0),  "changed_by_provider_id": PROV_MORGAN, "reason": "Provider began encounter"},
    {"id": uid("apptstat", f"{APPT_MARGARET_FU}-4"), "appointment_id": APPT_MARGARET_FU, "from_status": "in_progress","to_status": "completed","changed_at": ts(2025,1,13,9,38,0), "changed_by_provider_id": PROV_MORGAN, "reason": "Visit completed and documented"},
    {"id": uid("apptstat", f"{APPT_DAVID_NEWPT}-1"), "appointment_id": APPT_DAVID_NEWPT, "from_status": None,        "to_status": "scheduled","changed_at": ts(2024,12,10,14,15,0),"changed_by_provider_id": PROV_RAJESH, "reason": "Patient called for new appointment"},
    {"id": uid("apptstat", f"{APPT_DAVID_NEWPT}-2"), "appointment_id": APPT_DAVID_NEWPT, "from_status": "scheduled","to_status": "completed","changed_at": ts(2024,12,19,13,56,0),"changed_by_provider_id": PROV_RAJESH, "reason": "Visit completed"},
    {"id": uid("apptstat", f"{APPT_AISHA_OB}-1"),    "appointment_id": APPT_AISHA_OB,    "from_status": None,        "to_status": "scheduled","changed_at": ts(2024,12,16,11,0,0), "changed_by_provider_id": PROV_SARA,   "reason": "Routine prenatal schedule"},
    {"id": uid("apptstat", f"{APPT_AISHA_OB}-2"),    "appointment_id": APPT_AISHA_OB,    "from_status": "scheduled","to_status": "completed","changed_at": ts(2025,1,16,9,52,0),  "changed_by_provider_id": PROV_SARA,   "reason": "Visit completed"},
    {"id": uid("apptstat", f"{APPT_MARGARET_NEPHRO}-1"),"appointment_id":APPT_MARGARET_NEPHRO,"from_status":None,    "to_status":"scheduled", "changed_at": ts(2025,1,8,10,30,0),  "changed_by_provider_id": PROV_KOFI,   "reason": "Referral consultation scheduled"},
]

WAITLIST_ENTRIES = [
    {"id": uid("waitlist", "carlos-earlier"), "patient_id": P_CARLOS, "provider_id": PROV_RAJESH, "appointment_type_id": AT_FU_EST,  "requested_after": "2025-01-15", "requested_before": "2025-01-31", "priority": "routine", "status": "active"},
    {"id": uid("waitlist", "david-endo"),     "patient_id": P_DAVID,  "provider_id": PROV_HARPER, "appointment_type_id": AT_NEW_PT,  "requested_after": "2025-01-20", "requested_before": "2025-02-15", "priority": "routine", "status": "active"},
    {"id": uid("waitlist", "marcus-pcp-fu"),  "patient_id": P_MARCUS, "provider_id": PROV_MORGAN, "appointment_type_id": AT_FU_EST,  "requested_after": "2025-02-01", "requested_before": "2025-02-20", "priority": "low",     "status": "active"},
]

APPOINTMENT_REMINDERS = [
    {"id": uid("rem", f"{APPT_MARGARET_NEPHRO}-sms-7d"),"appointment_id": APPT_MARGARET_NEPHRO,"patient_id": P_MARGARET,"channel": "sms",  "scheduled_send_at": ts(2025,1,28,9,0,0),  "sent_at": ts(2025,1,28,9,0,12),  "delivery_status": "delivered","response": "confirmed"},
    {"id": uid("rem", f"{APPT_MARGARET_NEPHRO}-call-1d"),"appointment_id": APPT_MARGARET_NEPHRO,"patient_id": P_MARGARET,"channel": "phone","scheduled_send_at": ts(2025,2,3,15,0,0),  "sent_at": None,                "delivery_status": "pending",  "response": None},
    {"id": uid("rem", f"{APPT_ROBERT_CARDIO}-sms-3d"),  "appointment_id": APPT_ROBERT_CARDIO,  "patient_id": P_ROBERT,  "channel": "sms",  "scheduled_send_at": ts(2024,12,9,9,0,0),  "sent_at": ts(2024,12,9,9,0,8),   "delivery_status": "delivered","response": "confirmed"},
    {"id": uid("rem", f"{APPT_AISHA_OB}-sms-2d"),       "appointment_id": APPT_AISHA_OB,       "patient_id": P_AISHA,   "channel": "sms",  "scheduled_send_at": ts(2025,1,14,12,0,0), "sent_at": ts(2025,1,14,12,0,5),  "delivery_status": "delivered","response": "confirmed"},
    {"id": uid("rem", f"{APPT_DAVID_NEWPT}-email-7d"),  "appointment_id": APPT_DAVID_NEWPT,    "patient_id": P_DAVID,   "channel": "email","scheduled_send_at": ts(2024,12,12,9,0,0), "sent_at": ts(2024,12,12,9,0,3),  "delivery_status": "delivered","response": "no_response"},
    {"id": uid("rem", f"{APPT_DOROTHY_PREOP}-call-3d"), "appointment_id": APPT_DOROTHY_PREOP,  "patient_id": P_DOROTHY, "channel": "phone","scheduled_send_at": ts(2024,11,23,10,0,0),"sent_at": ts(2024,11,23,10,4,12),"delivery_status": "delivered","response": "confirmed"},
]

# ---------------------------------------------------------------------------
# 6. ENCOUNTERS — outpatient, ED, inpatient stays
# ---------------------------------------------------------------------------

# Outpatient encounters (mirror appointments) + ED + inpatient
ENC_MARGARET_FU      = uid("enc", "margaret-pcp-fu-2025-01-13")
ENC_MARGARET_NEPHRO  = uid("enc", "margaret-nephrology-2025-02-04")  # not yet started (scheduled)
ENC_ROBERT_CARDIO    = uid("enc", "robert-cardio-2024-12-12")
ENC_LAKSHMI_FU       = uid("enc", "lakshmi-asthma-fu-2025-01-13")
ENC_AISHA_OB         = uid("enc", "aisha-prenatal-2025-01-16")
ENC_MARCUS_LBP       = uid("enc", "marcus-lbp-2025-01-08")
ENC_DAVID_NEWPT      = uid("enc", "david-newdm-2024-12-19")
ENC_DOROTHY_PREOP    = uid("enc", "dorothy-preop-2024-11-26")
ENC_CARLOS_TELE      = uid("enc", "carlos-tele-2025-01-09")
ENC_EMILY_URGENT     = uid("enc", "emily-urgent-2024-10-22")

# ED encounters
ENC_ROBERT_ED        = uid("enc", "robert-ed-chestpain-2024-11-04")  # NSTEMI presentation
ENC_JAMES_ED         = uid("enc", "james-ed-copd-2025-01-04")        # COPD exacerbation -> MICU
ENC_DOROTHY_ED       = uid("enc", "dorothy-ed-hipfx-2024-11-18")     # hip fracture
ENC_CARLOS_ED        = uid("enc", "carlos-ed-chestpain-2024-12-28")  # negative
ENC_EMILY_ED         = uid("enc", "emily-ed-asthma-2024-12-15")      # asthma exacerbation

# Inpatient encounters
ENC_ROBERT_INPT      = uid("enc", "robert-inpatient-nstemi-2024-11-04")
ENC_JAMES_INPT       = uid("enc", "james-inpatient-copd-2025-01-04")
ENC_DOROTHY_INPT     = uid("enc", "dorothy-inpatient-hipfx-2024-11-18")

ENCOUNTERS = [
    # Outpatient
    {"id": ENC_MARGARET_FU,     "patient_id": P_MARGARET, "encounter_class": "ambulatory","status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "Routine diabetes/HTN follow-up",                                           "attending_provider_id": PROV_MORGAN, "admitting_provider_id": None,         "location_id": LOC_CLINIC,   "department_id": DEPT_FAMILY, "appointment_id": APPT_MARGARET_FU,    "admitted_at": ts(2025,1,13,9,5,0),   "discharged_at": ts(2025,1,13,9,38,0), "discharge_disposition": "home",          "triage_acuity": None},
    {"id": ENC_ROBERT_CARDIO,   "patient_id": P_ROBERT,   "encounter_class": "ambulatory","status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "Post-MI follow-up",                                                          "attending_provider_id": PROV_LISA,   "admitting_provider_id": None,         "location_id": LOC_HOSPITAL, "department_id": DEPT_CARDIO, "appointment_id": APPT_ROBERT_CARDIO,  "admitted_at": ts(2024,12,12,8,35,0), "discharged_at": ts(2024,12,12,9,10,0),"discharge_disposition": "home",         "triage_acuity": None},
    {"id": ENC_LAKSHMI_FU,      "patient_id": P_LAKSHMI,  "encounter_class": "ambulatory","status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "Asthma controller refill",                                                   "attending_provider_id": PROV_MORGAN, "admitting_provider_id": None,         "location_id": LOC_CLINIC,   "department_id": DEPT_FAMILY, "appointment_id": APPT_LAKSHMI_FU,     "admitted_at": ts(2025,1,13,10,2,0),  "discharged_at": ts(2025,1,13,10,22,0),"discharge_disposition": "home",         "triage_acuity": None},
    {"id": ENC_AISHA_OB,        "patient_id": P_AISHA,    "encounter_class": "ambulatory","status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "32-week prenatal visit",                                                     "attending_provider_id": PROV_SARA,   "admitting_provider_id": None,         "location_id": LOC_HOSPITAL, "department_id": DEPT_OBGYN,  "appointment_id": APPT_AISHA_OB,       "admitted_at": ts(2025,1,16,9,25,0),  "discharged_at": ts(2025,1,16,9,52,0), "discharge_disposition": "home",          "triage_acuity": None},
    {"id": ENC_MARCUS_LBP,      "patient_id": P_MARCUS,   "encounter_class": "ambulatory","status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "Low back pain flare",                                                        "attending_provider_id": PROV_MORGAN, "admitting_provider_id": None,         "location_id": LOC_CLINIC,   "department_id": DEPT_FAMILY, "appointment_id": APPT_MARCUS_LBP,     "admitted_at": ts(2025,1,8,15,10,0),  "discharged_at": ts(2025,1,8,15,32,0), "discharge_disposition": "home",          "triage_acuity": None},
    {"id": ENC_DAVID_NEWPT,     "patient_id": P_DAVID,    "encounter_class": "ambulatory","status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "Polyuria, polydipsia, weight loss",                                          "attending_provider_id": PROV_RAJESH, "admitting_provider_id": None,         "location_id": LOC_CLINIC,   "department_id": DEPT_IM,     "appointment_id": APPT_DAVID_NEWPT,    "admitted_at": ts(2024,12,19,13,8,0), "discharged_at": ts(2024,12,19,13,56,0),"discharge_disposition": "home",        "triage_acuity": None},
    {"id": ENC_DOROTHY_PREOP,   "patient_id": P_DOROTHY,  "encounter_class": "ambulatory","status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "Pre-op clearance for hip arthroplasty",                                       "attending_provider_id": PROV_RAJESH, "admitting_provider_id": None,         "location_id": LOC_CLINIC,   "department_id": DEPT_IM,     "appointment_id": APPT_DOROTHY_PREOP,  "admitted_at": ts(2024,11,26,14,5,0), "discharged_at": ts(2024,11,26,14,48,0),"discharge_disposition": "home",        "triage_acuity": None},
    {"id": ENC_CARLOS_TELE,     "patient_id": P_CARLOS,   "encounter_class": "virtual",   "status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "Telehealth anxiety follow-up",                                                "attending_provider_id": PROV_RAJESH, "admitting_provider_id": None,         "location_id": LOC_CLINIC,   "department_id": DEPT_IM,     "appointment_id": APPT_CARLOS_TELE,    "admitted_at": ts(2025,1,9,13,32,0),  "discharged_at": ts(2025,1,9,13,55,0), "discharge_disposition": "home",          "triage_acuity": None},
    {"id": ENC_EMILY_URGENT,    "patient_id": P_EMILY,    "encounter_class": "ambulatory","status": "finished",   "admission_type": None,        "admission_source": None,       "chief_complaint": "Asthma worsening with wildfire smoke",                                        "attending_provider_id": PROV_MORGAN, "admitting_provider_id": None,         "location_id": LOC_CLINIC,   "department_id": DEPT_FAMILY, "appointment_id": APPT_EMILY_URGENT,   "admitted_at": ts(2024,10,22,14,45,0),"discharged_at": ts(2024,10,22,15,5,0),"discharge_disposition": "home",         "triage_acuity": None},

    # ED visits
    {"id": ENC_ROBERT_ED,       "patient_id": P_ROBERT,   "encounter_class": "emergency", "status": "finished",   "admission_type": "emergency", "admission_source": "self",     "chief_complaint": "Substernal chest pressure x 90 min, radiating to L arm, diaphoresis",        "attending_provider_id": PROV_HASAN,  "admitting_provider_id": PROV_LISA,    "location_id": LOC_HOSPITAL, "department_id": DEPT_ED,     "appointment_id": None,                  "admitted_at": ts(2024,11,4,16,12,0), "discharged_at": ts(2024,11,4,19,42,0),"discharge_disposition": "admitted",     "triage_acuity": 2},
    {"id": ENC_JAMES_ED,        "patient_id": P_JAMES,    "encounter_class": "emergency", "status": "finished",   "admission_type": "emergency", "admission_source": "ems",      "chief_complaint": "Severe dyspnea x 3 days, sputum production, fever 38.4",                     "attending_provider_id": PROV_JOHANNA,"admitting_provider_id": PROV_AMANDA,  "location_id": LOC_HOSPITAL, "department_id": DEPT_ED,     "appointment_id": None,                  "admitted_at": ts(2025,1,4,2,30,0),   "discharged_at": ts(2025,1,4,5,15,0),  "discharge_disposition": "admitted",      "triage_acuity": 2},
    {"id": ENC_DOROTHY_ED,      "patient_id": P_DOROTHY,  "encounter_class": "emergency", "status": "finished",   "admission_type": "emergency", "admission_source": "ems",      "chief_complaint": "Fall at home with R hip pain, unable to bear weight",                        "attending_provider_id": PROV_HASAN,  "admitting_provider_id": PROV_DANIEL,  "location_id": LOC_HOSPITAL, "department_id": DEPT_ED,     "appointment_id": None,                  "admitted_at": ts(2024,11,18,11,40,0),"discharged_at": ts(2024,11,18,16,20,0),"discharge_disposition": "admitted",    "triage_acuity": 3},
    {"id": ENC_CARLOS_ED,       "patient_id": P_CARLOS,   "encounter_class": "emergency", "status": "finished",   "admission_type": "emergency", "admission_source": "self",     "chief_complaint": "Sudden chest pain with palpitations and shortness of breath",                "attending_provider_id": PROV_JOHANNA,"admitting_provider_id": None,         "location_id": LOC_HOSPITAL, "department_id": DEPT_ED,     "appointment_id": None,                  "admitted_at": ts(2024,12,28,21,15,0),"discharged_at": ts(2024,12,29,1,10,0),"discharge_disposition": "home",         "triage_acuity": 3},
    {"id": ENC_EMILY_ED,        "patient_id": P_EMILY,    "encounter_class": "emergency", "status": "finished",   "admission_type": "emergency", "admission_source": "self",     "chief_complaint": "Asthma exacerbation - SOB, wheezing, no relief from inhaler",                "attending_provider_id": PROV_HASAN,  "admitting_provider_id": None,         "location_id": LOC_HOSPITAL, "department_id": DEPT_ED,     "appointment_id": None,                  "admitted_at": ts(2024,12,15,22,5,0), "discharged_at": ts(2024,12,16,0,40,0),"discharge_disposition": "home",         "triage_acuity": 3},

    # Inpatient
    {"id": ENC_ROBERT_INPT,     "patient_id": P_ROBERT,   "encounter_class": "inpatient", "status": "finished",   "admission_type": "emergency", "admission_source": "ed",       "chief_complaint": "NSTEMI - cath/PCI to mid-LAD with DES",                                       "attending_provider_id": PROV_LISA,   "admitting_provider_id": PROV_LISA,    "location_id": LOC_HOSPITAL, "department_id": DEPT_CARDIO, "appointment_id": None,                  "admitted_at": ts(2024,11,4,19,42,0), "discharged_at": ts(2024,11,7,11,30,0),"discharge_disposition": "home",         "triage_acuity": None},
    {"id": ENC_JAMES_INPT,      "patient_id": P_JAMES,    "encounter_class": "inpatient", "status": "finished",   "admission_type": "emergency", "admission_source": "ed",       "chief_complaint": "COPD exacerbation requiring BiPAP and ICU monitoring",                       "attending_provider_id": PROV_AMANDA, "admitting_provider_id": PROV_AMANDA,  "location_id": LOC_HOSPITAL, "department_id": DEPT_ICU,    "appointment_id": None,                  "admitted_at": ts(2025,1,4,5,15,0),   "discharged_at": ts(2025,1,9,14,20,0), "discharge_disposition": "home_with_hh",  "triage_acuity": None},
    {"id": ENC_DOROTHY_INPT,    "patient_id": P_DOROTHY,  "encounter_class": "inpatient", "status": "finished",   "admission_type": "emergency", "admission_source": "ed",       "chief_complaint": "R femoral neck fracture - elective THA after pre-op clearance",              "attending_provider_id": PROV_DANIEL, "admitting_provider_id": PROV_DANIEL,  "location_id": LOC_HOSPITAL, "department_id": DEPT_ORTHO,  "appointment_id": None,                  "admitted_at": ts(2024,11,18,16,20,0),"discharged_at": ts(2024,11,22,13,0,0),"discharge_disposition": "snf",          "triage_acuity": None},
]

# Bed assignments (only for inpatient/ED admits we tracked beds for)
_ED_BED_A1 = BEDS[0]["id"]                                       # ED-01-A
_ED_BED_A2 = BEDS[1]["id"]                                       # ED-02-A
_ED_BED_A3 = BEDS[2]["id"]                                       # ED-03-A
_MS_BED_301A = next(b["id"] for b in BEDS if "301" in (next((r["room_number"] for r in ROOMS if r["id"]==b["room_id"]),"")) and b["bed_label"]=="A")
_MS_BED_402A = next(b["id"] for b in BEDS if "402" in (next((r["room_number"] for r in ROOMS if r["id"]==b["room_id"]),"")) and b["bed_label"]=="A")
_MS_BED_404B = next(b["id"] for b in BEDS if "404" in (next((r["room_number"] for r in ROOMS if r["id"]==b["room_id"]),"")) and b["bed_label"]=="B")
_ICU_BED_1A  = next(b["id"] for b in BEDS if next((r["room_number"] for r in ROOMS if r["id"]==b["room_id"]),"")=="ICU-1" and b["bed_label"]=="A")
_ICU_BED_2A  = next(b["id"] for b in BEDS if next((r["room_number"] for r in ROOMS if r["id"]==b["room_id"]),"")=="ICU-2" and b["bed_label"]=="A")

BED_ASSIGNMENTS = [
    {"id": uid("ba", "robert-ed"),       "encounter_id": ENC_ROBERT_ED,    "bed_id": _ED_BED_A1,    "assigned_at": ts(2024,11,4,16,20,0),  "released_at": ts(2024,11,4,19,42,0), "assigned_by_provider_id": PROV_RYAN_RN, "reason": "triage"},
    {"id": uid("ba", "robert-ms-3w"),    "encounter_id": ENC_ROBERT_INPT,  "bed_id": _MS_BED_301A,  "assigned_at": ts(2024,11,4,19,55,0),  "released_at": ts(2024,11,7,11,30,0), "assigned_by_provider_id": PROV_PRIYA,   "reason": "admission"},
    {"id": uid("ba", "james-ed"),        "encounter_id": ENC_JAMES_ED,     "bed_id": _ED_BED_A2,    "assigned_at": ts(2025,1,4,2,40,0),    "released_at": ts(2025,1,4,5,15,0),   "assigned_by_provider_id": PROV_RYAN_RN, "reason": "triage"},
    {"id": uid("ba", "james-micu"),      "encounter_id": ENC_JAMES_INPT,   "bed_id": _ICU_BED_1A,   "assigned_at": ts(2025,1,4,5,25,0),    "released_at": ts(2025,1,7,10,0,0),   "assigned_by_provider_id": PROV_MAYA_RN, "reason": "icu_admission"},
    {"id": uid("ba", "james-stepdown"),  "encounter_id": ENC_JAMES_INPT,   "bed_id": _MS_BED_402A,  "assigned_at": ts(2025,1,7,10,15,0),   "released_at": ts(2025,1,9,14,20,0),  "assigned_by_provider_id": PROV_PRIYA,   "reason": "transfer_out_icu"},
    {"id": uid("ba", "dorothy-ed"),      "encounter_id": ENC_DOROTHY_ED,   "bed_id": _ED_BED_A3,    "assigned_at": ts(2024,11,18,11,50,0), "released_at": ts(2024,11,18,16,20,0),"assigned_by_provider_id": PROV_RYAN_RN, "reason": "triage"},
    {"id": uid("ba", "dorothy-ortho"),   "encounter_id": ENC_DOROTHY_INPT, "bed_id": _MS_BED_404B,  "assigned_at": ts(2024,11,18,16,35,0), "released_at": ts(2024,11,22,13,0,0), "assigned_by_provider_id": PROV_DANIEL,  "reason": "post_op"},
]

ENCOUNTER_DIAGNOSES = [
    # Margaret routine fu
    {"id": uid("encdx", "margaret-fu-1"), "encounter_id": ENC_MARGARET_FU, "icd10_code": "E11.21", "diagnosis_type": "primary",   "present_on_admission": "Y", "rank": 1, "documented_by_provider_id": PROV_MORGAN, "documented_at": ts(2025,1,13,9,30,0)},
    {"id": uid("encdx", "margaret-fu-2"), "encounter_id": ENC_MARGARET_FU, "icd10_code": "I10",    "diagnosis_type": "secondary", "present_on_admission": "Y", "rank": 2, "documented_by_provider_id": PROV_MORGAN, "documented_at": ts(2025,1,13,9,30,0)},
    {"id": uid("encdx", "margaret-fu-3"), "encounter_id": ENC_MARGARET_FU, "icd10_code": "N18.3",  "diagnosis_type": "secondary", "present_on_admission": "Y", "rank": 3, "documented_by_provider_id": PROV_MORGAN, "documented_at": ts(2025,1,13,9,30,0)},
    {"id": uid("encdx", "margaret-fu-4"), "encounter_id": ENC_MARGARET_FU, "icd10_code": "E78.5",  "diagnosis_type": "secondary", "present_on_admission": "Y", "rank": 4, "documented_by_provider_id": PROV_MORGAN, "documented_at": ts(2025,1,13,9,30,0)},
    # Robert cardio fu
    {"id": uid("encdx", "robert-cardio-1"),"encounter_id": ENC_ROBERT_CARDIO,"icd10_code": "I25.10","diagnosis_type": "primary",  "present_on_admission": "Y", "rank": 1, "documented_by_provider_id": PROV_LISA,   "documented_at": ts(2024,12,12,9,0,0)},
    # Lakshmi asthma fu
    {"id": uid("encdx", "lakshmi-fu-1"),  "encounter_id": ENC_LAKSHMI_FU, "icd10_code": "J45.40", "diagnosis_type": "primary",   "present_on_admission": "Y", "rank": 1, "documented_by_provider_id": PROV_MORGAN, "documented_at": ts(2025,1,13,10,20,0)},
    {"id": uid("encdx", "lakshmi-fu-2"),  "encounter_id": ENC_LAKSHMI_FU, "icd10_code": "K21.9",  "diagnosis_type": "secondary", "present_on_admission": "Y", "rank": 2, "documented_by_provider_id": PROV_MORGAN, "documented_at": ts(2025,1,13,10,20,0)},
    # Aisha OB
    {"id": uid("encdx", "aisha-ob-1"),    "encounter_id": ENC_AISHA_OB,   "icd10_code": "Z34.83", "diagnosis_type": "primary",   "present_on_admission": "Y", "rank": 1, "documented_by_provider_id": PROV_SARA,   "documented_at": ts(2025,1,16,9,50,0)},
    # Marcus LBP
    {"id": uid("encdx", "marcus-lbp-1"),  "encounter_id": ENC_MARCUS_LBP, "icd10_code": "M54.50", "diagnosis_type": "primary",   "present_on_admission": "Y", "rank": 1, "documented_by_provider_id": PROV_MORGAN, "documented_at": ts(2025,1,8,15,30,0)},
    # David new DM
    {"id": uid("encdx", "david-newdm-1"), "encounter_id": ENC_DAVID_NEWPT,"icd10_code": "E11.65", "diagnosis_type": "primary",   "present_on_admission": "Y", "rank": 1, "documented_by_provider_id": PROV_RAJESH, "documented_at": ts(2024,12,19,13,55,0)},
    {"id": uid("encdx", "david-newdm-2"), "encounter_id": ENC_DAVID_NEWPT,"icd10_code": "E11.9",  "diagnosis_type": "secondary", "present_on_admission": "Y", "rank": 2, "documented_by_provider_id": PROV_RAJESH, "documented_at": ts(2024,12,19,13,55,0)},
    # Dorothy pre-op
    {"id": uid("encdx", "dorothy-preop-1"),"encounter_id": ENC_DOROTHY_PREOP,"icd10_code":"S72.001A","diagnosis_type":"primary","present_on_admission":"Y","rank":1,"documented_by_provider_id":PROV_RAJESH,"documented_at":ts(2024,11,26,14,40,0)},
    {"id": uid("encdx", "dorothy-preop-2"),"encounter_id": ENC_DOROTHY_PREOP,"icd10_code":"G30.9","diagnosis_type":"secondary","present_on_admission":"Y","rank":2,"documented_by_provider_id":PROV_RAJESH,"documented_at":ts(2024,11,26,14,40,0)},
    # Carlos tele anxiety (no F-coded ICD listed, but for completeness reuse R07.9 + F41.1)
    {"id": uid("encdx", "carlos-tele-1"), "encounter_id": ENC_CARLOS_TELE, "icd10_code": "F41.1", "diagnosis_type": "primary",   "present_on_admission": "Y", "rank": 1, "documented_by_provider_id": PROV_RAJESH, "documented_at": ts(2025,1,9,13,50,0)},
    # Emily urgent
    {"id": uid("encdx", "emily-urg-1"),   "encounter_id": ENC_EMILY_URGENT,"icd10_code": "J45.901","diagnosis_type": "primary",  "present_on_admission": "Y", "rank": 1, "documented_by_provider_id": PROV_MORGAN, "documented_at": ts(2024,10,22,15,3,0)},
    # ED + Inpatient
    {"id": uid("encdx", "robert-ed-1"),    "encounter_id": ENC_ROBERT_ED,    "icd10_code": "R07.9", "diagnosis_type": "primary",  "present_on_admission": "Y","rank": 1,"documented_by_provider_id": PROV_HASAN,  "documented_at": ts(2024,11,4,17,0,0)},
    {"id": uid("encdx", "robert-inpt-1"),  "encounter_id": ENC_ROBERT_INPT,  "icd10_code": "I21.4", "diagnosis_type": "primary",  "present_on_admission": "Y","rank": 1,"documented_by_provider_id": PROV_LISA,   "documented_at": ts(2024,11,5,8,30,0)},
    {"id": uid("encdx", "robert-inpt-2"),  "encounter_id": ENC_ROBERT_INPT,  "icd10_code": "I25.10","diagnosis_type": "secondary","present_on_admission": "Y","rank": 2,"documented_by_provider_id": PROV_LISA,   "documented_at": ts(2024,11,5,8,30,0)},
    {"id": uid("encdx", "robert-inpt-3"),  "encounter_id": ENC_ROBERT_INPT,  "icd10_code": "E78.5", "diagnosis_type": "secondary","present_on_admission": "Y","rank": 3,"documented_by_provider_id": PROV_LISA,   "documented_at": ts(2024,11,5,8,30,0)},
    {"id": uid("encdx", "james-ed-1"),     "encounter_id": ENC_JAMES_ED,     "icd10_code": "J44.1", "diagnosis_type": "primary",  "present_on_admission": "Y","rank": 1,"documented_by_provider_id": PROV_JOHANNA,"documented_at": ts(2025,1,4,3,0,0)},
    {"id": uid("encdx", "james-inpt-1"),   "encounter_id": ENC_JAMES_INPT,   "icd10_code": "J44.1", "diagnosis_type": "primary",  "present_on_admission": "Y","rank": 1,"documented_by_provider_id": PROV_AMANDA, "documented_at": ts(2025,1,4,6,0,0)},
    {"id": uid("encdx", "james-inpt-2"),   "encounter_id": ENC_JAMES_INPT,   "icd10_code": "I50.32","diagnosis_type": "secondary","present_on_admission": "Y","rank": 2,"documented_by_provider_id": PROV_AMANDA, "documented_at": ts(2025,1,4,6,0,0)},
    {"id": uid("encdx", "james-inpt-3"),   "encounter_id": ENC_JAMES_INPT,   "icd10_code": "J18.9", "diagnosis_type": "secondary","present_on_admission": "Y","rank": 3,"documented_by_provider_id": PROV_AMANDA, "documented_at": ts(2025,1,4,6,0,0)},
    {"id": uid("encdx", "dorothy-ed-1"),   "encounter_id": ENC_DOROTHY_ED,   "icd10_code": "S72.001A","diagnosis_type":"primary", "present_on_admission":"Y", "rank":1,"documented_by_provider_id": PROV_HASAN,  "documented_at": ts(2024,11,18,12,30,0)},
    {"id": uid("encdx", "dorothy-inpt-1"), "encounter_id": ENC_DOROTHY_INPT, "icd10_code": "S72.001A","diagnosis_type":"primary", "present_on_admission":"Y", "rank":1,"documented_by_provider_id": PROV_DANIEL, "documented_at": ts(2024,11,18,17,0,0)},
    {"id": uid("encdx", "dorothy-inpt-2"), "encounter_id": ENC_DOROTHY_INPT, "icd10_code": "G30.9", "diagnosis_type": "secondary","present_on_admission": "Y","rank": 2,"documented_by_provider_id": PROV_DANIEL, "documented_at": ts(2024,11,18,17,0,0)},
    {"id": uid("encdx", "carlos-ed-1"),    "encounter_id": ENC_CARLOS_ED,    "icd10_code": "R07.9", "diagnosis_type": "primary",  "present_on_admission": "Y","rank": 1,"documented_by_provider_id": PROV_JOHANNA,"documented_at": ts(2024,12,28,22,0,0)},
    {"id": uid("encdx", "carlos-ed-2"),    "encounter_id": ENC_CARLOS_ED,    "icd10_code": "F41.1", "diagnosis_type": "secondary","present_on_admission": "Y","rank": 2,"documented_by_provider_id": PROV_JOHANNA,"documented_at": ts(2024,12,28,22,0,0)},
    {"id": uid("encdx", "emily-ed-1"),     "encounter_id": ENC_EMILY_ED,     "icd10_code": "J45.901","diagnosis_type":"primary",  "present_on_admission":"Y","rank":1,"documented_by_provider_id": PROV_HASAN,  "documented_at": ts(2024,12,15,22,30,0)},
]

ENCOUNTER_PROCEDURES = [
    # Margaret routine fu (E/M only)
    {"id": uid("encproc", "margaret-fu-em"),     "encounter_id": ENC_MARGARET_FU,     "cpt_code": "99214", "performed_at": ts(2025,1,13,9,30,0),  "performing_provider_id": PROV_MORGAN, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "30-min established patient visit, moderate MDM"},
    {"id": uid("encproc", "margaret-fu-bld"),    "encounter_id": ENC_MARGARET_FU,     "cpt_code": "36415", "performed_at": ts(2025,1,13,9,35,0),  "performing_provider_id": PROV_MORGAN, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "Routine venipuncture for CMP/CBC/A1c"},
    # Lakshmi
    {"id": uid("encproc", "lakshmi-fu-em"),      "encounter_id": ENC_LAKSHMI_FU,      "cpt_code": "99213", "performed_at": ts(2025,1,13,10,20,0), "performing_provider_id": PROV_MORGAN, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "Asthma follow-up, low MDM"},
    # Aisha OB - prenatal E/M + US
    {"id": uid("encproc", "aisha-em"),           "encounter_id": ENC_AISHA_OB,        "cpt_code": "99214", "performed_at": ts(2025,1,16,9,50,0),  "performing_provider_id": PROV_SARA,   "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Routine 32-week prenatal visit"},
    {"id": uid("encproc", "aisha-tdap"),         "encounter_id": ENC_AISHA_OB,        "cpt_code": "90471", "performed_at": ts(2025,1,16,9,55,0),  "performing_provider_id": PROV_SARA,   "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Tdap administered IM L deltoid"},
    # Marcus
    {"id": uid("encproc", "marcus-lbp-em"),      "encounter_id": ENC_MARCUS_LBP,      "cpt_code": "99213", "performed_at": ts(2025,1,8,15,30,0),  "performing_provider_id": PROV_MORGAN, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "LBP flare, conservative mgmt"},
    # David
    {"id": uid("encproc", "david-newpt-em"),     "encounter_id": ENC_DAVID_NEWPT,     "cpt_code": "99203", "performed_at": ts(2024,12,19,13,55,0),"performing_provider_id": PROV_RAJESH, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "New patient eval, low MDM"},
    {"id": uid("encproc", "david-bld"),          "encounter_id": ENC_DAVID_NEWPT,     "cpt_code": "36415", "performed_at": ts(2024,12,19,13,50,0),"performing_provider_id": PROV_RAJESH, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "Venipuncture for CMP/A1c/lipid"},
    # Dorothy pre-op
    {"id": uid("encproc", "dorothy-preop-em"),   "encounter_id": ENC_DOROTHY_PREOP,   "cpt_code": "99214", "performed_at": ts(2024,11,26,14,45,0),"performing_provider_id": PROV_RAJESH, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "Pre-op clearance moderate MDM"},
    {"id": uid("encproc", "dorothy-ecg"),        "encounter_id": ENC_DOROTHY_PREOP,   "cpt_code": "93000", "performed_at": ts(2024,11,26,14,30,0),"performing_provider_id": PROV_RAJESH, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "Routine ECG, NSR rate 76, no acute changes"},
    # Carlos tele
    {"id": uid("encproc", "carlos-tele-em"),     "encounter_id": ENC_CARLOS_TELE,     "cpt_code": "99213", "performed_at": ts(2025,1,9,13,50,0),  "performing_provider_id": PROV_RAJESH, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": "95",    "notes": "Telehealth E/M (mod 95)"},
    # Emily urgent (in clinic)
    {"id": uid("encproc", "emily-urg-em"),       "encounter_id": ENC_EMILY_URGENT,    "cpt_code": "99213", "performed_at": ts(2024,10,22,15,3,0), "performing_provider_id": PROV_MORGAN, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "Same-day urgent visit"},
    {"id": uid("encproc", "emily-urg-neb"),      "encounter_id": ENC_EMILY_URGENT,    "cpt_code": "94640", "performed_at": ts(2024,10,22,14,55,0),"performing_provider_id": PROV_MORGAN, "assistant_provider_id": None,        "location_id": LOC_CLINIC,   "modifier": None,    "notes": "Albuterol nebulization x 1"},
    # ED + Inpatient
    {"id": uid("encproc", "robert-ed-em"),       "encounter_id": ENC_ROBERT_ED,       "cpt_code": "99285", "performed_at": ts(2024,11,4,17,0,0),  "performing_provider_id": PROV_HASAN,  "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "ED E/M high severity"},
    {"id": uid("encproc", "robert-ed-ekg"),      "encounter_id": ENC_ROBERT_ED,       "cpt_code": "93000", "performed_at": ts(2024,11,4,16,25,0), "performing_provider_id": PROV_HASAN,  "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Serial EKGs - dynamic ST changes V4-V6"},
    {"id": uid("encproc", "robert-ed-trop"),     "encounter_id": ENC_ROBERT_ED,       "cpt_code": "84484", "performed_at": ts(2024,11,4,16,45,0), "performing_provider_id": PROV_HASAN,  "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Troponin elevated 0.42 ng/mL"},
    {"id": uid("encproc", "robert-inpt-em"),     "encounter_id": ENC_ROBERT_INPT,     "cpt_code": "99221", "performed_at": ts(2024,11,4,20,30,0), "performing_provider_id": PROV_LISA,   "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Admission H&P"},
    {"id": uid("encproc", "robert-inpt-em2"),    "encounter_id": ENC_ROBERT_INPT,     "cpt_code": "99232", "performed_at": ts(2024,11,5,8,30,0),  "performing_provider_id": PROV_LISA,   "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "HD 1 progress note"},
    {"id": uid("encproc", "robert-inpt-em3"),    "encounter_id": ENC_ROBERT_INPT,     "cpt_code": "99232", "performed_at": ts(2024,11,6,8,30,0),  "performing_provider_id": PROV_LISA,   "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "HD 2 progress note"},
    {"id": uid("encproc", "robert-echo"),        "encounter_id": ENC_ROBERT_CARDIO,   "cpt_code": "93306", "performed_at": ts(2024,12,12,8,40,0), "performing_provider_id": PROV_LISA,   "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "TTE: LVEF 50%, mild anterior hypokinesis"},
    {"id": uid("encproc", "james-ed-em"),        "encounter_id": ENC_JAMES_ED,        "cpt_code": "99285", "performed_at": ts(2025,1,4,3,0,0),    "performing_provider_id": PROV_JOHANNA,"assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "ED critical illness"},
    {"id": uid("encproc", "james-ed-cxr"),       "encounter_id": ENC_JAMES_ED,        "cpt_code": "71046", "performed_at": ts(2025,1,4,3,15,0),   "performing_provider_id": PROV_KATHRYN,"assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "CXR PA/lat: hyperinflation, RLL infiltrate"},
    {"id": uid("encproc", "james-micu-crit"),    "encounter_id": ENC_JAMES_INPT,      "cpt_code": "99291", "performed_at": ts(2025,1,4,6,15,0),   "performing_provider_id": PROV_AMANDA, "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Critical care 75 min - BiPAP titration"},
    {"id": uid("encproc", "james-inpt-em"),      "encounter_id": ENC_JAMES_INPT,      "cpt_code": "99232", "performed_at": ts(2025,1,5,8,0,0),    "performing_provider_id": PROV_AMANDA, "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "HD 1 progress note"},
    {"id": uid("encproc", "james-inpt-em2"),     "encounter_id": ENC_JAMES_INPT,      "cpt_code": "99232", "performed_at": ts(2025,1,6,8,0,0),    "performing_provider_id": PROV_AMANDA, "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "HD 2 progress note - weaning BiPAP"},
    {"id": uid("encproc", "dorothy-ed-em"),      "encounter_id": ENC_DOROTHY_ED,      "cpt_code": "99284", "performed_at": ts(2024,11,18,12,30,0),"performing_provider_id": PROV_HASAN,  "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "ED moderate severity"},
    {"id": uid("encproc", "dorothy-ed-xr"),      "encounter_id": ENC_DOROTHY_ED,      "cpt_code": "71046", "performed_at": ts(2024,11,18,12,45,0),"performing_provider_id": PROV_KATHRYN,"assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": "RT",  "notes": "R hip XR: displaced femoral neck fx"},
    {"id": uid("encproc", "dorothy-tha"),        "encounter_id": ENC_DOROTHY_INPT,    "cpt_code": "27130", "performed_at": ts(2024,11,19,8,0,0),  "performing_provider_id": PROV_DANIEL, "assistant_provider_id": PROV_BRIAN_PA,"location_id": LOC_HOSPITAL, "modifier": "RT",  "notes": "R total hip arthroplasty, posterior approach, uncemented stem"},
    {"id": uid("encproc", "dorothy-inpt-em"),    "encounter_id": ENC_DOROTHY_INPT,    "cpt_code": "99221", "performed_at": ts(2024,11,18,17,30,0),"performing_provider_id": PROV_DANIEL, "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Pre-op H&P"},
    {"id": uid("encproc", "carlos-ed-em"),       "encounter_id": ENC_CARLOS_ED,       "cpt_code": "99284", "performed_at": ts(2024,12,28,22,0,0), "performing_provider_id": PROV_JOHANNA,"assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "ED moderate severity"},
    {"id": uid("encproc", "carlos-ed-trop"),     "encounter_id": ENC_CARLOS_ED,       "cpt_code": "84484", "performed_at": ts(2024,12,28,22,30,0),"performing_provider_id": PROV_JOHANNA,"assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Troponin x 2 negative"},
    {"id": uid("encproc", "carlos-ed-ekg"),      "encounter_id": ENC_CARLOS_ED,       "cpt_code": "93000", "performed_at": ts(2024,12,28,21,45,0),"performing_provider_id": PROV_JOHANNA,"assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "ECG: NSR, no ST changes"},
    {"id": uid("encproc", "emily-ed-em"),        "encounter_id": ENC_EMILY_ED,        "cpt_code": "99284", "performed_at": ts(2024,12,15,22,30,0),"performing_provider_id": PROV_HASAN,  "assistant_provider_id": None,        "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "ED moderate severity asthma"},
    {"id": uid("encproc", "emily-ed-neb"),       "encounter_id": ENC_EMILY_ED,        "cpt_code": "94640", "performed_at": ts(2024,12,15,22,40,0),"performing_provider_id": PROV_BRIAN_PA,"assistant_provider_id": None,       "location_id": LOC_HOSPITAL, "modifier": None,    "notes": "Albuterol/ipratropium neb x 3"},
]

# ---------------------------------------------------------------------------
# 7. CLINICAL — problems, allergies, vitals, observations, care plans
# ---------------------------------------------------------------------------

PROBLEM_LIST_ENTRIES = [
    # Margaret
    {"id": uid("prob", "margaret-t2dm"),  "patient_id": P_MARGARET, "icd10_code": "E11.21", "snomed_code": "44054006",  "description": "Type 2 diabetes mellitus with diabetic nephropathy",      "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2003-09-12", "resolved_date": None, "recorded_by_provider_id": PROV_MORGAN},
    {"id": uid("prob", "margaret-htn"),   "patient_id": P_MARGARET, "icd10_code": "I10",    "snomed_code": "38341003",  "description": "Essential hypertension",                                  "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2001-04-22", "resolved_date": None, "recorded_by_provider_id": PROV_MORGAN},
    {"id": uid("prob", "margaret-ckd"),   "patient_id": P_MARGARET, "icd10_code": "N18.3",  "snomed_code": "709044004", "description": "Chronic kidney disease stage 3",                          "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2019-06-04", "resolved_date": None, "recorded_by_provider_id": PROV_MORGAN},
    {"id": uid("prob", "margaret-hld"),   "patient_id": P_MARGARET, "icd10_code": "E78.5",  "snomed_code": "13644009",  "description": "Hyperlipidemia",                                          "clinical_status": "active",   "verification_status": "confirmed", "severity": "mild",     "onset_date": "2008-11-15", "resolved_date": None, "recorded_by_provider_id": PROV_MORGAN},
    # Robert
    {"id": uid("prob", "robert-cad"),     "patient_id": P_ROBERT,   "icd10_code": "I25.10", "snomed_code": None,        "description": "Atherosclerotic heart disease native coronary, no angina","clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2024-11-04", "resolved_date": None, "recorded_by_provider_id": PROV_LISA},
    {"id": uid("prob", "robert-hld"),     "patient_id": P_ROBERT,   "icd10_code": "E78.5",  "snomed_code": "13644009",  "description": "Hyperlipidemia",                                          "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2018-09-10", "resolved_date": None, "recorded_by_provider_id": PROV_RAJESH},
    {"id": uid("prob", "robert-htn"),     "patient_id": P_ROBERT,   "icd10_code": "I10",    "snomed_code": "38341003",  "description": "Essential hypertension",                                  "clinical_status": "active",   "verification_status": "confirmed", "severity": "mild",     "onset_date": "2020-03-18", "resolved_date": None, "recorded_by_provider_id": PROV_RAJESH},
    # Lakshmi
    {"id": uid("prob", "lakshmi-asthma"), "patient_id": P_LAKSHMI,  "icd10_code": "J45.40", "snomed_code": "195967001", "description": "Moderate persistent asthma",                              "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2001-05-22", "resolved_date": None, "recorded_by_provider_id": PROV_MORGAN},
    {"id": uid("prob", "lakshmi-gerd"),   "patient_id": P_LAKSHMI,  "icd10_code": "K21.9",  "snomed_code": "235595009", "description": "Gastroesophageal reflux disease without esophagitis",     "clinical_status": "active",   "verification_status": "confirmed", "severity": "mild",     "onset_date": "2017-02-14", "resolved_date": None, "recorded_by_provider_id": PROV_MORGAN},
    # James
    {"id": uid("prob", "james-copd"),     "patient_id": P_JAMES,    "icd10_code": "J44.1",  "snomed_code": "13645005",  "description": "COPD GOLD stage D with acute exacerbations",              "clinical_status": "active",   "verification_status": "confirmed", "severity": "severe",   "onset_date": "2014-08-05", "resolved_date": None, "recorded_by_provider_id": PROV_AMANDA},
    {"id": uid("prob", "james-chf"),      "patient_id": P_JAMES,    "icd10_code": "I50.32", "snomed_code": "84114007",  "description": "Chronic diastolic heart failure",                         "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2021-01-20", "resolved_date": None, "recorded_by_provider_id": PROV_LISA},
    {"id": uid("prob", "james-htn"),      "patient_id": P_JAMES,    "icd10_code": "I10",    "snomed_code": "38341003",  "description": "Essential hypertension",                                  "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "1998-11-04", "resolved_date": None, "recorded_by_provider_id": PROV_RAJESH},
    # Aisha
    {"id": uid("prob", "aisha-preg"),     "patient_id": P_AISHA,    "icd10_code": "Z34.83", "snomed_code": None,        "description": "Encounter for supervision of normal pregnancy, third trimester","clinical_status":"active","verification_status":"confirmed","severity":"mild","onset_date":"2024-09-01","resolved_date":None,"recorded_by_provider_id":PROV_SARA},
    # Marcus
    {"id": uid("prob", "marcus-lbp"),     "patient_id": P_MARCUS,   "icd10_code": "M54.50", "snomed_code": "279039007", "description": "Chronic low back pain, mechanical",                       "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2021-06-10", "resolved_date": None, "recorded_by_provider_id": PROV_MORGAN},
    # Dorothy
    {"id": uid("prob", "dorothy-alz"),    "patient_id": P_DOROTHY,  "icd10_code": "G30.9",  "snomed_code": "26929004",  "description": "Alzheimer's disease, moderate cognitive impairment",      "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2020-04-12", "resolved_date": None, "recorded_by_provider_id": PROV_RAJESH},
    {"id": uid("prob", "dorothy-hipfx"),  "patient_id": P_DOROTHY,  "icd10_code": "S72.001A","snomed_code": None,       "description": "Right femoral neck fracture, status post THA",            "clinical_status": "resolved", "verification_status": "confirmed", "severity": "severe",   "onset_date": "2024-11-18", "resolved_date": "2024-11-22", "recorded_by_provider_id": PROV_DANIEL},
    # Carlos
    {"id": uid("prob", "carlos-gad"),     "patient_id": P_CARLOS,   "icd10_code": "F41.1",  "snomed_code": None,        "description": "Generalized anxiety disorder",                            "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2020-08-15", "resolved_date": None, "recorded_by_provider_id": PROV_RAJESH},
    # Emily
    {"id": uid("prob", "emily-asthma"),   "patient_id": P_EMILY,    "icd10_code": "J45.40", "snomed_code": "195967001", "description": "Moderate persistent asthma since childhood",              "clinical_status": "active",   "verification_status": "confirmed", "severity": "moderate", "onset_date": "2009-03-04", "resolved_date": None, "recorded_by_provider_id": PROV_MORGAN},
    # David
    {"id": uid("prob", "david-newdm"),    "patient_id": P_DAVID,    "icd10_code": "E11.65", "snomed_code": "44054006",  "description": "Type 2 diabetes mellitus with hyperglycemia, newly diagnosed","clinical_status":"active","verification_status":"confirmed","severity":"moderate","onset_date":"2024-12-19","resolved_date":None,"recorded_by_provider_id":PROV_RAJESH},
]

ALG_MARGARET_PCN = uid("allergy", "margaret-penicillin")
ALG_ROBERT_SULFA = uid("allergy", "robert-sulfa")
ALG_LAKSHMI_LATEX= uid("allergy", "lakshmi-latex")
ALG_JAMES_NKDA   = uid("allergy", "james-nkda")
ALG_AISHA_PB     = uid("allergy", "aisha-peanut")
ALG_DOROTHY_CDN  = uid("allergy", "dorothy-codeine")
ALG_DAVID_IODINE = uid("allergy", "david-iodine-contrast")
ALG_EMILY_ASA    = uid("allergy", "emily-aspirin")

ALLERGIES = [
    {"id": ALG_MARGARET_PCN, "patient_id": P_MARGARET, "allergen_type": "medication", "allergen_name": "Penicillin V",       "rxnorm_code": "308191","criticality": "high",   "clinical_status": "active", "verification_status": "confirmed","recorded_date": "1992-06-15", "last_occurrence_date": "1992-06-15"},
    {"id": ALG_ROBERT_SULFA, "patient_id": P_ROBERT,   "allergen_type": "medication", "allergen_name": "Sulfa drugs",         "rxnorm_code": None,    "criticality": "high",   "clinical_status": "active", "verification_status": "confirmed","recorded_date": "2010-04-08", "last_occurrence_date": "2010-04-08"},
    {"id": ALG_LAKSHMI_LATEX,"patient_id": P_LAKSHMI,  "allergen_type": "environment","allergen_name": "Natural rubber latex",  "rxnorm_code": None,    "criticality": "low",    "clinical_status": "active", "verification_status": "confirmed","recorded_date": "2015-11-20", "last_occurrence_date": "2022-08-14"},
    {"id": ALG_JAMES_NKDA,   "patient_id": P_JAMES,    "allergen_type": "no_known",   "allergen_name": "No known drug allergies","rxnorm_code": None,  "criticality": None,     "clinical_status": "active", "verification_status": "confirmed","recorded_date": "2014-08-10", "last_occurrence_date": None},
    {"id": ALG_AISHA_PB,     "patient_id": P_AISHA,    "allergen_type": "food",       "allergen_name": "Peanut",               "rxnorm_code": None,    "criticality": "high",   "clinical_status": "active", "verification_status": "confirmed","recorded_date": "1999-09-12", "last_occurrence_date": "2019-07-22"},
    {"id": ALG_DOROTHY_CDN,  "patient_id": P_DOROTHY,  "allergen_type": "medication", "allergen_name": "Codeine",              "rxnorm_code": None,    "criticality": "moderate","clinical_status": "active","verification_status": "confirmed","recorded_date": "2005-02-14", "last_occurrence_date": "2005-02-14"},
    {"id": ALG_DAVID_IODINE, "patient_id": P_DAVID,    "allergen_type": "medication", "allergen_name": "Iodinated IV contrast","rxnorm_code": None,    "criticality": "high",   "clinical_status": "active", "verification_status": "confirmed","recorded_date": "2018-04-30", "last_occurrence_date": "2018-04-30"},
    {"id": ALG_EMILY_ASA,    "patient_id": P_EMILY,    "allergen_type": "medication", "allergen_name": "Aspirin",              "rxnorm_code": "855332","criticality": "moderate","clinical_status": "active","verification_status": "confirmed","recorded_date": "2018-09-04", "last_occurrence_date": "2018-09-04"},
]

ALLERGY_REACTIONS = [
    {"id": uid("alrx", "margaret-pcn-hives"),"allergy_id": ALG_MARGARET_PCN, "manifestation": "urticaria",      "severity": "moderate","onset_minutes": 30},
    {"id": uid("alrx", "margaret-pcn-dysp"), "allergy_id": ALG_MARGARET_PCN, "manifestation": "dyspnea",        "severity": "severe",  "onset_minutes": 45},
    {"id": uid("alrx", "robert-sulfa-rash"),"allergy_id": ALG_ROBERT_SULFA,  "manifestation": "maculopapular_rash","severity": "moderate","onset_minutes": 240},
    {"id": uid("alrx", "lakshmi-latex"),    "allergy_id": ALG_LAKSHMI_LATEX, "manifestation": "contact_dermatitis","severity": "mild","onset_minutes": 60},
    {"id": uid("alrx", "aisha-peanut-anaphyl"),"allergy_id": ALG_AISHA_PB,    "manifestation": "anaphylaxis",     "severity": "severe",  "onset_minutes": 10},
    {"id": uid("alrx", "dorothy-cdn-nv"),   "allergy_id": ALG_DOROTHY_CDN,   "manifestation": "nausea_vomiting", "severity": "mild",    "onset_minutes": 30},
    {"id": uid("alrx", "david-iodine-rash"),"allergy_id": ALG_DAVID_IODINE,  "manifestation": "urticaria",       "severity": "moderate","onset_minutes": 15},
    {"id": uid("alrx", "emily-asa-bronch"), "allergy_id": ALG_EMILY_ASA,     "manifestation": "bronchospasm",    "severity": "moderate","onset_minutes": 20},
]

VITAL_SIGNS = [
    # Margaret PCP follow-up
    {"id": uid("vs", "margaret-fu"),  "patient_id": P_MARGARET, "encounter_id": ENC_MARGARET_FU, "measured_at": ts(2025,1,13,9,10,0),  "systolic_bp": 148, "diastolic_bp": 84, "heart_rate": 78, "respiratory_rate": 16, "temperature_c": 36.8, "spo2": 97, "pain_score": 2, "height_cm": 162.5, "weight_kg": 78.4, "bmi": 29.7, "recorded_by_provider_id": PROV_MORGAN},
    # Robert ED + cardio FU
    {"id": uid("vs", "robert-ed"),    "patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_ED,    "measured_at": ts(2024,11,4,16,15,0), "systolic_bp": 162, "diastolic_bp": 94, "heart_rate": 102,"respiratory_rate": 22, "temperature_c": 36.9, "spo2": 95, "pain_score": 8, "height_cm": 178.0, "weight_kg": 92.0, "bmi": 29.0, "recorded_by_provider_id": PROV_RYAN_RN},
    {"id": uid("vs", "robert-inpt"),  "patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_INPT,  "measured_at": ts(2024,11,5,7,30,0),  "systolic_bp": 132, "diastolic_bp": 78, "heart_rate": 72, "respiratory_rate": 16, "temperature_c": 36.7, "spo2": 98, "pain_score": 1, "height_cm": 178.0, "weight_kg": 91.4, "bmi": 28.9, "recorded_by_provider_id": PROV_MAYA_RN},
    {"id": uid("vs", "robert-cardio"),"patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_CARDIO,"measured_at": ts(2024,12,12,8,40,0), "systolic_bp": 124, "diastolic_bp": 76, "heart_rate": 68, "respiratory_rate": 14, "temperature_c": 36.6, "spo2": 99, "pain_score": 0, "height_cm": 178.0, "weight_kg": 89.8, "bmi": 28.3, "recorded_by_provider_id": PROV_LISA},
    # Lakshmi
    {"id": uid("vs", "lakshmi-fu"),   "patient_id": P_LAKSHMI,  "encounter_id": ENC_LAKSHMI_FU,   "measured_at": ts(2025,1,13,10,5,0),  "systolic_bp": 118, "diastolic_bp": 74, "heart_rate": 72, "respiratory_rate": 14, "temperature_c": 36.6, "spo2": 99, "pain_score": 0, "height_cm": 163.0, "weight_kg": 61.5, "bmi": 23.1, "recorded_by_provider_id": PROV_MORGAN},
    # James ED + MICU
    {"id": uid("vs", "james-ed"),     "patient_id": P_JAMES,    "encounter_id": ENC_JAMES_ED,     "measured_at": ts(2025,1,4,2,35,0),   "systolic_bp": 152, "diastolic_bp": 88, "heart_rate": 114,"respiratory_rate": 28, "temperature_c": 38.4, "spo2": 86, "pain_score": 3, "height_cm": 175.0, "weight_kg": 70.1, "bmi": 22.9, "recorded_by_provider_id": PROV_RYAN_RN},
    {"id": uid("vs", "james-micu"),   "patient_id": P_JAMES,    "encounter_id": ENC_JAMES_INPT,   "measured_at": ts(2025,1,5,7,30,0),   "systolic_bp": 134, "diastolic_bp": 78, "heart_rate": 92, "respiratory_rate": 22, "temperature_c": 37.4, "spo2": 92, "pain_score": 1, "height_cm": 175.0, "weight_kg": 70.5, "bmi": 23.0, "recorded_by_provider_id": PROV_MAYA_RN},
    # Aisha
    {"id": uid("vs", "aisha-ob"),     "patient_id": P_AISHA,    "encounter_id": ENC_AISHA_OB,     "measured_at": ts(2025,1,16,9,30,0),  "systolic_bp": 118, "diastolic_bp": 70, "heart_rate": 84, "respiratory_rate": 16, "temperature_c": 36.7, "spo2": 98, "pain_score": 0, "height_cm": 165.0, "weight_kg": 74.8, "bmi": 27.5, "recorded_by_provider_id": PROV_SARA},
    # Marcus
    {"id": uid("vs", "marcus-lbp"),   "patient_id": P_MARCUS,   "encounter_id": ENC_MARCUS_LBP,   "measured_at": ts(2025,1,8,15,15,0),  "systolic_bp": 126, "diastolic_bp": 80, "heart_rate": 76, "respiratory_rate": 14, "temperature_c": 36.7, "spo2": 99, "pain_score": 5, "height_cm": 183.0, "weight_kg": 95.5, "bmi": 28.5, "recorded_by_provider_id": PROV_MORGAN},
    # Dorothy ED + inpt
    {"id": uid("vs", "dorothy-ed"),   "patient_id": P_DOROTHY,  "encounter_id": ENC_DOROTHY_ED,   "measured_at": ts(2024,11,18,11,50,0),"systolic_bp": 142, "diastolic_bp": 78, "heart_rate": 96, "respiratory_rate": 20, "temperature_c": 36.9, "spo2": 96, "pain_score": 9, "height_cm": 155.0, "weight_kg": 52.3, "bmi": 21.8, "recorded_by_provider_id": PROV_RYAN_RN},
    {"id": uid("vs", "dorothy-postop"),"patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_INPT, "measured_at": ts(2024,11,20,7,30,0), "systolic_bp": 128, "diastolic_bp": 74, "heart_rate": 80, "respiratory_rate": 18, "temperature_c": 37.1, "spo2": 96, "pain_score": 4, "height_cm": 155.0, "weight_kg": 51.8, "bmi": 21.6, "recorded_by_provider_id": PROV_DANIEL},
    # Carlos ED + tele
    {"id": uid("vs", "carlos-ed"),    "patient_id": P_CARLOS,   "encounter_id": ENC_CARLOS_ED,    "measured_at": ts(2024,12,28,21,30,0),"systolic_bp": 138, "diastolic_bp": 86, "heart_rate": 108,"respiratory_rate": 22, "temperature_c": 36.8, "spo2": 99, "pain_score": 6, "height_cm": 170.0, "weight_kg": 84.0, "bmi": 29.1, "recorded_by_provider_id": PROV_RYAN_RN},
    # Emily ED + urgent
    {"id": uid("vs", "emily-ed"),     "patient_id": P_EMILY,    "encounter_id": ENC_EMILY_ED,     "measured_at": ts(2024,12,15,22,15,0),"systolic_bp": 124, "diastolic_bp": 80, "heart_rate": 118,"respiratory_rate": 28, "temperature_c": 37.0, "spo2": 91, "pain_score": 4, "height_cm": 167.5, "weight_kg": 60.2, "bmi": 21.5, "recorded_by_provider_id": PROV_RYAN_RN},
    {"id": uid("vs", "emily-urg"),    "patient_id": P_EMILY,    "encounter_id": ENC_EMILY_URGENT, "measured_at": ts(2024,10,22,14,48,0),"systolic_bp": 118, "diastolic_bp": 76, "heart_rate": 96, "respiratory_rate": 22, "temperature_c": 36.7, "spo2": 95, "pain_score": 2, "height_cm": 167.5, "weight_kg": 60.8, "bmi": 21.7, "recorded_by_provider_id": PROV_MORGAN},
    # David
    {"id": uid("vs", "david-newpt"),  "patient_id": P_DAVID,    "encounter_id": ENC_DAVID_NEWPT,  "measured_at": ts(2024,12,19,13,15,0),"systolic_bp": 136, "diastolic_bp": 84, "heart_rate": 82, "respiratory_rate": 16, "temperature_c": 36.7, "spo2": 98, "pain_score": 0, "height_cm": 172.0, "weight_kg": 88.5, "bmi": 29.9, "recorded_by_provider_id": PROV_RAJESH},
]

CLINICAL_OBSERVATIONS = [
    # Examples of non-lab observations: A1c values, BP narrative, depression screen, peak flow
    {"id": uid("obs", "margaret-a1c"),   "patient_id": P_MARGARET, "encounter_id": ENC_MARGARET_FU,    "loinc_code": "4548-4", "value_numeric": 7.8,    "value_text": None,    "units": "%",        "interpretation": "H", "observed_at": ts(2025,1,13,9,20,0),  "recorded_by_provider_id": PROV_MORGAN},
    {"id": uid("obs", "margaret-bp-sys"),"patient_id": P_MARGARET, "encounter_id": ENC_MARGARET_FU,    "loinc_code": "8480-6", "value_numeric": 148,    "value_text": None,    "units": "mm[Hg]",   "interpretation": "H", "observed_at": ts(2025,1,13,9,10,0),  "recorded_by_provider_id": PROV_MORGAN},
    {"id": uid("obs", "margaret-bp-dia"),"patient_id": P_MARGARET, "encounter_id": ENC_MARGARET_FU,    "loinc_code": "8462-4", "value_numeric": 84,     "value_text": None,    "units": "mm[Hg]",   "interpretation": "H", "observed_at": ts(2025,1,13,9,10,0),  "recorded_by_provider_id": PROV_MORGAN},
    {"id": uid("obs", "robert-a1c"),     "patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_CARDIO,  "loinc_code": "4548-4", "value_numeric": 5.6,    "value_text": None,    "units": "%",        "interpretation": "N", "observed_at": ts(2024,12,12,9,0,0),  "recorded_by_provider_id": PROV_LISA},
    {"id": uid("obs", "lakshmi-peakflow"),"patient_id": P_LAKSHMI, "encounter_id": ENC_LAKSHMI_FU,     "loinc_code": "33452-4","value_numeric": 410,    "value_text": "best of 3 measurements","units": "L/min", "interpretation": "N", "observed_at": ts(2025,1,13,10,12,0), "recorded_by_provider_id": PROV_MORGAN},
    {"id": uid("obs", "james-spo2-ed"),  "patient_id": P_JAMES,    "encounter_id": ENC_JAMES_ED,       "loinc_code": "2708-6", "value_numeric": 86,     "value_text": None,    "units": "%",        "interpretation": "L", "observed_at": ts(2025,1,4,2,35,0),   "recorded_by_provider_id": PROV_RYAN_RN},
    {"id": uid("obs", "james-fev1"),     "patient_id": P_JAMES,    "encounter_id": ENC_JAMES_INPT,     "loinc_code": "20150-9","value_numeric": 0.95,   "value_text": "Post-bronchodilator FEV1","units": "L",  "interpretation": "L", "observed_at": ts(2025,1,7,10,30,0),  "recorded_by_provider_id": PROV_AMANDA},
    {"id": uid("obs", "aisha-fheartrate"),"patient_id":P_AISHA,    "encounter_id": ENC_AISHA_OB,       "loinc_code": "11631-9","value_numeric": 148,    "value_text": "Doppler", "units": "{beats}/min","interpretation":"N", "observed_at": ts(2025,1,16,9,40,0),  "recorded_by_provider_id": PROV_SARA},
    {"id": uid("obs", "aisha-fundal"),   "patient_id": P_AISHA,    "encounter_id": ENC_AISHA_OB,       "loinc_code": "11879-4","value_numeric": 31,     "value_text": "appropriate for GA","units": "cm","interpretation": "N", "observed_at": ts(2025,1,16,9,42,0),  "recorded_by_provider_id": PROV_SARA},
    {"id": uid("obs", "marcus-phq9"),    "patient_id": P_MARCUS,   "encounter_id": ENC_MARCUS_LBP,     "loinc_code": "44261-6","value_numeric": 4,      "value_text": "Minimal depressive symptoms","units": "{score}","interpretation":"N","observed_at": ts(2025,1,8,15,18,0),"recorded_by_provider_id":PROV_MORGAN},
    {"id": uid("obs", "dorothy-moca"),   "patient_id": P_DOROTHY,  "encounter_id": ENC_DOROTHY_PREOP,  "loinc_code": "72172-0","value_numeric": 14,     "value_text": "Moderate cognitive impairment","units": "{score}","interpretation":"L","observed_at": ts(2024,11,26,14,30,0),"recorded_by_provider_id":PROV_RAJESH},
    {"id": uid("obs", "carlos-gad7"),    "patient_id": P_CARLOS,   "encounter_id": ENC_CARLOS_TELE,    "loinc_code": "70274-6","value_numeric": 13,     "value_text": "Moderate anxiety","units": "{score}","interpretation":"H","observed_at": ts(2025,1,9,13,40,0),"recorded_by_provider_id":PROV_RAJESH},
    {"id": uid("obs", "emily-peakflow"), "patient_id": P_EMILY,    "encounter_id": ENC_EMILY_ED,       "loinc_code": "33452-4","value_numeric": 240,    "value_text": "63% of personal best 380","units": "L/min","interpretation":"L","observed_at": ts(2024,12,15,22,20,0),"recorded_by_provider_id":PROV_HASAN},
    {"id": uid("obs", "david-a1c"),      "patient_id": P_DAVID,    "encounter_id": ENC_DAVID_NEWPT,    "loinc_code": "4548-4", "value_numeric": 9.2,    "value_text": None,    "units": "%",        "interpretation": "H", "observed_at": ts(2024,12,19,13,40,0),"recorded_by_provider_id": PROV_RAJESH},
]

CARE_PLANS = [
    {"id": uid("cp", "margaret-dm-htn"),"patient_id": P_MARGARET,"title": "Diabetes & Hypertension Co-Management","category": "chronic_disease","status": "active",    "start_date": "2023-06-15","end_date": None,         "responsible_provider_id": PROV_MORGAN, "description": "Goal A1c < 7.0%, BP < 130/80. Q3 month follow-up. Annual eye/foot exam. Metformin/glipizide + lisinopril/HCTZ + atorvastatin."},
    {"id": uid("cp", "robert-cad"),     "patient_id": P_ROBERT,  "title": "Post-NSTEMI / CAD Secondary Prevention","category": "cardiovascular","status": "active",    "start_date": "2024-11-07","end_date": None,         "responsible_provider_id": PROV_LISA,   "description": "DAPT x 12 months, high-intensity statin, beta-blocker, ACE-I, cardiac rehab 36 sessions, lifestyle counseling."},
    {"id": uid("cp", "lakshmi-asthma"), "patient_id": P_LAKSHMI, "title": "Asthma Action Plan",                    "category": "respiratory",    "status": "active",    "start_date": "2024-01-15","end_date": None,         "responsible_provider_id": PROV_MORGAN, "description": "Green zone: fluticasone 220mcg 2 puffs BID. Yellow: add albuterol q4h x 48h. Red: prednisone 40mg + call PCP/ED. Peak flow log."},
    {"id": uid("cp", "james-copd-chf"), "patient_id": P_JAMES,   "title": "COPD/CHF Bundle",                       "category": "respiratory",    "status": "active",    "start_date": "2024-01-04","end_date": None,         "responsible_provider_id": PROV_AMANDA, "description": "Tiotropium + ICS/LABA, home O2 2L PRN sat<88%, daily weights, fluid restriction 2L, pulmonary rehab, annual flu/pneumococcal."},
    {"id": uid("cp", "aisha-pregnancy"),"patient_id": P_AISHA,   "title": "Prenatal Care Plan",                    "category": "ob",             "status": "active",    "start_date": "2024-09-01","end_date": "2025-04-30", "responsible_provider_id": PROV_SARA,   "description": "Q4 wk visits to 28wk, q2wk to 36, weekly thereafter. GBS at 36wk. Tdap administered. Birth plan filed."},
    {"id": uid("cp", "dorothy-thr"),    "patient_id": P_DOROTHY, "title": "Post-THA Recovery",                     "category": "post_surgical",  "status": "active",    "start_date": "2024-11-22","end_date": "2025-02-22", "responsible_provider_id": PROV_DANIEL, "description": "6 weeks PT, hip precautions (no flex >90, no adduction past midline), DVT prophylaxis 4wk, follow-up 2/6/12 weeks."},
    {"id": uid("cp", "carlos-gad"),     "patient_id": P_CARLOS,  "title": "Anxiety Management",                    "category": "behavioral_health","status": "active",  "start_date": "2024-09-10","end_date": None,         "responsible_provider_id": PROV_RAJESH, "description": "Sertraline titration to 100mg, CBT referral, daily relaxation practice, monitor GAD-7 q4wk."},
    {"id": uid("cp", "david-newdm"),    "patient_id": P_DAVID,   "title": "Newly Diagnosed T2DM",                  "category": "chronic_disease","status": "active",    "start_date": "2024-12-19","end_date": None,         "responsible_provider_id": PROV_RAJESH, "description": "Start metformin 500mg BID titrate to 1000mg BID. DM education class. SMBG fasting + 2hr post. Refer endocrinology. Repeat A1c 3 months."},
]

CARE_PLAN_GOALS = [
    {"id": uid("cpg", "margaret-a1c"),"care_plan_id": uid("cp", "margaret-dm-htn"),"description": "Achieve A1c < 7.0% within 6 months",          "target_date": "2025-07-15","achievement_status": "in_progress","priority": "high"},
    {"id": uid("cpg", "margaret-bp"), "care_plan_id": uid("cp", "margaret-dm-htn"),"description": "Maintain BP < 130/80 at 80% of readings",     "target_date": "2025-04-15","achievement_status": "in_progress","priority": "high"},
    {"id": uid("cpg", "margaret-eye"),"care_plan_id": uid("cp", "margaret-dm-htn"),"description": "Annual dilated eye exam",                      "target_date": "2025-06-15","achievement_status": "not_started","priority": "medium"},
    {"id": uid("cpg", "robert-dapt"), "care_plan_id": uid("cp", "robert-cad"),     "description": "Complete 12 months DAPT without bleeding event","target_date": "2025-11-04","achievement_status": "in_progress","priority": "high"},
    {"id": uid("cpg", "robert-ldl"),  "care_plan_id": uid("cp", "robert-cad"),     "description": "LDL < 70 mg/dL on high-intensity statin",      "target_date": "2025-03-01","achievement_status": "achieved",   "priority": "high"},
    {"id": uid("cpg", "robert-rehab"),"care_plan_id": uid("cp", "robert-cad"),     "description": "Complete 36 cardiac rehab sessions",           "target_date": "2025-05-01","achievement_status": "in_progress","priority": "medium"},
    {"id": uid("cpg", "lakshmi-pf"),  "care_plan_id": uid("cp", "lakshmi-asthma"), "description": "Peak flow stays in green zone (>320)",         "target_date": "2025-12-31","achievement_status": "achieved",   "priority": "high"},
    {"id": uid("cpg", "james-noadm"), "care_plan_id": uid("cp", "james-copd-chf"), "description": "Avoid further hospitalization x 6 months",     "target_date": "2025-07-04","achievement_status": "in_progress","priority": "high"},
    {"id": uid("cpg", "james-rehab"), "care_plan_id": uid("cp", "james-copd-chf"), "description": "Attend ≥80% pulmonary rehab sessions",         "target_date": "2025-04-01","achievement_status": "in_progress","priority": "medium"},
    {"id": uid("cpg", "aisha-deliv"), "care_plan_id": uid("cp", "aisha-pregnancy"),"description": "Term vaginal delivery without complications",  "target_date": "2025-04-18","achievement_status": "in_progress","priority": "high"},
    {"id": uid("cpg", "dorothy-amb"), "care_plan_id": uid("cp", "dorothy-thr"),    "description": "Ambulate 100ft with walker independently",     "target_date": "2025-01-22","achievement_status": "achieved",   "priority": "high"},
    {"id": uid("cpg", "carlos-gad7"), "care_plan_id": uid("cp", "carlos-gad"),     "description": "GAD-7 score below 10 sustained x 2 visits",    "target_date": "2025-04-09","achievement_status": "in_progress","priority": "medium"},
    {"id": uid("cpg", "david-a1c"),   "care_plan_id": uid("cp", "david-newdm"),    "description": "A1c reduction to < 7.5% by 3 months",          "target_date": "2025-03-19","achievement_status": "in_progress","priority": "high"},
]

# ---------------------------------------------------------------------------
# 8. MEDICATIONS — formulary, prescriptions, administrations, reconciliations
# ---------------------------------------------------------------------------

MED_METFORMIN_1000 = uid("med", "metformin-1000-tab")
MED_LISIN_10       = uid("med", "lisinopril-10-tab")
MED_LISIN_20       = uid("med", "lisinopril-20-tab")
MED_ATORV_40       = uid("med", "atorvastatin-40-tab")
MED_ATORV_80       = uid("med", "atorvastatin-80-tab")
MED_ALBUT_NEB      = uid("med", "albuterol-neb")
MED_ALBUT_MDI      = uid("med", "albuterol-mdi")
MED_TIOTROPIUM     = uid("med", "tiotropium-handihaler")
MED_OMEPRAZOLE     = uid("med", "omeprazole-20-cap")
MED_LEVOTHYRX_50   = uid("med", "levothyroxine-50-tab")
MED_ASA_81         = uid("med", "aspirin-81-tab")
MED_CLOPIDOGREL_75 = uid("med", "clopidogrel-75-tab")
MED_METOP_50       = uid("med", "metoprolol-50-tab")
MED_FUROSEMIDE_40  = uid("med", "furosemide-40-tab")
MED_HCTZ_25        = uid("med", "hctz-25-tab")
MED_AMOX_500       = uid("med", "amoxicillin-500-cap")
MED_APAP_325       = uid("med", "acetaminophen-325-tab")
MED_IBUPROFEN_600  = uid("med", "ibuprofen-600-tab")
MED_MORPHINE_4     = uid("med", "morphine-iv-4")
MED_OXYCODONE_5    = uid("med", "oxycodone-5-tab")
MED_INSULIN_GLAR   = uid("med", "insulin-glargine-100")
MED_SERTRALINE_50  = uid("med", "sertraline-50-tab")

MEDICATIONS = [
    {"id": MED_METFORMIN_1000,"rxcui": "860975","ndc": "00093-1048-01","name": "Metformin HCl 1000 mg tablet",                       "strength": "1000 mg","dosage_form": "tablet",  "route": "oral",      "manufacturer": "Teva",            "is_on_formulary": True},
    {"id": MED_LISIN_10,      "rxcui": "314076","ndc": "00378-1810-01","name": "Lisinopril 10 mg tablet",                            "strength": "10 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Mylan",           "is_on_formulary": True},
    {"id": MED_LISIN_20,      "rxcui": "314077","ndc": "00378-1820-01","name": "Lisinopril 20 mg tablet",                            "strength": "20 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Mylan",           "is_on_formulary": True},
    {"id": MED_ATORV_40,      "rxcui": "617314","ndc": "00071-0157-23","name": "Atorvastatin calcium 40 mg tablet",                  "strength": "40 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Pfizer",          "is_on_formulary": True},
    {"id": MED_ATORV_80,      "rxcui": "617318","ndc": "00071-0158-23","name": "Atorvastatin calcium 80 mg tablet",                  "strength": "80 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Pfizer",          "is_on_formulary": True},
    {"id": MED_ALBUT_NEB,     "rxcui": "308136","ndc": "00487-9501-25","name": "Albuterol sulfate 0.083% inhalation solution",       "strength": "2.5 mg/3 mL","dosage_form": "solution","route": "inhalation","manufacturer": "Nephron",      "is_on_formulary": True},
    {"id": MED_ALBUT_MDI,     "rxcui": "745679","ndc": "59310-0579-22","name": "Albuterol 90 mcg/actuation MDI (ProAir HFA)",        "strength": "90 mcg", "dosage_form": "inhaler", "route": "inhalation","manufacturer": "Teva",            "is_on_formulary": True},
    {"id": MED_TIOTROPIUM,    "rxcui": "856987","ndc": "00597-0075-41","name": "Tiotropium bromide 18 mcg Handihaler",               "strength": "18 mcg", "dosage_form": "inhaler", "route": "inhalation","manufacturer": "Boehringer",      "is_on_formulary": True},
    {"id": MED_OMEPRAZOLE,    "rxcui": "313988","ndc": "00378-7090-01","name": "Omeprazole 20 mg capsule DR",                        "strength": "20 mg",  "dosage_form": "capsule", "route": "oral",      "manufacturer": "Mylan",           "is_on_formulary": True},
    {"id": MED_LEVOTHYRX_50,  "rxcui": "197604","ndc": "00074-7068-19","name": "Levothyroxine sodium 50 mcg tablet",                 "strength": "50 mcg", "dosage_form": "tablet",  "route": "oral",      "manufacturer": "AbbVie (Synthroid)","is_on_formulary": True},
    {"id": MED_ASA_81,        "rxcui": "855332","ndc": "63824-0103-71","name": "Aspirin 81 mg chewable tablet",                      "strength": "81 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Bayer",           "is_on_formulary": True},
    {"id": MED_CLOPIDOGREL_75,"rxcui": "309362","ndc": "63629-1126-01","name": "Clopidogrel bisulfate 75 mg tablet",                 "strength": "75 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Bristol-Myers Squibb","is_on_formulary": True},
    {"id": MED_METOP_50,      "rxcui": "866412","ndc": "00378-0014-01","name": "Metoprolol tartrate 50 mg tablet",                   "strength": "50 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Mylan",           "is_on_formulary": True},
    {"id": MED_FUROSEMIDE_40, "rxcui": "310429","ndc": "00781-1818-01","name": "Furosemide 40 mg tablet",                            "strength": "40 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Sandoz",          "is_on_formulary": True},
    {"id": MED_HCTZ_25,       "rxcui": "313850","ndc": "00603-4225-21","name": "Hydrochlorothiazide 25 mg tablet",                   "strength": "25 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Qualitest",       "is_on_formulary": True},
    {"id": MED_AMOX_500,      "rxcui": "308191","ndc": "65862-0017-05","name": "Amoxicillin 500 mg capsule",                         "strength": "500 mg", "dosage_form": "capsule", "route": "oral",      "manufacturer": "Aurobindo",       "is_on_formulary": True},
    {"id": MED_APAP_325,      "rxcui": "313782","ndc": "70000-0017-01","name": "Acetaminophen 325 mg tablet",                        "strength": "325 mg", "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Cardinal Health", "is_on_formulary": True},
    {"id": MED_IBUPROFEN_600, "rxcui": "310965","ndc": "00904-7929-80","name": "Ibuprofen 600 mg tablet",                            "strength": "600 mg", "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Major",           "is_on_formulary": True},
    {"id": MED_MORPHINE_4,    "rxcui": "892494","ndc": "00641-6072-25","name": "Morphine sulfate 4 mg/mL injection",                 "strength": "4 mg/mL","dosage_form": "solution","route": "iv",       "manufacturer": "Hospira",         "is_on_formulary": True},
    {"id": MED_OXYCODONE_5,   "rxcui": "1014676","ndc":"00603-4988-21","name": "Oxycodone HCl 5 mg tablet",                          "strength": "5 mg",   "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Qualitest",       "is_on_formulary": True},
    {"id": MED_INSULIN_GLAR,  "rxcui": "966177","ndc": "00088-2220-33","name": "Insulin glargine 100 units/mL (Lantus SoloStar)",    "strength": "100 unit/mL","dosage_form": "pen", "route": "subq",      "manufacturer": "Sanofi",          "is_on_formulary": True},
    {"id": MED_SERTRALINE_50, "rxcui": "847232","ndc": "00378-3375-01","name": "Sertraline 50 mg tablet",                            "strength": "50 mg",  "dosage_form": "tablet",  "route": "oral",      "manufacturer": "Greenstone",      "is_on_formulary": True},
]

RX_MARGARET_METFORMIN = uid("rx", "margaret-metformin")
RX_MARGARET_LISIN    = uid("rx", "margaret-lisinopril")
RX_MARGARET_ATORV    = uid("rx", "margaret-atorvastatin")
RX_MARGARET_LEVO     = uid("rx", "margaret-levothyroxine")
RX_ROBERT_ASA        = uid("rx", "robert-aspirin")
RX_ROBERT_CLOPI      = uid("rx", "robert-clopidogrel")
RX_ROBERT_ATORV      = uid("rx", "robert-atorvastatin-80")
RX_ROBERT_METOP      = uid("rx", "robert-metoprolol")
RX_ROBERT_LISIN      = uid("rx", "robert-lisinopril-10")
RX_LAKSHMI_ALBUT     = uid("rx", "lakshmi-albuterol-mdi")
RX_LAKSHMI_OMEP      = uid("rx", "lakshmi-omeprazole")
RX_JAMES_TIO         = uid("rx", "james-tiotropium")
RX_JAMES_FURO        = uid("rx", "james-furosemide")
RX_JAMES_ALBUT       = uid("rx", "james-albuterol-mdi")
RX_JAMES_AMOX        = uid("rx", "james-amoxicillin-inpt")
RX_JAMES_MORPHINE    = uid("rx", "james-morphine-inpt")
RX_MARCUS_IBU        = uid("rx", "marcus-ibuprofen")
RX_MARCUS_APAP       = uid("rx", "marcus-apap")
RX_DOROTHY_OXY       = uid("rx", "dorothy-oxycodone-postop")
RX_DOROTHY_APAP      = uid("rx", "dorothy-apap-postop")
RX_CARLOS_SERT       = uid("rx", "carlos-sertraline")
RX_EMILY_ALBUT       = uid("rx", "emily-albuterol-mdi")
RX_DAVID_METFORMIN   = uid("rx", "david-metformin")

PRESCRIPTIONS = [
    {"id": RX_MARGARET_METFORMIN,"patient_id": P_MARGARET,"prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_MARGARET_FU, "medication_id": MED_METFORMIN_1000,"dose": "1000 mg","route": "oral","frequency": "BID with meals", "duration_days": None, "quantity": 180, "refills": 3, "start_date": "2023-06-15", "end_date": None,        "status": "active",   "indication": "Type 2 diabetes mellitus", "prn_reason": None,                          "pharmacy_id": PHARM_CVS,    "is_electronic": True},
    {"id": RX_MARGARET_LISIN,    "patient_id": P_MARGARET,"prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_MARGARET_FU, "medication_id": MED_LISIN_20,      "dose": "20 mg",  "route": "oral","frequency": "daily",          "duration_days": None, "quantity": 90,  "refills": 3, "start_date": "2024-08-01", "end_date": None,        "status": "active",   "indication": "Hypertension + CKD",       "prn_reason": None,                          "pharmacy_id": PHARM_CVS,    "is_electronic": True},
    {"id": RX_MARGARET_ATORV,    "patient_id": P_MARGARET,"prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_MARGARET_FU, "medication_id": MED_ATORV_40,      "dose": "40 mg",  "route": "oral","frequency": "at bedtime",     "duration_days": None, "quantity": 90,  "refills": 3, "start_date": "2022-03-12", "end_date": None,        "status": "active",   "indication": "Hyperlipidemia",           "prn_reason": None,                          "pharmacy_id": PHARM_CVS,    "is_electronic": True},
    {"id": RX_MARGARET_LEVO,     "patient_id": P_MARGARET,"prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_MARGARET_FU, "medication_id": MED_LEVOTHYRX_50,  "dose": "50 mcg", "route": "oral","frequency": "daily AM fasting","duration_days":None,  "quantity": 90,  "refills": 3, "start_date": "2020-11-20", "end_date": None,        "status": "active",   "indication": "Hypothyroidism",           "prn_reason": None,                          "pharmacy_id": PHARM_CVS,    "is_electronic": True},
    {"id": RX_ROBERT_ASA,        "patient_id": P_ROBERT,  "prescriber_provider_id": PROV_LISA,  "encounter_id": ENC_ROBERT_INPT, "medication_id": MED_ASA_81,        "dose": "81 mg",  "route": "oral","frequency": "daily",          "duration_days": None, "quantity": 90,  "refills": 5, "start_date": "2024-11-07", "end_date": None,        "status": "active",   "indication": "Post-NSTEMI",              "prn_reason": None,                          "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_ROBERT_CLOPI,      "patient_id": P_ROBERT,  "prescriber_provider_id": PROV_LISA,  "encounter_id": ENC_ROBERT_INPT, "medication_id": MED_CLOPIDOGREL_75,"dose": "75 mg",  "route": "oral","frequency": "daily",          "duration_days": 365,  "quantity": 90,  "refills": 3, "start_date": "2024-11-07", "end_date": "2025-11-07","status": "active",   "indication": "DAPT post-DES",            "prn_reason": None,                          "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_ROBERT_ATORV,      "patient_id": P_ROBERT,  "prescriber_provider_id": PROV_LISA,  "encounter_id": ENC_ROBERT_INPT, "medication_id": MED_ATORV_80,      "dose": "80 mg",  "route": "oral","frequency": "at bedtime",     "duration_days": None, "quantity": 90,  "refills": 3, "start_date": "2024-11-07", "end_date": None,        "status": "active",   "indication": "High-intensity statin post-MI","prn_reason": None,                       "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_ROBERT_METOP,      "patient_id": P_ROBERT,  "prescriber_provider_id": PROV_LISA,  "encounter_id": ENC_ROBERT_INPT, "medication_id": MED_METOP_50,      "dose": "50 mg",  "route": "oral","frequency": "BID",            "duration_days": None, "quantity": 180, "refills": 3, "start_date": "2024-11-07", "end_date": None,        "status": "active",   "indication": "Beta-blocker post-MI",     "prn_reason": None,                          "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_ROBERT_LISIN,      "patient_id": P_ROBERT,  "prescriber_provider_id": PROV_LISA,  "encounter_id": ENC_ROBERT_INPT, "medication_id": MED_LISIN_10,      "dose": "10 mg",  "route": "oral","frequency": "daily",          "duration_days": None, "quantity": 90,  "refills": 3, "start_date": "2024-11-07", "end_date": None,        "status": "active",   "indication": "ACE-I post-MI",            "prn_reason": None,                          "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_LAKSHMI_ALBUT,     "patient_id": P_LAKSHMI, "prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_LAKSHMI_FU,  "medication_id": MED_ALBUT_MDI,     "dose": "2 puffs","route": "inhalation","frequency": "PRN q4h",   "duration_days": None, "quantity": 1,   "refills": 5, "start_date": "2024-01-10", "end_date": None,        "status": "active",   "indication": "Asthma rescue",            "prn_reason": "Wheezing or SOB",             "pharmacy_id": PHARM_CVS,    "is_electronic": True},
    {"id": RX_LAKSHMI_OMEP,      "patient_id": P_LAKSHMI, "prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_LAKSHMI_FU,  "medication_id": MED_OMEPRAZOLE,    "dose": "20 mg",  "route": "oral","frequency": "daily AM",       "duration_days": None, "quantity": 90,  "refills": 3, "start_date": "2017-02-20", "end_date": None,        "status": "active",   "indication": "GERD",                     "prn_reason": None,                          "pharmacy_id": PHARM_CVS,    "is_electronic": True},
    {"id": RX_JAMES_TIO,         "patient_id": P_JAMES,   "prescriber_provider_id": PROV_AMANDA,"encounter_id": ENC_JAMES_INPT,  "medication_id": MED_TIOTROPIUM,    "dose": "18 mcg", "route": "inhalation","frequency": "1 cap daily","duration_days":None,"quantity":1,    "refills": 5, "start_date": "2024-09-04", "end_date": None,        "status": "active",   "indication": "COPD maintenance",         "prn_reason": None,                          "pharmacy_id": PHARM_MAIL,   "is_electronic": True},
    {"id": RX_JAMES_FURO,        "patient_id": P_JAMES,   "prescriber_provider_id": PROV_AMANDA,"encounter_id": ENC_JAMES_INPT,  "medication_id": MED_FUROSEMIDE_40, "dose": "40 mg",  "route": "oral","frequency": "daily AM",       "duration_days": None, "quantity": 90,  "refills": 3, "start_date": "2024-01-04", "end_date": None,        "status": "active",   "indication": "CHF volume management",    "prn_reason": None,                          "pharmacy_id": PHARM_MAIL,   "is_electronic": True},
    {"id": RX_JAMES_ALBUT,       "patient_id": P_JAMES,   "prescriber_provider_id": PROV_AMANDA,"encounter_id": ENC_JAMES_INPT,  "medication_id": MED_ALBUT_MDI,     "dose": "2 puffs","route": "inhalation","frequency": "PRN q4h",   "duration_days": None, "quantity": 1,   "refills": 5, "start_date": "2024-01-04", "end_date": None,        "status": "active",   "indication": "COPD rescue",              "prn_reason": "Dyspnea",                     "pharmacy_id": PHARM_MAIL,   "is_electronic": True},
    {"id": RX_JAMES_AMOX,        "patient_id": P_JAMES,   "prescriber_provider_id": PROV_AMANDA,"encounter_id": ENC_JAMES_INPT,  "medication_id": MED_AMOX_500,      "dose": "500 mg", "route": "oral","frequency": "TID",            "duration_days": 7,    "quantity": 21,  "refills": 0, "start_date": "2025-01-04", "end_date": "2025-01-11","status": "completed","indication": "CAP / COPD exacerbation",  "prn_reason": None,                          "pharmacy_id": PHARM_INHOUSE,"is_electronic": True},
    {"id": RX_JAMES_MORPHINE,    "patient_id": P_JAMES,   "prescriber_provider_id": PROV_AMANDA,"encounter_id": ENC_JAMES_INPT,  "medication_id": MED_MORPHINE_4,    "dose": "2 mg",   "route": "iv",  "frequency": "PRN q4h",        "duration_days": 3,    "quantity": None,"refills": 0, "start_date": "2025-01-04", "end_date": "2025-01-07","status": "completed","indication": "Severe dyspnea palliation","prn_reason": "Refractory air hunger",       "pharmacy_id": PHARM_INHOUSE,"is_electronic": True},
    {"id": RX_MARCUS_IBU,        "patient_id": P_MARCUS,  "prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_MARCUS_LBP,  "medication_id": MED_IBUPROFEN_600, "dose": "600 mg", "route": "oral","frequency": "TID with food",  "duration_days": 7,    "quantity": 21,  "refills": 0, "start_date": "2025-01-08", "end_date": "2025-01-15","status": "completed","indication": "LBP flare",                "prn_reason": None,                          "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_MARCUS_APAP,       "patient_id": P_MARCUS,  "prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_MARCUS_LBP,  "medication_id": MED_APAP_325,      "dose": "650 mg", "route": "oral","frequency": "PRN q6h",        "duration_days": None, "quantity": 100, "refills": 1, "start_date": "2025-01-08", "end_date": None,        "status": "active",   "indication": "Pain",                     "prn_reason": "LBP",                         "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_DOROTHY_OXY,       "patient_id": P_DOROTHY, "prescriber_provider_id": PROV_DANIEL,"encounter_id": ENC_DOROTHY_INPT,"medication_id": MED_OXYCODONE_5,   "dose": "5 mg",   "route": "oral","frequency": "PRN q6h",        "duration_days": 7,    "quantity": 28,  "refills": 0, "start_date": "2024-11-22", "end_date": "2024-11-29","status": "completed","indication": "Post-op pain",             "prn_reason": "Severe pain after PT",        "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_DOROTHY_APAP,      "patient_id": P_DOROTHY, "prescriber_provider_id": PROV_DANIEL,"encounter_id": ENC_DOROTHY_INPT,"medication_id": MED_APAP_325,      "dose": "650 mg", "route": "oral","frequency": "scheduled q6h",  "duration_days": 14,   "quantity": 56,  "refills": 0, "start_date": "2024-11-22", "end_date": "2024-12-06","status": "completed","indication": "Post-op pain baseline",    "prn_reason": None,                          "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
    {"id": RX_CARLOS_SERT,       "patient_id": P_CARLOS,  "prescriber_provider_id": PROV_RAJESH,"encounter_id": ENC_CARLOS_TELE, "medication_id": MED_SERTRALINE_50, "dose": "100 mg", "route": "oral","frequency": "daily AM",       "duration_days": None, "quantity": 90,  "refills": 3, "start_date": "2025-01-09", "end_date": None,        "status": "active",   "indication": "Generalized anxiety",      "prn_reason": None,                          "pharmacy_id": PHARM_CVS,    "is_electronic": True},
    {"id": RX_EMILY_ALBUT,       "patient_id": P_EMILY,   "prescriber_provider_id": PROV_MORGAN,"encounter_id": ENC_EMILY_URGENT,"medication_id": MED_ALBUT_MDI,     "dose": "2 puffs","route": "inhalation","frequency": "PRN q4h",   "duration_days": None, "quantity": 1,   "refills": 5, "start_date": "2024-10-22", "end_date": None,        "status": "active",   "indication": "Asthma rescue",            "prn_reason": "Wheezing/SOB",                "pharmacy_id": PHARM_CVS,    "is_electronic": True},
    {"id": RX_DAVID_METFORMIN,   "patient_id": P_DAVID,   "prescriber_provider_id": PROV_RAJESH,"encounter_id": ENC_DAVID_NEWPT, "medication_id": MED_METFORMIN_1000,"dose": "500 mg", "route": "oral","frequency": "BID titrate to 1000 mg BID","duration_days":None,"quantity":60,"refills": 5,"start_date":"2024-12-19","end_date":None,"status":"active","indication": "New T2DM",                 "prn_reason": None,                          "pharmacy_id": PHARM_WALGRN, "is_electronic": True},
]

MEDICATION_ADMINISTRATIONS = [
    # Inpatient admins for Robert, James, Dorothy
    {"id": uid("medadmin", "robert-asa-d1"),     "prescription_id": RX_ROBERT_ASA,     "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_INPT, "administered_at": ts(2024,11,5,8,0,0),   "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "81 mg",  "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "robert-asa-d2"),     "prescription_id": RX_ROBERT_ASA,     "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_INPT, "administered_at": ts(2024,11,6,8,0,0),   "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "81 mg",  "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "robert-clopi-d1"),   "prescription_id": RX_ROBERT_CLOPI,   "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_INPT, "administered_at": ts(2024,11,5,8,0,0),   "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "75 mg",  "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "robert-atorv-d1"),   "prescription_id": RX_ROBERT_ATORV,   "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_INPT, "administered_at": ts(2024,11,4,22,0,0),  "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "80 mg",  "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "robert-metop-d1-am"),"prescription_id": RX_ROBERT_METOP,   "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_INPT, "administered_at": ts(2024,11,5,8,0,0),   "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "50 mg",  "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "robert-metop-d1-pm"),"prescription_id": RX_ROBERT_METOP,   "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_INPT, "administered_at": ts(2024,11,5,20,0,0),  "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "50 mg",  "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "james-tio-d1"),      "prescription_id": RX_JAMES_TIO,      "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "administered_at": ts(2025,1,5,8,0,0),    "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "18 mcg", "route": "inhalation","status": "completed","not_done_reason": None,"site": None},
    {"id": uid("medadmin", "james-furo-d1"),     "prescription_id": RX_JAMES_FURO,     "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "administered_at": ts(2025,1,5,8,0,0),    "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "40 mg",  "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "james-amox-d1-am"),  "prescription_id": RX_JAMES_AMOX,     "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "administered_at": ts(2025,1,4,12,0,0),   "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "500 mg", "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "james-amox-d1-pm"),  "prescription_id": RX_JAMES_AMOX,     "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "administered_at": ts(2025,1,4,18,0,0),   "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "500 mg", "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "james-amox-d1-night"),"prescription_id": RX_JAMES_AMOX,    "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "administered_at": ts(2025,1,5,0,0,0),    "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "500 mg", "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "james-morphine-prn1"),"prescription_id": RX_JAMES_MORPHINE,"patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "administered_at": ts(2025,1,4,9,30,0),   "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "2 mg",   "route": "iv",   "status": "completed", "not_done_reason": None, "site": "Left antecubital"},
    {"id": uid("medadmin", "james-morphine-prn2"),"prescription_id": RX_JAMES_MORPHINE,"patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "administered_at": ts(2025,1,4,15,45,0),  "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "2 mg",   "route": "iv",   "status": "completed", "not_done_reason": None, "site": "Left antecubital"},
    {"id": uid("medadmin", "dorothy-oxy-postop"),"prescription_id": RX_DOROTHY_OXY,    "patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_INPT,"administered_at": ts(2024,11,19,14,0,0), "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "5 mg",   "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "dorothy-oxy-night1"),"prescription_id": RX_DOROTHY_OXY,    "patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_INPT,"administered_at": ts(2024,11,19,22,0,0), "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "5 mg",   "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "dorothy-apap-d1"),   "prescription_id": RX_DOROTHY_APAP,   "patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_INPT,"administered_at": ts(2024,11,19,6,0,0),  "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "650 mg", "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "dorothy-apap-d1-12"),"prescription_id": RX_DOROTHY_APAP,   "patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_INPT,"administered_at": ts(2024,11,19,12,0,0), "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "650 mg", "route": "oral", "status": "completed", "not_done_reason": None, "site": None},
    {"id": uid("medadmin", "dorothy-apap-d2-18"),"prescription_id": RX_DOROTHY_APAP,   "patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_INPT,"administered_at": ts(2024,11,19,18,0,0), "administered_by_provider_id": PROV_MAYA_RN, "dose_given": "650 mg", "route": "oral", "status": "held",       "not_done_reason": "patient_refused", "site": None},
]

MEDICATION_RECONCILIATIONS = [
    {"id": uid("medrec", "robert-admit"),   "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_INPT, "reconciliation_type": "admission","performed_by_provider_id": PROV_PRIYA, "performed_at": ts(2024,11,4,20,15,0), "notes": "Pre-admission: lisinopril 10, atorvastatin 20. Continued lisinopril, increased atorvastatin to 80, added ASA 81 + clopidogrel 75 + metoprolol 50 BID."},
    {"id": uid("medrec", "robert-dc"),      "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_INPT, "reconciliation_type": "discharge","performed_by_provider_id": PROV_PRIYA, "performed_at": ts(2024,11,7,11,0,0),  "notes": "Discharge meds: ASA 81 daily, clopidogrel 75 daily x 12 months, atorvastatin 80 qHS, metoprolol 50 BID, lisinopril 10 daily. Patient verbalized understanding."},
    {"id": uid("medrec", "james-admit"),    "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "reconciliation_type": "admission","performed_by_provider_id": PROV_PRIYA, "performed_at": ts(2025,1,4,6,15,0),   "notes": "Pre-admission: tiotropium, ICS/LABA, furosemide 40, ASA 81, lisinopril 5. Held lisinopril for AKI Cr 1.8, added amoxicillin 500 TID x 7d for CAP, started solumedrol IV 60 q6."},
    {"id": uid("medrec", "james-dc"),       "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,  "reconciliation_type": "discharge","performed_by_provider_id": PROV_AMANDA,"performed_at": ts(2025,1,9,13,30,0),  "notes": "Discharged on home meds + prednisone 40 mg taper over 10d + amoxicillin x 3 more days. Restart lisinopril when Cr<1.5."},
    {"id": uid("medrec", "dorothy-admit"),  "patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_INPT,"reconciliation_type": "admission","performed_by_provider_id": PROV_DANIEL, "performed_at": ts(2024,11,18,17,0,0), "notes": "Home: donepezil 10, memantine 10 BID, sertraline 50, calcium/D. NPO for surgery. Held all PO meds, will resume post-op."},
    {"id": uid("medrec", "dorothy-dc"),     "patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_INPT,"reconciliation_type": "discharge","performed_by_provider_id": PROV_DANIEL, "performed_at": ts(2024,11,22,11,30,0),"notes": "SNF transfer with oxycodone 5 PRN, scheduled APAP 650 q6, restart home meds. Enoxaparin 40 daily x 28d for DVT ppx."},
    {"id": uid("medrec", "margaret-fu"),    "patient_id": P_MARGARET,"encounter_id": ENC_MARGARET_FU, "reconciliation_type": "ambulatory","performed_by_provider_id": PROV_MORGAN,"performed_at": ts(2025,1,13,9,25,0),  "notes": "Reviewed home meds; all continued. Discussed glipizide hold if hypoglycemic episodes recur."},
    {"id": uid("medrec", "david-newpt"),    "patient_id": P_DAVID,   "encounter_id": ENC_DAVID_NEWPT, "reconciliation_type": "ambulatory","performed_by_provider_id": PROV_RAJESH,"performed_at": ts(2024,12,19,13,55,0),"notes": "No prior chronic meds. Initiated metformin 500 BID, titrate over 2 weeks to 1000 BID. Counseled GI side effects."},
]

# ---------------------------------------------------------------------------
# 9. LABS & IMAGING
# ---------------------------------------------------------------------------

LO_MARGARET_CMP   = uid("laborder", "margaret-cmp-2025-01-13")
LO_MARGARET_CBC   = uid("laborder", "margaret-cbc-2025-01-13")
LO_MARGARET_A1C   = uid("laborder", "margaret-a1c-2025-01-13")
LO_ROBERT_TROP    = uid("laborder", "robert-trop-2024-11-04")
LO_ROBERT_CMP_ED  = uid("laborder", "robert-cmp-2024-11-04")
LO_ROBERT_LIPID   = uid("laborder", "robert-lipid-2024-12-12")
LO_JAMES_CBC_ED   = uid("laborder", "james-cbc-2025-01-04")
LO_JAMES_BMP_ED   = uid("laborder", "james-bmp-2025-01-04")
LO_DAVID_A1C      = uid("laborder", "david-a1c-2024-12-19")
LO_DAVID_LIPID    = uid("laborder", "david-lipid-2024-12-19")
LO_AISHA_CBC      = uid("laborder", "aisha-cbc-2025-01-16")
LO_CARLOS_TROP    = uid("laborder", "carlos-trop-2024-12-28")

LAB_ORDERS = [
    {"id": LO_MARGARET_CMP,  "patient_id": P_MARGARET, "encounter_id": ENC_MARGARET_FU,   "ordering_provider_id": PROV_MORGAN,  "ordered_at": ts(2025,1,13,9,30,0),  "priority": "routine", "status": "completed", "panel_loinc_code": "24323-8", "clinical_question": "DM/HTN/CKD follow-up - eGFR, K+",                "fasting_required": True},
    {"id": LO_MARGARET_CBC,  "patient_id": P_MARGARET, "encounter_id": ENC_MARGARET_FU,   "ordering_provider_id": PROV_MORGAN,  "ordered_at": ts(2025,1,13,9,30,0),  "priority": "routine", "status": "completed", "panel_loinc_code": "57021-8", "clinical_question": "Annual screening CBC",                          "fasting_required": False},
    {"id": LO_MARGARET_A1C,  "patient_id": P_MARGARET, "encounter_id": ENC_MARGARET_FU,   "ordering_provider_id": PROV_MORGAN,  "ordered_at": ts(2025,1,13,9,30,0),  "priority": "routine", "status": "completed", "panel_loinc_code": "4548-4",  "clinical_question": "DM control - target <7.0%",                     "fasting_required": False},
    {"id": LO_ROBERT_TROP,   "patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_ED,     "ordering_provider_id": PROV_HASAN,   "ordered_at": ts(2024,11,4,16,30,0), "priority": "stat",    "status": "completed", "panel_loinc_code": "10839-9", "clinical_question": "R/O ACS - serial troponins",                    "fasting_required": False},
    {"id": LO_ROBERT_CMP_ED, "patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_ED,     "ordering_provider_id": PROV_HASAN,   "ordered_at": ts(2024,11,4,16,30,0), "priority": "stat",    "status": "completed", "panel_loinc_code": "24323-8", "clinical_question": "Baseline CMP for chest pain workup",            "fasting_required": False},
    {"id": LO_ROBERT_LIPID,  "patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_CARDIO, "ordering_provider_id": PROV_LISA,    "ordered_at": ts(2024,12,12,9,0,0),  "priority": "routine", "status": "completed", "panel_loinc_code": None,      "clinical_question": "On-statin lipid panel - confirm goal LDL<70",   "fasting_required": True},
    {"id": LO_JAMES_CBC_ED,  "patient_id": P_JAMES,    "encounter_id": ENC_JAMES_ED,      "ordering_provider_id": PROV_JOHANNA, "ordered_at": ts(2025,1,4,3,0,0),    "priority": "stat",    "status": "completed", "panel_loinc_code": "57021-8", "clinical_question": "Sepsis workup - CBC with diff",                 "fasting_required": False},
    {"id": LO_JAMES_BMP_ED,  "patient_id": P_JAMES,    "encounter_id": ENC_JAMES_ED,      "ordering_provider_id": PROV_JOHANNA, "ordered_at": ts(2025,1,4,3,0,0),    "priority": "stat",    "status": "completed", "panel_loinc_code": "24323-8", "clinical_question": "Electrolyte/renal function in CHF/COPD exac",    "fasting_required": False},
    {"id": LO_DAVID_A1C,     "patient_id": P_DAVID,    "encounter_id": ENC_DAVID_NEWPT,   "ordering_provider_id": PROV_RAJESH,  "ordered_at": ts(2024,12,19,13,50,0),"priority": "routine", "status": "completed", "panel_loinc_code": "4548-4",  "clinical_question": "New DM diagnosis confirmation",                 "fasting_required": False},
    {"id": LO_DAVID_LIPID,   "patient_id": P_DAVID,    "encounter_id": ENC_DAVID_NEWPT,   "ordering_provider_id": PROV_RAJESH,  "ordered_at": ts(2024,12,19,13,50,0),"priority": "routine", "status": "completed", "panel_loinc_code": None,      "clinical_question": "Baseline lipids for new DM",                     "fasting_required": True},
    {"id": LO_AISHA_CBC,     "patient_id": P_AISHA,    "encounter_id": ENC_AISHA_OB,      "ordering_provider_id": PROV_SARA,    "ordered_at": ts(2025,1,16,9,50,0),  "priority": "routine", "status": "completed", "panel_loinc_code": "57021-8", "clinical_question": "Third trimester anemia screen",                 "fasting_required": False},
    {"id": LO_CARLOS_TROP,   "patient_id": P_CARLOS,   "encounter_id": ENC_CARLOS_ED,     "ordering_provider_id": PROV_JOHANNA, "ordered_at": ts(2024,12,28,21,40,0),"priority": "stat",    "status": "completed", "panel_loinc_code": "10839-9", "clinical_question": "R/O ACS - serial troponins q3h",                "fasting_required": False},
]

LAB_SPECIMENS = [
    {"id": uid("spec", "margaret-cmp"),  "lab_order_id": LO_MARGARET_CMP,  "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2025,1,13,9,35,0),  "collected_by_provider_id": PROV_MORGAN,  "received_in_lab_at": ts(2025,1,13,10,2,0),  "volume_ml": 5.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "margaret-cbc"),  "lab_order_id": LO_MARGARET_CBC,  "specimen_type": "blood_edta", "container": "Lavender (EDTA)","collected_at": ts(2025,1,13,9,35,0),  "collected_by_provider_id": PROV_MORGAN,  "received_in_lab_at": ts(2025,1,13,10,2,0),  "volume_ml": 4.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "margaret-a1c"),  "lab_order_id": LO_MARGARET_A1C,  "specimen_type": "blood_edta", "container": "Lavender (EDTA)","collected_at": ts(2025,1,13,9,35,0),  "collected_by_provider_id": PROV_MORGAN,  "received_in_lab_at": ts(2025,1,13,10,2,0),  "volume_ml": 3.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "robert-trop-1"), "lab_order_id": LO_ROBERT_TROP,   "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2024,11,4,16,40,0), "collected_by_provider_id": PROV_RYAN_RN, "received_in_lab_at": ts(2024,11,4,16,55,0), "volume_ml": 5.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "robert-trop-2"), "lab_order_id": LO_ROBERT_TROP,   "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2024,11,4,19,30,0), "collected_by_provider_id": PROV_RYAN_RN, "received_in_lab_at": ts(2024,11,4,19,50,0), "volume_ml": 5.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "robert-cmp-ed"),"lab_order_id": LO_ROBERT_CMP_ED,  "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2024,11,4,16,40,0), "collected_by_provider_id": PROV_RYAN_RN, "received_in_lab_at": ts(2024,11,4,16,55,0), "volume_ml": 7.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "robert-lipid"), "lab_order_id": LO_ROBERT_LIPID,   "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2024,12,12,8,30,0), "collected_by_provider_id": PROV_LISA,    "received_in_lab_at": ts(2024,12,12,9,0,0),  "volume_ml": 5.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "james-cbc-ed"), "lab_order_id": LO_JAMES_CBC_ED,   "specimen_type": "blood_edta", "container": "Lavender (EDTA)","collected_at": ts(2025,1,4,3,5,0),    "collected_by_provider_id": PROV_RYAN_RN, "received_in_lab_at": ts(2025,1,4,3,20,0),   "volume_ml": 4.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "james-bmp-ed"), "lab_order_id": LO_JAMES_BMP_ED,   "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2025,1,4,3,5,0),    "collected_by_provider_id": PROV_RYAN_RN, "received_in_lab_at": ts(2025,1,4,3,20,0),   "volume_ml": 5.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "david-a1c"),    "lab_order_id": LO_DAVID_A1C,      "specimen_type": "blood_edta", "container": "Lavender (EDTA)","collected_at": ts(2024,12,19,13,52,0),"collected_by_provider_id": PROV_RAJESH,  "received_in_lab_at": ts(2024,12,19,14,15,0),"volume_ml": 3.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "david-lipid"),  "lab_order_id": LO_DAVID_LIPID,    "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2024,12,19,13,52,0),"collected_by_provider_id": PROV_RAJESH,  "received_in_lab_at": ts(2024,12,19,14,15,0),"volume_ml": 5.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "aisha-cbc"),    "lab_order_id": LO_AISHA_CBC,      "specimen_type": "blood_edta", "container": "Lavender (EDTA)","collected_at": ts(2025,1,16,9,55,0),  "collected_by_provider_id": PROV_SARA,    "received_in_lab_at": ts(2025,1,16,10,20,0), "volume_ml": 4.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "carlos-trop-1"),"lab_order_id": LO_CARLOS_TROP,    "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2024,12,28,21,45,0),"collected_by_provider_id": PROV_RYAN_RN, "received_in_lab_at": ts(2024,12,28,22,0,0), "volume_ml": 5.0, "is_rejected": False, "rejection_reason": None},
    {"id": uid("spec", "carlos-trop-2"),"lab_order_id": LO_CARLOS_TROP,    "specimen_type": "blood_serum","container": "SST (gold)",     "collected_at": ts(2024,12,29,0,45,0), "collected_by_provider_id": PROV_RYAN_RN, "received_in_lab_at": ts(2024,12,29,1,0,0),  "volume_ml": 5.0, "is_rejected": False, "rejection_reason": None},
]

LAB_RESULTS = [
    # Margaret CMP - sodium, K, Cl, CO2, Glu, BUN, Cr
    {"id": uid("res", "margaret-na"),  "lab_order_id": LO_MARGARET_CMP, "lab_specimen_id": uid("spec","margaret-cmp"), "loinc_code": "2951-2", "value_numeric": 138, "value_text": None, "units": "mmol/L", "reference_range_low": 136, "reference_range_high": 145, "interpretation": "N", "is_critical": False, "resulted_at": ts(2025,1,13,11,30,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-k"),   "lab_order_id": LO_MARGARET_CMP, "lab_specimen_id": uid("spec","margaret-cmp"), "loinc_code": "2823-3", "value_numeric": 5.1, "value_text": None, "units": "mmol/L", "reference_range_low": 3.5, "reference_range_high": 5.0, "interpretation": "H", "is_critical": False, "resulted_at": ts(2025,1,13,11,30,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-cl"),  "lab_order_id": LO_MARGARET_CMP, "lab_specimen_id": uid("spec","margaret-cmp"), "loinc_code": "2075-0", "value_numeric": 103, "value_text": None, "units": "mmol/L", "reference_range_low": 98,  "reference_range_high": 107, "interpretation": "N", "is_critical": False, "resulted_at": ts(2025,1,13,11,30,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-co2"), "lab_order_id": LO_MARGARET_CMP, "lab_specimen_id": uid("spec","margaret-cmp"), "loinc_code": "2028-9", "value_numeric": 24,  "value_text": None, "units": "mmol/L", "reference_range_low": 22,  "reference_range_high": 30, "interpretation": "N", "is_critical": False, "resulted_at": ts(2025,1,13,11,30,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-glu"), "lab_order_id": LO_MARGARET_CMP, "lab_specimen_id": uid("spec","margaret-cmp"), "loinc_code": "2345-7", "value_numeric": 158, "value_text": None, "units": "mg/dL",  "reference_range_low": 70,  "reference_range_high": 99, "interpretation": "H", "is_critical": False, "resulted_at": ts(2025,1,13,11,30,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-bun"), "lab_order_id": LO_MARGARET_CMP, "lab_specimen_id": uid("spec","margaret-cmp"), "loinc_code": "3094-0", "value_numeric": 32,  "value_text": None, "units": "mg/dL",  "reference_range_low": 7,   "reference_range_high": 20, "interpretation": "H", "is_critical": False, "resulted_at": ts(2025,1,13,11,30,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-cr"),  "lab_order_id": LO_MARGARET_CMP, "lab_specimen_id": uid("spec","margaret-cmp"), "loinc_code": "2160-0", "value_numeric": 1.42,"value_text": None, "units": "mg/dL",  "reference_range_low": 0.6, "reference_range_high": 1.1,"interpretation": "H", "is_critical": False, "resulted_at": ts(2025,1,13,11,30,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-egfr"),"lab_order_id": LO_MARGARET_CMP, "lab_specimen_id": uid("spec","margaret-cmp"), "loinc_code": "33914-3","value_numeric": 44,  "value_text": "Stage 3b CKD","units":"mL/min/{1.73_m2}","reference_range_low":60,"reference_range_high":None,"interpretation":"L","is_critical":False,"resulted_at":ts(2025,1,13,11,30,0),"verified_by_provider_id":PROV_MORGAN,"status":"final"},
    # Margaret CBC
    {"id": uid("res", "margaret-hgb"), "lab_order_id": LO_MARGARET_CBC, "lab_specimen_id": uid("spec","margaret-cbc"), "loinc_code": "718-7",  "value_numeric": 10.8,"value_text": None, "units": "g/dL",   "reference_range_low": 12.0,"reference_range_high": 16.0,"interpretation": "L","is_critical": False, "resulted_at": ts(2025,1,13,11,15,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-hct"), "lab_order_id": LO_MARGARET_CBC, "lab_specimen_id": uid("spec","margaret-cbc"), "loinc_code": "4544-3", "value_numeric": 32.5,"value_text": None, "units": "%",      "reference_range_low": 36.0,"reference_range_high": 46.0,"interpretation": "L","is_critical": False, "resulted_at": ts(2025,1,13,11,15,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-wbc"), "lab_order_id": LO_MARGARET_CBC, "lab_specimen_id": uid("spec","margaret-cbc"), "loinc_code": "6690-2", "value_numeric": 7.2, "value_text": None, "units": "10*3/uL","reference_range_low": 4.5, "reference_range_high": 11.0,"interpretation": "N","is_critical": False, "resulted_at": ts(2025,1,13,11,15,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    {"id": uid("res", "margaret-plt"), "lab_order_id": LO_MARGARET_CBC, "lab_specimen_id": uid("spec","margaret-cbc"), "loinc_code": "777-3",  "value_numeric": 245, "value_text": None, "units": "10*3/uL","reference_range_low": 150, "reference_range_high": 400,"interpretation": "N","is_critical": False, "resulted_at": ts(2025,1,13,11,15,0), "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    # Margaret A1c
    {"id": uid("res", "margaret-a1cv"),"lab_order_id": LO_MARGARET_A1C, "lab_specimen_id": uid("spec","margaret-a1c"), "loinc_code": "4548-4", "value_numeric": 7.8, "value_text": None, "units": "%",      "reference_range_low": 4.0, "reference_range_high": 5.6, "interpretation": "H","is_critical": False, "resulted_at": ts(2025,1,13,12,5,0),  "verified_by_provider_id": PROV_MORGAN, "status": "final"},
    # Robert troponin - rising
    {"id": uid("res", "robert-trop1"), "lab_order_id": LO_ROBERT_TROP,  "lab_specimen_id": uid("spec","robert-trop-1"),"loinc_code": "10839-9","value_numeric": 0.42,"value_text": None, "units": "ng/mL",  "reference_range_low": 0.0, "reference_range_high": 0.04,"interpretation": "H","is_critical": True,  "resulted_at": ts(2024,11,4,17,30,0), "verified_by_provider_id": PROV_HASAN,  "status": "final"},
    {"id": uid("res", "robert-trop2"), "lab_order_id": LO_ROBERT_TROP,  "lab_specimen_id": uid("spec","robert-trop-2"),"loinc_code": "10839-9","value_numeric": 1.85,"value_text": None, "units": "ng/mL",  "reference_range_low": 0.0, "reference_range_high": 0.04,"interpretation": "H","is_critical": True,  "resulted_at": ts(2024,11,4,20,45,0), "verified_by_provider_id": PROV_LISA,   "status": "final"},
    # Robert CMP - K and Cr normal
    {"id": uid("res", "robert-na-ed"), "lab_order_id": LO_ROBERT_CMP_ED,"lab_specimen_id": uid("spec","robert-cmp-ed"),"loinc_code": "2951-2", "value_numeric": 140, "value_text": None, "units": "mmol/L", "reference_range_low": 136, "reference_range_high": 145, "interpretation": "N","is_critical": False, "resulted_at": ts(2024,11,4,17,30,0), "verified_by_provider_id": PROV_HASAN,  "status": "final"},
    {"id": uid("res", "robert-k-ed"),  "lab_order_id": LO_ROBERT_CMP_ED,"lab_specimen_id": uid("spec","robert-cmp-ed"),"loinc_code": "2823-3", "value_numeric": 4.2, "value_text": None, "units": "mmol/L", "reference_range_low": 3.5, "reference_range_high": 5.0, "interpretation": "N","is_critical": False, "resulted_at": ts(2024,11,4,17,30,0), "verified_by_provider_id": PROV_HASAN,  "status": "final"},
    {"id": uid("res", "robert-cr-ed"), "lab_order_id": LO_ROBERT_CMP_ED,"lab_specimen_id": uid("spec","robert-cmp-ed"),"loinc_code": "2160-0", "value_numeric": 1.05,"value_text": None, "units": "mg/dL",  "reference_range_low": 0.7, "reference_range_high": 1.3, "interpretation": "N","is_critical": False, "resulted_at": ts(2024,11,4,17,30,0), "verified_by_provider_id": PROV_HASAN,  "status": "final"},
    # Robert lipid panel post-statin
    {"id": uid("res", "robert-ldl"),   "lab_order_id": LO_ROBERT_LIPID, "lab_specimen_id": uid("spec","robert-lipid"), "loinc_code": "2089-1", "value_numeric": 58,  "value_text": None, "units": "mg/dL",  "reference_range_low": 0,   "reference_range_high": 100,"interpretation": "N","is_critical": False, "resulted_at": ts(2024,12,12,14,0,0), "verified_by_provider_id": PROV_LISA,   "status": "final"},
    {"id": uid("res", "robert-hdl"),   "lab_order_id": LO_ROBERT_LIPID, "lab_specimen_id": uid("spec","robert-lipid"), "loinc_code": "2085-9", "value_numeric": 38,  "value_text": None, "units": "mg/dL",  "reference_range_low": 40,  "reference_range_high": None,"interpretation": "L","is_critical": False, "resulted_at": ts(2024,12,12,14,0,0), "verified_by_provider_id": PROV_LISA,   "status": "final"},
    {"id": uid("res", "robert-trig"),  "lab_order_id": LO_ROBERT_LIPID, "lab_specimen_id": uid("spec","robert-lipid"), "loinc_code": "2571-8", "value_numeric": 142, "value_text": None, "units": "mg/dL",  "reference_range_low": 0,   "reference_range_high": 150,"interpretation": "N","is_critical": False, "resulted_at": ts(2024,12,12,14,0,0), "verified_by_provider_id": PROV_LISA,   "status": "final"},
    # James ED - leukocytosis + AKI
    {"id": uid("res", "james-wbc"),    "lab_order_id": LO_JAMES_CBC_ED, "lab_specimen_id": uid("spec","james-cbc-ed"), "loinc_code": "6690-2", "value_numeric": 18.4,"value_text": None, "units": "10*3/uL","reference_range_low": 4.5, "reference_range_high": 11.0,"interpretation": "H","is_critical": True, "resulted_at": ts(2025,1,4,4,0,0),   "verified_by_provider_id": PROV_JOHANNA,"status": "final"},
    {"id": uid("res", "james-hgb"),    "lab_order_id": LO_JAMES_CBC_ED, "lab_specimen_id": uid("spec","james-cbc-ed"), "loinc_code": "718-7",  "value_numeric": 13.8,"value_text": None, "units": "g/dL",   "reference_range_low": 13.5,"reference_range_high": 17.5,"interpretation": "N","is_critical": False,"resulted_at": ts(2025,1,4,4,0,0),    "verified_by_provider_id": PROV_JOHANNA,"status": "final"},
    {"id": uid("res", "james-cr-ed"),  "lab_order_id": LO_JAMES_BMP_ED, "lab_specimen_id": uid("spec","james-bmp-ed"), "loinc_code": "2160-0", "value_numeric": 1.82,"value_text": "AKI baseline 1.2","units":"mg/dL","reference_range_low":0.7,"reference_range_high":1.3,"interpretation":"H","is_critical":False,"resulted_at":ts(2025,1,4,4,0,0),"verified_by_provider_id":PROV_JOHANNA,"status":"final"},
    {"id": uid("res", "james-k-ed"),   "lab_order_id": LO_JAMES_BMP_ED, "lab_specimen_id": uid("spec","james-bmp-ed"), "loinc_code": "2823-3", "value_numeric": 3.4, "value_text": None, "units": "mmol/L", "reference_range_low": 3.5, "reference_range_high": 5.0, "interpretation": "L","is_critical": False, "resulted_at": ts(2025,1,4,4,0,0),   "verified_by_provider_id": PROV_JOHANNA,"status": "final"},
    # David A1c new dx
    {"id": uid("res", "david-a1c"),    "lab_order_id": LO_DAVID_A1C,    "lab_specimen_id": uid("spec","david-a1c"),    "loinc_code": "4548-4", "value_numeric": 9.2, "value_text": None, "units": "%",      "reference_range_low": 4.0, "reference_range_high": 5.6, "interpretation": "H","is_critical": False, "resulted_at": ts(2024,12,19,16,30,0),"verified_by_provider_id": PROV_RAJESH, "status": "final"},
    # David Lipid
    {"id": uid("res", "david-ldl"),    "lab_order_id": LO_DAVID_LIPID,  "lab_specimen_id": uid("spec","david-lipid"),  "loinc_code": "2089-1", "value_numeric": 142, "value_text": None, "units": "mg/dL",  "reference_range_low": 0,   "reference_range_high": 130,"interpretation": "H","is_critical": False, "resulted_at": ts(2024,12,19,16,30,0),"verified_by_provider_id": PROV_RAJESH, "status": "final"},
    {"id": uid("res", "david-chol"),   "lab_order_id": LO_DAVID_LIPID,  "lab_specimen_id": uid("spec","david-lipid"),  "loinc_code": "2093-3", "value_numeric": 224, "value_text": None, "units": "mg/dL",  "reference_range_low": 0,   "reference_range_high": 200,"interpretation": "H","is_critical": False, "resulted_at": ts(2024,12,19,16,30,0),"verified_by_provider_id": PROV_RAJESH, "status": "final"},
    {"id": uid("res", "david-trig"),   "lab_order_id": LO_DAVID_LIPID,  "lab_specimen_id": uid("spec","david-lipid"),  "loinc_code": "2571-8", "value_numeric": 198, "value_text": None, "units": "mg/dL",  "reference_range_low": 0,   "reference_range_high": 150,"interpretation": "H","is_critical": False, "resulted_at": ts(2024,12,19,16,30,0),"verified_by_provider_id": PROV_RAJESH, "status": "final"},
    # Aisha CBC pregnancy
    {"id": uid("res", "aisha-hgb"),    "lab_order_id": LO_AISHA_CBC,    "lab_specimen_id": uid("spec","aisha-cbc"),    "loinc_code": "718-7",  "value_numeric": 11.4,"value_text": "mild dilutional", "units":"g/dL","reference_range_low":11.0,"reference_range_high":15.0,"interpretation":"N","is_critical":False,"resulted_at":ts(2025,1,16,11,30,0),"verified_by_provider_id":PROV_SARA,"status":"final"},
    {"id": uid("res", "aisha-plt"),    "lab_order_id": LO_AISHA_CBC,    "lab_specimen_id": uid("spec","aisha-cbc"),    "loinc_code": "777-3",  "value_numeric": 198, "value_text": None, "units": "10*3/uL","reference_range_low": 150, "reference_range_high": 400,"interpretation": "N","is_critical": False, "resulted_at": ts(2025,1,16,11,30,0),"verified_by_provider_id": PROV_SARA,   "status": "final"},
    # Carlos troponin - negative
    {"id": uid("res", "carlos-trop1"), "lab_order_id": LO_CARLOS_TROP,  "lab_specimen_id": uid("spec","carlos-trop-1"),"loinc_code": "10839-9","value_numeric": 0.02,"value_text": None, "units": "ng/mL",  "reference_range_low": 0.0, "reference_range_high": 0.04,"interpretation": "N","is_critical": False, "resulted_at": ts(2024,12,28,22,45,0),"verified_by_provider_id": PROV_JOHANNA,"status": "final"},
    {"id": uid("res", "carlos-trop2"), "lab_order_id": LO_CARLOS_TROP,  "lab_specimen_id": uid("spec","carlos-trop-2"),"loinc_code": "10839-9","value_numeric": 0.02,"value_text": None, "units": "ng/mL",  "reference_range_low": 0.0, "reference_range_high": 0.04,"interpretation": "N","is_critical": False, "resulted_at": ts(2024,12,29,1,45,0), "verified_by_provider_id": PROV_JOHANNA,"status": "final"},
]

# Imaging
IO_ROBERT_ECHO   = uid("imorder", "robert-echo-2024-12-12")
IO_JAMES_CXR     = uid("imorder", "james-cxr-2025-01-04")
IO_DOROTHY_HIPXR = uid("imorder", "dorothy-hipxr-2024-11-18")
IO_AISHA_USOB    = uid("imorder", "aisha-usob-2025-01-16")
IO_CARLOS_CTA    = uid("imorder", "carlos-cta-2024-12-28")

IMAGING_ORDERS = [
    {"id": IO_ROBERT_ECHO,   "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_CARDIO,  "ordering_provider_id": PROV_LISA,    "modality": "US",  "body_part": "Heart (TTE)",      "cpt_code": "93306", "clinical_indication": "Post-NSTEMI surveillance, assess LV function and wall motion abnormalities","priority": "routine","status": "completed","requires_contrast": False, "ordered_at": ts(2024,12,5,9,0,0)},
    {"id": IO_JAMES_CXR,     "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_ED,       "ordering_provider_id": PROV_JOHANNA, "modality": "XR",  "body_part": "Chest PA/Lat",     "cpt_code": "71046", "clinical_indication": "Severe SOB, fever, productive cough - evaluate for pneumonia vs CHF",       "priority": "stat",   "status": "completed","requires_contrast": False, "ordered_at": ts(2025,1,4,3,0,0)},
    {"id": IO_DOROTHY_HIPXR, "patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_ED,     "ordering_provider_id": PROV_HASAN,   "modality": "XR",  "body_part": "Right Hip 3-views","cpt_code": "71046", "clinical_indication": "S/p fall, R hip pain, unable to bear weight - r/o fracture",                 "priority": "stat",   "status": "completed","requires_contrast": False, "ordered_at": ts(2024,11,18,12,0,0)},
    {"id": IO_AISHA_USOB,    "patient_id": P_AISHA,   "encounter_id": ENC_AISHA_OB,       "ordering_provider_id": PROV_SARA,    "modality": "US",  "body_part": "Obstetric (Complete 3rd trimester)","cpt_code":"76700","clinical_indication":"32-week growth ultrasound, AFI, biophysical profile","priority":"routine","status": "completed","requires_contrast": False, "ordered_at": ts(2025,1,10,9,0,0)},
    {"id": IO_CARLOS_CTA,    "patient_id": P_CARLOS,  "encounter_id": ENC_CARLOS_ED,      "ordering_provider_id": PROV_JOHANNA, "modality": "CT",  "body_part": "Chest with contrast","cpt_code": "71250","clinical_indication": "Acute pleuritic chest pain - r/o PE","priority": "stat", "status": "completed","requires_contrast": True,  "ordered_at": ts(2024,12,28,22,15,0)},
]

IS_ROBERT_ECHO   = uid("imstudy", "robert-echo")
IS_JAMES_CXR     = uid("imstudy", "james-cxr")
IS_DOROTHY_HIPXR = uid("imstudy", "dorothy-hipxr")
IS_AISHA_USOB    = uid("imstudy", "aisha-usob")
IS_CARLOS_CTA    = uid("imstudy", "carlos-cta")

IMAGING_STUDIES = [
    {"id": IS_ROBERT_ECHO,   "imaging_order_id": IO_ROBERT_ECHO,   "study_uid": "1.2.840.113619.2.55.3.604688310.811.1734005405.1",  "accession_number": "SRMC-IM-20241212-001145", "performed_at": ts(2024,12,12,8,40,0),  "performed_by_provider_id": PROV_LISA,    "location_id": LOC_HOSPITAL, "series_count": 12, "image_count": 280,  "dicom_metadata_doc_id": "dicom_echo_robert_chen"},
    {"id": IS_JAMES_CXR,     "imaging_order_id": IO_JAMES_CXR,     "study_uid": "1.2.840.113619.2.55.3.604688310.811.1735965310.2",  "accession_number": "SRMC-IM-20250104-000087", "performed_at": ts(2025,1,4,3,15,0),    "performed_by_provider_id": PROV_KATHRYN, "location_id": LOC_HOSPITAL, "series_count": 2,  "image_count": 2,    "dicom_metadata_doc_id": "dicom_cxr_james_oconnor"},
    {"id": IS_DOROTHY_HIPXR, "imaging_order_id": IO_DOROTHY_HIPXR, "study_uid": "1.2.840.113619.2.55.3.604688310.811.1731950900.3",  "accession_number": "SRMC-IM-20241118-001872", "performed_at": ts(2024,11,18,12,45,0), "performed_by_provider_id": PROV_KATHRYN, "location_id": LOC_HOSPITAL, "series_count": 3,  "image_count": 3,    "dicom_metadata_doc_id": "dicom_hipxr_dorothy_kim"},
    {"id": IS_AISHA_USOB,    "imaging_order_id": IO_AISHA_USOB,    "study_uid": "1.2.840.113619.2.55.3.604688310.811.1737023400.4",  "accession_number": "SRMC-IM-20250116-001408", "performed_at": ts(2025,1,16,10,30,0),  "performed_by_provider_id": PROV_KATHRYN, "location_id": LOC_HOSPITAL, "series_count": 8,  "image_count": 140,  "dicom_metadata_doc_id": "dicom_us_aisha_rodriguez"},
    {"id": IS_CARLOS_CTA,    "imaging_order_id": IO_CARLOS_CTA,    "study_uid": "1.2.840.113619.2.55.3.604688310.811.1735419300.5",  "accession_number": "SRMC-IM-20241228-002209", "performed_at": ts(2024,12,28,22,40,0), "performed_by_provider_id": PROV_KATHRYN, "location_id": LOC_HOSPITAL, "series_count": 5,  "image_count": 412,  "dicom_metadata_doc_id": "dicom_cta_carlos_mendoza"},
]

IMAGING_REPORTS = [
    {"id": uid("imrpt", "robert-echo"),   "imaging_study_id": IS_ROBERT_ECHO,   "reading_radiologist_id": PROV_LISA,    "dictated_at": ts(2024,12,12,9,5,0),    "signed_at": ts(2024,12,12,9,30,0),  "findings":  "LVEDD 5.0 cm. LVEF 50% (mildly reduced) with mild anterior wall hypokinesis. RV size/function normal. No pericardial effusion. Mild mitral regurgitation. PASP 28 mmHg.","impression": "1. Mildly reduced LV systolic function (LVEF 50%) with mild anterior hypokinesis consistent with prior anterior MI. 2. Mild mitral regurgitation. 3. No pericardial effusion.","recommendation": "Continue current cardiac medications. Consider repeat TTE in 6 months.","status": "final"},
    {"id": uid("imrpt", "james-cxr"),     "imaging_study_id": IS_JAMES_CXR,     "reading_radiologist_id": PROV_KATHRYN, "dictated_at": ts(2025,1,4,3,45,0),     "signed_at": ts(2025,1,4,4,15,0),    "findings":  "Hyperinflation of bilateral lung fields with flattened diaphragms consistent with COPD. New patchy airspace opacity in the right lower lobe consistent with pneumonia. Mild cardiomegaly. No pneumothorax. No pleural effusion.","impression": "1. RLL pneumonia. 2. Severe COPD with hyperinflation. 3. Mild cardiomegaly, stable.","recommendation": "Clinical correlation; consider repeat CXR in 4-6 weeks to confirm resolution.","status": "final"},
    {"id": uid("imrpt", "dorothy-hipxr"), "imaging_study_id": IS_DOROTHY_HIPXR, "reading_radiologist_id": PROV_KATHRYN, "dictated_at": ts(2024,11,18,13,15,0),  "signed_at": ts(2024,11,18,13,40,0), "findings":  "Displaced subcapital fracture of the right femoral neck with mild varus angulation. No additional fractures identified. Joint space preserved. Soft tissues unremarkable.","impression": "Acute displaced right femoral neck fracture.","recommendation": "Orthopedic surgical consultation - candidate for hemiarthroplasty vs THA.","status": "final"},
    {"id": uid("imrpt", "aisha-usob"),    "imaging_study_id": IS_AISHA_USOB,    "reading_radiologist_id": PROV_KATHRYN, "dictated_at": ts(2025,1,16,11,0,0),    "signed_at": ts(2025,1,16,11,20,0),  "findings":  "Single live intrauterine pregnancy. EFW 1850 g (47th percentile). AFI 14 cm (normal). Placenta posterior, not previa. BPP 8/8. FHR 148 bpm. Cephalic presentation.","impression": "1. Singleton viable pregnancy at 32w0d by composite dating. 2. Appropriate fetal growth. 3. Normal amniotic fluid. 4. Reassuring BPP.","recommendation": "Routine prenatal care; next scheduled visit at 34 weeks.","status": "final"},
    {"id": uid("imrpt", "carlos-cta"),    "imaging_study_id": IS_CARLOS_CTA,    "reading_radiologist_id": PROV_KATHRYN, "dictated_at": ts(2024,12,28,23,10,0),  "signed_at": ts(2024,12,28,23,30,0), "findings":  "No filling defect in main, lobar, segmental, or subsegmental pulmonary arteries. Lung parenchyma clear. No pneumothorax. No pleural effusion. Heart size normal. No mediastinal lymphadenopathy.","impression": "Negative CTPA for pulmonary embolism. No alternative cause of chest pain identified.","recommendation": "Clinical correlation - consider non-thromboembolic etiologies (anxiety, GERD, MSK).","status": "final"},
]

# ---------------------------------------------------------------------------
# 10. BILLING — claims, lines, charges, payments, adjustments, denials, appeals, statements
# ---------------------------------------------------------------------------

CLM_MARGARET_FU      = uid("claim", "margaret-fu-2025-01-13")
CLM_ROBERT_INPT      = uid("claim", "robert-nstemi-2024-11-04")
CLM_ROBERT_CARDIO    = uid("claim", "robert-cardio-2024-12-12")
CLM_LAKSHMI_FU       = uid("claim", "lakshmi-fu-2025-01-13")
CLM_JAMES_INPT       = uid("claim", "james-copd-2025-01-04")
CLM_DOROTHY_INPT     = uid("claim", "dorothy-tha-2024-11-18")
CLM_DAVID_NEWPT      = uid("claim", "david-newpt-2024-12-19")
CLM_CARLOS_ED        = uid("claim", "carlos-ed-2024-12-28")
CLM_EMILY_ED         = uid("claim", "emily-ed-2024-12-15")
CLM_AISHA_OB         = uid("claim", "aisha-ob-2025-01-16")
CLM_MARCUS_LBP       = uid("claim", "marcus-lbp-2025-01-08")
CLM_DOROTHY_PREOP    = uid("claim", "dorothy-preop-2024-11-26")

CLAIMS = [
    {"id": CLM_MARGARET_FU,   "patient_id": P_MARGARET, "encounter_id": ENC_MARGARET_FU,   "coverage_id": COV_MARGARET_MCARE,"claim_number": "SRMC-CLM-2025-000142", "payer_claim_id": "MCARE-202501-3982017", "claim_type": "professional", "status": "paid",     "service_start_date": "2025-01-13", "service_end_date": "2025-01-13", "total_charge": 285.00, "total_allowed": 142.18,"total_paid": 113.74, "patient_responsibility": 28.44, "submitted_at": ts(2025,1,14,9,30,0),  "paid_at": ts(2025,1,28,14,15,0), "billing_provider_id": PROV_MORGAN},
    {"id": CLM_ROBERT_INPT,   "patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_INPT,   "coverage_id": COV_ROBERT_AETNA,  "claim_number": "SRMC-CLM-2024-009823", "payer_claim_id": "AET-2024-7281047",     "claim_type": "institutional","status": "paid",     "service_start_date": "2024-11-04", "service_end_date": "2024-11-07", "total_charge": 38420.00,"total_allowed": 24587.50,"total_paid": 22928.75,"patient_responsibility": 1658.75,"submitted_at": ts(2024,11,8,16,0,0),  "paid_at": ts(2024,12,5,10,30,0), "billing_provider_id": PROV_LISA},
    {"id": CLM_ROBERT_CARDIO, "patient_id": P_ROBERT,   "encounter_id": ENC_ROBERT_CARDIO, "coverage_id": COV_ROBERT_AETNA,  "claim_number": "SRMC-CLM-2024-010472", "payer_claim_id": "AET-2024-7298314",     "claim_type": "professional", "status": "paid",     "service_start_date": "2024-12-12", "service_end_date": "2024-12-12", "total_charge": 640.00, "total_allowed": 425.20,"total_paid": 375.20, "patient_responsibility": 50.00, "submitted_at": ts(2024,12,13,10,0,0), "paid_at": ts(2025,1,9,11,15,0),  "billing_provider_id": PROV_LISA},
    {"id": CLM_LAKSHMI_FU,    "patient_id": P_LAKSHMI,  "encounter_id": ENC_LAKSHMI_FU,    "coverage_id": COV_LAKSHMI_BCBS,  "claim_number": "SRMC-CLM-2025-000148", "payer_claim_id": "BSC-2025-9023841",     "claim_type": "professional", "status": "paid",     "service_start_date": "2025-01-13", "service_end_date": "2025-01-13", "total_charge": 145.00, "total_allowed": 92.50, "total_paid": 62.50,  "patient_responsibility": 30.00, "submitted_at": ts(2025,1,14,10,0,0),  "paid_at": ts(2025,1,30,15,0,0),  "billing_provider_id": PROV_MORGAN},
    {"id": CLM_JAMES_INPT,    "patient_id": P_JAMES,    "encounter_id": ENC_JAMES_INPT,    "coverage_id": COV_JAMES_MCARE,   "claim_number": "SRMC-CLM-2025-000209", "payer_claim_id": "MCARE-202501-3984708", "claim_type": "institutional","status": "submitted","service_start_date": "2025-01-04", "service_end_date": "2025-01-09", "total_charge": 52814.00,"total_allowed": None,    "total_paid": None,    "patient_responsibility": None,   "submitted_at": ts(2025,1,11,14,30,0), "paid_at": None,                  "billing_provider_id": PROV_AMANDA},
    {"id": CLM_DOROTHY_INPT,  "patient_id": P_DOROTHY,  "encounter_id": ENC_DOROTHY_INPT,  "coverage_id": COV_DOROTHY_MCARE, "claim_number": "SRMC-CLM-2024-010038", "payer_claim_id": "MCARE-202411-3978293", "claim_type": "institutional","status": "paid",     "service_start_date": "2024-11-18", "service_end_date": "2024-11-22", "total_charge": 32850.00,"total_allowed": 21420.00,"total_paid": 19847.00,"patient_responsibility": 1573.00,"submitted_at": ts(2024,11,25,11,0,0), "paid_at": ts(2024,12,18,9,15,0), "billing_provider_id": PROV_DANIEL},
    {"id": CLM_DAVID_NEWPT,   "patient_id": P_DAVID,    "encounter_id": ENC_DAVID_NEWPT,   "coverage_id": COV_DAVID_AETNA,   "claim_number": "SRMC-CLM-2024-010154", "payer_claim_id": "AET-2024-7301478",     "claim_type": "professional", "status": "paid",     "service_start_date": "2024-12-19", "service_end_date": "2024-12-19", "total_charge": 425.00, "total_allowed": 218.50,"total_paid": 168.50, "patient_responsibility": 50.00, "submitted_at": ts(2024,12,20,9,30,0), "paid_at": ts(2025,1,8,11,20,0),  "billing_provider_id": PROV_RAJESH},
    {"id": CLM_CARLOS_ED,     "patient_id": P_CARLOS,   "encounter_id": ENC_CARLOS_ED,     "coverage_id": COV_CARLOS_KAISER, "claim_number": "SRMC-CLM-2024-010237", "payer_claim_id": "KP-2024-8472019",      "claim_type": "institutional","status": "denied",   "service_start_date": "2024-12-28", "service_end_date": "2024-12-29", "total_charge": 4285.00,"total_allowed": 0.00,    "total_paid": 0.00,    "patient_responsibility": 4285.00,"submitted_at": ts(2024,12,31,10,0,0), "paid_at": None,                  "billing_provider_id": PROV_JOHANNA},
    {"id": CLM_EMILY_ED,      "patient_id": P_EMILY,    "encounter_id": ENC_EMILY_ED,      "coverage_id": COV_EMILY_AETNA,   "claim_number": "SRMC-CLM-2024-010172", "payer_claim_id": "AET-2024-7305821",     "claim_type": "institutional","status": "paid",     "service_start_date": "2024-12-15", "service_end_date": "2024-12-16", "total_charge": 1850.00,"total_allowed": 1248.50,"total_paid": 0.00,    "patient_responsibility": 1248.50,"submitted_at": ts(2024,12,17,11,0,0), "paid_at": ts(2025,1,12,14,0,0),  "billing_provider_id": PROV_HASAN},
    {"id": CLM_AISHA_OB,      "patient_id": P_AISHA,    "encounter_id": ENC_AISHA_OB,      "coverage_id": COV_AISHA_MEDICAL, "claim_number": "SRMC-CLM-2025-000158", "payer_claim_id": "MCAL-2025-1024571",    "claim_type": "professional", "status": "paid",     "service_start_date": "2025-01-16", "service_end_date": "2025-01-16", "total_charge": 528.00, "total_allowed": 198.40,"total_paid": 198.40, "patient_responsibility": 0.00,   "submitted_at": ts(2025,1,17,9,0,0),   "paid_at": ts(2025,2,7,10,0,0),   "billing_provider_id": PROV_SARA},
    {"id": CLM_MARCUS_LBP,    "patient_id": P_MARCUS,   "encounter_id": ENC_MARCUS_LBP,    "coverage_id": COV_MARCUS_UHC,    "claim_number": "SRMC-CLM-2025-000138", "payer_claim_id": "UHC-2025-2389104",     "claim_type": "professional", "status": "paid",     "service_start_date": "2025-01-08", "service_end_date": "2025-01-08", "total_charge": 145.00, "total_allowed": 88.40, "total_paid": 53.40,  "patient_responsibility": 35.00, "submitted_at": ts(2025,1,9,10,30,0),  "paid_at": ts(2025,1,26,13,0,0),  "billing_provider_id": PROV_MORGAN},
    {"id": CLM_DOROTHY_PREOP, "patient_id": P_DOROTHY,  "encounter_id": ENC_DOROTHY_PREOP, "coverage_id": COV_DOROTHY_MCARE, "claim_number": "SRMC-CLM-2024-010012", "payer_claim_id": "MCARE-202411-3977208", "claim_type": "professional", "status": "paid",     "service_start_date": "2024-11-26", "service_end_date": "2024-11-26", "total_charge": 310.00, "total_allowed": 162.45,"total_paid": 129.96, "patient_responsibility": 32.49, "submitted_at": ts(2024,11,27,9,0,0),  "paid_at": ts(2024,12,15,11,0,0), "billing_provider_id": PROV_RAJESH},
]

# Charges (per encounter, per CPT)
CHARGES = [
    {"id": uid("chg", "margaret-fu-99214"),"encounter_id": ENC_MARGARET_FU, "patient_id": P_MARGARET, "cpt_code": "99214", "quantity": 1, "charge_amount": 215.00, "posted_at": ts(2025,1,13,9,40,0),  "posted_by_provider_id": PROV_MORGAN, "claim_id": CLM_MARGARET_FU},
    {"id": uid("chg", "margaret-fu-36415"),"encounter_id": ENC_MARGARET_FU, "patient_id": P_MARGARET, "cpt_code": "36415", "quantity": 1, "charge_amount": 25.00,  "posted_at": ts(2025,1,13,9,40,0),  "posted_by_provider_id": PROV_MORGAN, "claim_id": CLM_MARGARET_FU},
    {"id": uid("chg", "margaret-fu-80053"),"encounter_id": ENC_MARGARET_FU, "patient_id": P_MARGARET, "cpt_code": "80053", "quantity": 1, "charge_amount": 65.00,  "posted_at": ts(2025,1,13,12,0,0),  "posted_by_provider_id": PROV_MORGAN, "claim_id": CLM_MARGARET_FU},
    {"id": uid("chg", "robert-cardio-99214"),"encounter_id":ENC_ROBERT_CARDIO,"patient_id":P_ROBERT,  "cpt_code":"99214","quantity":1,"charge_amount":215.00,"posted_at":ts(2024,12,12,9,30,0),"posted_by_provider_id":PROV_LISA,"claim_id":CLM_ROBERT_CARDIO},
    {"id": uid("chg", "robert-cardio-93306"),"encounter_id":ENC_ROBERT_CARDIO,"patient_id":P_ROBERT,  "cpt_code":"93306","quantity":1,"charge_amount":425.00,"posted_at":ts(2024,12,12,9,30,0),"posted_by_provider_id":PROV_LISA,"claim_id":CLM_ROBERT_CARDIO},
    {"id": uid("chg", "lakshmi-fu-99213"), "encounter_id": ENC_LAKSHMI_FU,  "patient_id": P_LAKSHMI,  "cpt_code": "99213", "quantity": 1, "charge_amount": 145.00, "posted_at": ts(2025,1,13,10,30,0), "posted_by_provider_id": PROV_MORGAN, "claim_id": CLM_LAKSHMI_FU},
    {"id": uid("chg", "marcus-lbp-99213"), "encounter_id": ENC_MARCUS_LBP,  "patient_id": P_MARCUS,   "cpt_code": "99213", "quantity": 1, "charge_amount": 145.00, "posted_at": ts(2025,1,8,15,40,0),  "posted_by_provider_id": PROV_MORGAN, "claim_id": CLM_MARCUS_LBP},
    {"id": uid("chg", "david-newpt-99203"),"encounter_id": ENC_DAVID_NEWPT, "patient_id": P_DAVID,    "cpt_code": "99203", "quantity": 1, "charge_amount": 215.00, "posted_at": ts(2024,12,19,14,0,0), "posted_by_provider_id": PROV_RAJESH, "claim_id": CLM_DAVID_NEWPT},
    {"id": uid("chg", "david-newpt-36415"),"encounter_id": ENC_DAVID_NEWPT, "patient_id": P_DAVID,    "cpt_code": "36415", "quantity": 1, "charge_amount": 25.00,  "posted_at": ts(2024,12,19,14,0,0), "posted_by_provider_id": PROV_RAJESH, "claim_id": CLM_DAVID_NEWPT},
    {"id": uid("chg", "david-newpt-83036"),"encounter_id": ENC_DAVID_NEWPT, "patient_id": P_DAVID,    "cpt_code": "83036", "quantity": 1, "charge_amount": 55.00,  "posted_at": ts(2024,12,19,16,30,0),"posted_by_provider_id": PROV_RAJESH, "claim_id": CLM_DAVID_NEWPT},
    {"id": uid("chg", "david-newpt-80061"),"encounter_id": ENC_DAVID_NEWPT, "patient_id": P_DAVID,    "cpt_code": "80061", "quantity": 1, "charge_amount": 75.00,  "posted_at": ts(2024,12,19,16,30,0),"posted_by_provider_id": PROV_RAJESH, "claim_id": CLM_DAVID_NEWPT},
    {"id": uid("chg", "robert-inpt-99221"),"encounter_id": ENC_ROBERT_INPT, "patient_id": P_ROBERT,   "cpt_code": "99221", "quantity": 1, "charge_amount": 312.00, "posted_at": ts(2024,11,4,21,0,0),  "posted_by_provider_id": PROV_LISA,   "claim_id": CLM_ROBERT_INPT},
    {"id": uid("chg", "robert-inpt-99232"),"encounter_id": ENC_ROBERT_INPT, "patient_id": P_ROBERT,   "cpt_code": "99232", "quantity": 2, "charge_amount": 350.00, "posted_at": ts(2024,11,6,9,0,0),   "posted_by_provider_id": PROV_LISA,   "claim_id": CLM_ROBERT_INPT},
    {"id": uid("chg", "james-inpt-99221"), "encounter_id": ENC_JAMES_INPT,  "patient_id": P_JAMES,    "cpt_code": "99221", "quantity": 1, "charge_amount": 312.00, "posted_at": ts(2025,1,4,7,0,0),    "posted_by_provider_id": PROV_AMANDA, "claim_id": CLM_JAMES_INPT},
    {"id": uid("chg", "james-inpt-99291"), "encounter_id": ENC_JAMES_INPT,  "patient_id": P_JAMES,    "cpt_code": "99291", "quantity": 1, "charge_amount": 685.00, "posted_at": ts(2025,1,4,7,0,0),    "posted_by_provider_id": PROV_AMANDA, "claim_id": CLM_JAMES_INPT},
    {"id": uid("chg", "james-inpt-99232"), "encounter_id": ENC_JAMES_INPT,  "patient_id": P_JAMES,    "cpt_code": "99232", "quantity": 4, "charge_amount": 700.00, "posted_at": ts(2025,1,9,9,0,0),    "posted_by_provider_id": PROV_AMANDA, "claim_id": CLM_JAMES_INPT},
    {"id": uid("chg", "dorothy-tha-27130"),"encounter_id": ENC_DOROTHY_INPT,"patient_id": P_DOROTHY,  "cpt_code": "27130", "quantity": 1, "charge_amount": 8500.00,"posted_at": ts(2024,11,19,10,0,0), "posted_by_provider_id": PROV_DANIEL, "claim_id": CLM_DOROTHY_INPT},
    {"id": uid("chg", "dorothy-inpt-99221"),"encounter_id": ENC_DOROTHY_INPT,"patient_id": P_DOROTHY, "cpt_code": "99221", "quantity": 1, "charge_amount": 312.00, "posted_at": ts(2024,11,18,18,0,0), "posted_by_provider_id": PROV_DANIEL, "claim_id": CLM_DOROTHY_INPT},
    {"id": uid("chg", "carlos-ed-99284"),  "encounter_id": ENC_CARLOS_ED,   "patient_id": P_CARLOS,   "cpt_code": "99284", "quantity": 1, "charge_amount": 565.00, "posted_at": ts(2024,12,29,1,30,0), "posted_by_provider_id": PROV_JOHANNA,"claim_id": CLM_CARLOS_ED},
    {"id": uid("chg", "carlos-ed-71250"),  "encounter_id": ENC_CARLOS_ED,   "patient_id": P_CARLOS,   "cpt_code": "71250", "quantity": 1, "charge_amount": 525.00, "posted_at": ts(2024,12,29,1,30,0), "posted_by_provider_id": PROV_JOHANNA,"claim_id": CLM_CARLOS_ED},
    {"id": uid("chg", "carlos-ed-84484"),  "encounter_id": ENC_CARLOS_ED,   "patient_id": P_CARLOS,   "cpt_code": "84484", "quantity": 2, "charge_amount": 190.00, "posted_at": ts(2024,12,29,1,30,0), "posted_by_provider_id": PROV_JOHANNA,"claim_id": CLM_CARLOS_ED},
    {"id": uid("chg", "emily-ed-99284"),   "encounter_id": ENC_EMILY_ED,    "patient_id": P_EMILY,    "cpt_code": "99284", "quantity": 1, "charge_amount": 565.00, "posted_at": ts(2024,12,16,0,45,0), "posted_by_provider_id": PROV_HASAN,  "claim_id": CLM_EMILY_ED},
    {"id": uid("chg", "emily-ed-94640"),   "encounter_id": ENC_EMILY_ED,    "patient_id": P_EMILY,    "cpt_code": "94640", "quantity": 3, "charge_amount": 195.00, "posted_at": ts(2024,12,16,0,45,0), "posted_by_provider_id": PROV_HASAN,  "claim_id": CLM_EMILY_ED},
    {"id": uid("chg", "aisha-ob-99214"),   "encounter_id": ENC_AISHA_OB,    "patient_id": P_AISHA,    "cpt_code": "99214", "quantity": 1, "charge_amount": 215.00, "posted_at": ts(2025,1,16,10,0,0),  "posted_by_provider_id": PROV_SARA,   "claim_id": CLM_AISHA_OB},
    {"id": uid("chg", "aisha-ob-90471"),   "encounter_id": ENC_AISHA_OB,    "patient_id": P_AISHA,    "cpt_code": "90471", "quantity": 1, "charge_amount": 28.00,  "posted_at": ts(2025,1,16,10,0,0),  "posted_by_provider_id": PROV_SARA,   "claim_id": CLM_AISHA_OB},
    {"id": uid("chg", "aisha-ob-76700"),   "encounter_id": ENC_AISHA_OB,    "patient_id": P_AISHA,    "cpt_code": "76700", "quantity": 1, "charge_amount": 285.00, "posted_at": ts(2025,1,16,11,0,0),  "posted_by_provider_id": PROV_KATHRYN,"claim_id": CLM_AISHA_OB},
    {"id": uid("chg", "dorothy-preop-99214"),"encounter_id": ENC_DOROTHY_PREOP,"patient_id": P_DOROTHY,"cpt_code":"99214","quantity":1,"charge_amount":215.00,"posted_at":ts(2024,11,26,15,0,0),"posted_by_provider_id":PROV_RAJESH,"claim_id":CLM_DOROTHY_PREOP},
    {"id": uid("chg", "dorothy-preop-93000"),"encounter_id": ENC_DOROTHY_PREOP,"patient_id": P_DOROTHY,"cpt_code":"93000","quantity":1,"charge_amount":95.00,"posted_at":ts(2024,11,26,15,0,0),"posted_by_provider_id":PROV_RAJESH,"claim_id":CLM_DOROTHY_PREOP},
]

CLAIM_LINES = [
    {"id": uid("cline", "margaret-fu-1"),       "claim_id": CLM_MARGARET_FU,    "line_number": 1, "cpt_code": "99214", "modifier": None, "icd10_pointer": "1,2,3,4", "service_date": "2025-01-13", "units": 1, "charge_amount": 215.00, "allowed_amount": 107.18, "paid_amount": 85.74,  "adjustment_amount": 107.82, "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_MORGAN},
    {"id": uid("cline", "margaret-fu-2"),       "claim_id": CLM_MARGARET_FU,    "line_number": 2, "cpt_code": "36415", "modifier": None, "icd10_pointer": "1",       "service_date": "2025-01-13", "units": 1, "charge_amount": 25.00,  "allowed_amount": 5.00,   "paid_amount": 4.00,   "adjustment_amount": 20.00,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_MORGAN},
    {"id": uid("cline", "margaret-fu-3"),       "claim_id": CLM_MARGARET_FU,    "line_number": 3, "cpt_code": "80053", "modifier": None, "icd10_pointer": "1,3",     "service_date": "2025-01-13", "units": 1, "charge_amount": 65.00,  "allowed_amount": 30.00,  "paid_amount": 24.00,  "adjustment_amount": 35.00,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_MORGAN},
    {"id": uid("cline", "robert-cardio-1"),     "claim_id": CLM_ROBERT_CARDIO,  "line_number": 1, "cpt_code": "99214", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-12-12", "units": 1, "charge_amount": 215.00, "allowed_amount": 175.20, "paid_amount": 125.20, "adjustment_amount": 39.80,  "denial_code": None, "place_of_service": "22", "rendering_provider_id": PROV_LISA},
    {"id": uid("cline", "robert-cardio-2"),     "claim_id": CLM_ROBERT_CARDIO,  "line_number": 2, "cpt_code": "93306", "modifier": "26", "icd10_pointer": "1",       "service_date": "2024-12-12", "units": 1, "charge_amount": 425.00, "allowed_amount": 250.00, "paid_amount": 250.00, "adjustment_amount": 175.00, "denial_code": None, "place_of_service": "22", "rendering_provider_id": PROV_LISA},
    {"id": uid("cline", "lakshmi-fu-1"),        "claim_id": CLM_LAKSHMI_FU,     "line_number": 1, "cpt_code": "99213", "modifier": None, "icd10_pointer": "1,2",     "service_date": "2025-01-13", "units": 1, "charge_amount": 145.00, "allowed_amount": 92.50,  "paid_amount": 62.50,  "adjustment_amount": 52.50,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_MORGAN},
    {"id": uid("cline", "marcus-lbp-1"),        "claim_id": CLM_MARCUS_LBP,     "line_number": 1, "cpt_code": "99213", "modifier": None, "icd10_pointer": "1",       "service_date": "2025-01-08", "units": 1, "charge_amount": 145.00, "allowed_amount": 88.40,  "paid_amount": 53.40,  "adjustment_amount": 56.60,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_MORGAN},
    {"id": uid("cline", "david-newpt-1"),       "claim_id": CLM_DAVID_NEWPT,    "line_number": 1, "cpt_code": "99203", "modifier": None, "icd10_pointer": "1,2",     "service_date": "2024-12-19", "units": 1, "charge_amount": 215.00, "allowed_amount": 158.50, "paid_amount": 108.50, "adjustment_amount": 56.50,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_RAJESH},
    {"id": uid("cline", "david-newpt-2"),       "claim_id": CLM_DAVID_NEWPT,    "line_number": 2, "cpt_code": "36415", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-12-19", "units": 1, "charge_amount": 25.00,  "allowed_amount": 5.00,   "paid_amount": 4.00,   "adjustment_amount": 20.00,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_RAJESH},
    {"id": uid("cline", "david-newpt-3"),       "claim_id": CLM_DAVID_NEWPT,    "line_number": 3, "cpt_code": "83036", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-12-19", "units": 1, "charge_amount": 55.00,  "allowed_amount": 22.00,  "paid_amount": 18.00,  "adjustment_amount": 33.00,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_RAJESH},
    {"id": uid("cline", "david-newpt-4"),       "claim_id": CLM_DAVID_NEWPT,    "line_number": 4, "cpt_code": "80061", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-12-19", "units": 1, "charge_amount": 75.00,  "allowed_amount": 33.00,  "paid_amount": 26.00,  "adjustment_amount": 42.00,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_RAJESH},
    {"id": uid("cline", "robert-inpt-em-1"),    "claim_id": CLM_ROBERT_INPT,    "line_number": 1, "cpt_code": "99221", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-11-04", "units": 1, "charge_amount": 312.00, "allowed_amount": 215.00, "paid_amount": 215.00, "adjustment_amount": 97.00,  "denial_code": None, "place_of_service": "21", "rendering_provider_id": PROV_LISA},
    {"id": uid("cline", "robert-inpt-em-2"),    "claim_id": CLM_ROBERT_INPT,    "line_number": 2, "cpt_code": "99232", "modifier": None, "icd10_pointer": "1,2",     "service_date": "2024-11-06", "units": 2, "charge_amount": 350.00, "allowed_amount": 240.00, "paid_amount": 240.00, "adjustment_amount": 110.00, "denial_code": None, "place_of_service": "21", "rendering_provider_id": PROV_LISA},
    {"id": uid("cline", "james-inpt-em-1"),     "claim_id": CLM_JAMES_INPT,     "line_number": 1, "cpt_code": "99221", "modifier": None, "icd10_pointer": "1",       "service_date": "2025-01-04", "units": 1, "charge_amount": 312.00, "allowed_amount": None,   "paid_amount": None,   "adjustment_amount": None,   "denial_code": None, "place_of_service": "21", "rendering_provider_id": PROV_AMANDA},
    {"id": uid("cline", "james-inpt-em-2"),     "claim_id": CLM_JAMES_INPT,     "line_number": 2, "cpt_code": "99291", "modifier": None, "icd10_pointer": "1,2",     "service_date": "2025-01-04", "units": 1, "charge_amount": 685.00, "allowed_amount": None,   "paid_amount": None,   "adjustment_amount": None,   "denial_code": None, "place_of_service": "21", "rendering_provider_id": PROV_AMANDA},
    {"id": uid("cline", "james-inpt-em-3"),     "claim_id": CLM_JAMES_INPT,     "line_number": 3, "cpt_code": "99232", "modifier": None, "icd10_pointer": "1,2,3",   "service_date": "2025-01-06", "units": 4, "charge_amount": 700.00, "allowed_amount": None,   "paid_amount": None,   "adjustment_amount": None,   "denial_code": None, "place_of_service": "21", "rendering_provider_id": PROV_AMANDA},
    {"id": uid("cline", "dorothy-tha-1"),       "claim_id": CLM_DOROTHY_INPT,   "line_number": 1, "cpt_code": "27130", "modifier": "RT", "icd10_pointer": "1",       "service_date": "2024-11-19", "units": 1, "charge_amount": 8500.00,"allowed_amount": 5800.00,"paid_amount": 5800.00,"adjustment_amount": 2700.00,"denial_code": None, "place_of_service": "21", "rendering_provider_id": PROV_DANIEL},
    {"id": uid("cline", "dorothy-tha-2"),       "claim_id": CLM_DOROTHY_INPT,   "line_number": 2, "cpt_code": "99221", "modifier": None, "icd10_pointer": "1,2",     "service_date": "2024-11-18", "units": 1, "charge_amount": 312.00, "allowed_amount": 215.00, "paid_amount": 215.00, "adjustment_amount": 97.00,  "denial_code": None, "place_of_service": "21", "rendering_provider_id": PROV_DANIEL},
    {"id": uid("cline", "carlos-ed-em-1"),      "claim_id": CLM_CARLOS_ED,      "line_number": 1, "cpt_code": "99284", "modifier": None, "icd10_pointer": "1,2",     "service_date": "2024-12-28", "units": 1, "charge_amount": 565.00, "allowed_amount": 0.00,   "paid_amount": 0.00,   "adjustment_amount": 565.00, "denial_code": "97",  "place_of_service": "23", "rendering_provider_id": PROV_JOHANNA},
    {"id": uid("cline", "carlos-ed-cta-2"),     "claim_id": CLM_CARLOS_ED,      "line_number": 2, "cpt_code": "71250", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-12-28", "units": 1, "charge_amount": 525.00, "allowed_amount": 0.00,   "paid_amount": 0.00,   "adjustment_amount": 525.00, "denial_code": "197", "place_of_service": "23", "rendering_provider_id": PROV_KATHRYN},
    {"id": uid("cline", "carlos-ed-trop-3"),    "claim_id": CLM_CARLOS_ED,      "line_number": 3, "cpt_code": "84484", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-12-28", "units": 2, "charge_amount": 190.00, "allowed_amount": 0.00,   "paid_amount": 0.00,   "adjustment_amount": 190.00, "denial_code": "97",  "place_of_service": "23", "rendering_provider_id": PROV_JOHANNA},
    {"id": uid("cline", "emily-ed-em-1"),       "claim_id": CLM_EMILY_ED,       "line_number": 1, "cpt_code": "99284", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-12-15", "units": 1, "charge_amount": 565.00, "allowed_amount": 412.50, "paid_amount": 0.00,   "adjustment_amount": 152.50, "denial_code": None, "place_of_service": "23", "rendering_provider_id": PROV_HASAN},
    {"id": uid("cline", "emily-ed-neb-2"),      "claim_id": CLM_EMILY_ED,       "line_number": 2, "cpt_code": "94640", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-12-15", "units": 3, "charge_amount": 195.00, "allowed_amount": 132.00, "paid_amount": 0.00,   "adjustment_amount": 63.00,  "denial_code": None, "place_of_service": "23", "rendering_provider_id": PROV_BRIAN_PA},
    {"id": uid("cline", "aisha-ob-em-1"),       "claim_id": CLM_AISHA_OB,       "line_number": 1, "cpt_code": "99214", "modifier": None, "icd10_pointer": "1",       "service_date": "2025-01-16", "units": 1, "charge_amount": 215.00, "allowed_amount": 110.40, "paid_amount": 110.40, "adjustment_amount": 104.60, "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_SARA},
    {"id": uid("cline", "aisha-ob-imm-2"),      "claim_id": CLM_AISHA_OB,       "line_number": 2, "cpt_code": "90471", "modifier": None, "icd10_pointer": "1",       "service_date": "2025-01-16", "units": 1, "charge_amount": 28.00,  "allowed_amount": 14.00,  "paid_amount": 14.00,  "adjustment_amount": 14.00,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_SARA},
    {"id": uid("cline", "aisha-ob-us-3"),       "claim_id": CLM_AISHA_OB,       "line_number": 3, "cpt_code": "76700", "modifier": "26", "icd10_pointer": "1",       "service_date": "2025-01-16", "units": 1, "charge_amount": 285.00, "allowed_amount": 74.00,  "paid_amount": 74.00,  "adjustment_amount": 211.00, "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_KATHRYN},
    {"id": uid("cline", "dorothy-preop-em-1"),  "claim_id": CLM_DOROTHY_PREOP,  "line_number": 1, "cpt_code": "99214", "modifier": "57", "icd10_pointer": "1,2",     "service_date": "2024-11-26", "units": 1, "charge_amount": 215.00, "allowed_amount": 116.45, "paid_amount": 93.16,  "adjustment_amount": 98.55,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_RAJESH},
    {"id": uid("cline", "dorothy-preop-ekg-2"), "claim_id": CLM_DOROTHY_PREOP,  "line_number": 2, "cpt_code": "93000", "modifier": None, "icd10_pointer": "1",       "service_date": "2024-11-26", "units": 1, "charge_amount": 95.00,  "allowed_amount": 46.00,  "paid_amount": 36.80,  "adjustment_amount": 49.00,  "denial_code": None, "place_of_service": "11", "rendering_provider_id": PROV_RAJESH},
]

PAYMENTS = [
    {"id": uid("pay", "margaret-fu"),     "claim_id": CLM_MARGARET_FU,    "patient_id": P_MARGARET, "payer_id": PAYER_MEDICARE, "payment_type": "insurance","payment_method": "eft",   "amount": 113.74,  "received_at": ts(2025,1,28,14,15,0), "reference_number": "ERA-MCARE-202501-0029","era_835_id": "835-MCARE-2025028-A"},
    {"id": uid("pay", "robert-inpt"),     "claim_id": CLM_ROBERT_INPT,    "patient_id": P_ROBERT,   "payer_id": PAYER_AETNA,    "payment_type": "insurance","payment_method": "eft",   "amount": 22928.75,"received_at": ts(2024,12,5,10,30,0), "reference_number": "ERA-AET-202412-0145",  "era_835_id": "835-AET-2024339-A"},
    {"id": uid("pay", "robert-cardio"),   "claim_id": CLM_ROBERT_CARDIO,  "patient_id": P_ROBERT,   "payer_id": PAYER_AETNA,    "payment_type": "insurance","payment_method": "eft",   "amount": 375.20,  "received_at": ts(2025,1,9,11,15,0),  "reference_number": "ERA-AET-202501-0021",  "era_835_id": "835-AET-2025009-A"},
    {"id": uid("pay", "robert-pat-copay"),"claim_id": CLM_ROBERT_CARDIO,  "patient_id": P_ROBERT,   "payer_id": None,           "payment_type": "patient",  "payment_method": "credit_card","amount": 50.00,"received_at": ts(2024,12,12,8,30,0), "reference_number": "AUTH-238104",          "era_835_id": None},
    {"id": uid("pay", "lakshmi-fu"),      "claim_id": CLM_LAKSHMI_FU,     "patient_id": P_LAKSHMI,  "payer_id": PAYER_BCBS,     "payment_type": "insurance","payment_method": "eft",   "amount": 62.50,   "received_at": ts(2025,1,30,15,0,0),  "reference_number": "ERA-BSC-202501-0312",  "era_835_id": "835-BSC-2025030-A"},
    {"id": uid("pay", "dorothy-inpt"),    "claim_id": CLM_DOROTHY_INPT,   "patient_id": P_DOROTHY,  "payer_id": PAYER_MEDICARE, "payment_type": "insurance","payment_method": "eft",   "amount": 19847.00,"received_at": ts(2024,12,18,9,15,0), "reference_number": "ERA-MCARE-202412-0018","era_835_id": "835-MCARE-2024353-A"},
    {"id": uid("pay", "david-newpt"),     "claim_id": CLM_DAVID_NEWPT,    "patient_id": P_DAVID,    "payer_id": PAYER_AETNA,    "payment_type": "insurance","payment_method": "eft",   "amount": 168.50,  "received_at": ts(2025,1,8,11,20,0),  "reference_number": "ERA-AET-202501-0020",  "era_835_id": "835-AET-2025008-A"},
    {"id": uid("pay", "david-pat-ded"),   "claim_id": CLM_DAVID_NEWPT,    "patient_id": P_DAVID,    "payer_id": None,           "payment_type": "patient",  "payment_method": "credit_card","amount": 50.00,"received_at": ts(2025,1,12,16,0,0),  "reference_number": "AUTH-251847",          "era_835_id": None},
    {"id": uid("pay", "marcus-lbp"),      "claim_id": CLM_MARCUS_LBP,     "patient_id": P_MARCUS,   "payer_id": PAYER_UHC,      "payment_type": "insurance","payment_method": "eft",   "amount": 53.40,   "received_at": ts(2025,1,26,13,0,0),  "reference_number": "ERA-UHC-202501-0228",  "era_835_id": "835-UHC-2025026-A"},
    {"id": uid("pay", "aisha-ob"),        "claim_id": CLM_AISHA_OB,       "patient_id": P_AISHA,    "payer_id": PAYER_MEDICAL,  "payment_type": "insurance","payment_method": "eft",   "amount": 198.40,  "received_at": ts(2025,2,7,10,0,0),   "reference_number": "ERA-MCAL-202502-0045", "era_835_id": "835-MCAL-2025038-A"},
    {"id": uid("pay", "dorothy-preop"),   "claim_id": CLM_DOROTHY_PREOP,  "patient_id": P_DOROTHY,  "payer_id": PAYER_MEDICARE, "payment_type": "insurance","payment_method": "eft",   "amount": 129.96,  "received_at": ts(2024,12,15,11,0,0), "reference_number": "ERA-MCARE-202412-0009","era_835_id": "835-MCARE-2024350-A"},
]

ADJUSTMENTS = [
    {"id": uid("adj", "margaret-fu-1"),       "claim_line_id": uid("cline", "margaret-fu-1"),     "adjustment_group": "CO",  "reason_code": "45",  "amount": 107.82, "note": "Charge exceeds fee schedule (Medicare allowed)"},
    {"id": uid("adj", "margaret-fu-1-cb"),    "claim_line_id": uid("cline", "margaret-fu-1"),     "adjustment_group": "PR",  "reason_code": "2",   "amount": 21.44,  "note": "Coinsurance (20% of allowed amount)"},
    {"id": uid("adj", "margaret-fu-2"),       "claim_line_id": uid("cline", "margaret-fu-2"),     "adjustment_group": "CO",  "reason_code": "45",  "amount": 20.00,  "note": "Charge exceeds fee schedule"},
    {"id": uid("adj", "margaret-fu-3"),       "claim_line_id": uid("cline", "margaret-fu-3"),     "adjustment_group": "CO",  "reason_code": "45",  "amount": 35.00,  "note": "Charge exceeds fee schedule"},
    {"id": uid("adj", "robert-cardio-1"),     "claim_line_id": uid("cline", "robert-cardio-1"),   "adjustment_group": "CO",  "reason_code": "45",  "amount": 39.80,  "note": "Charge exceeds Aetna contracted rate"},
    {"id": uid("adj", "robert-cardio-1-pr"),  "claim_line_id": uid("cline", "robert-cardio-1"),   "adjustment_group": "PR",  "reason_code": "3",   "amount": 50.00,  "note": "Specialist copay"},
    {"id": uid("adj", "robert-cardio-2"),     "claim_line_id": uid("cline", "robert-cardio-2"),   "adjustment_group": "CO",  "reason_code": "45",  "amount": 175.00, "note": "Modifier 26 - professional component only allowed"},
    {"id": uid("adj", "lakshmi-fu-1"),        "claim_line_id": uid("cline", "lakshmi-fu-1"),     "adjustment_group": "CO",  "reason_code": "45",  "amount": 52.50,  "note": "Charge exceeds BSC contracted rate"},
    {"id": uid("adj", "lakshmi-fu-1-pr"),     "claim_line_id": uid("cline", "lakshmi-fu-1"),     "adjustment_group": "PR",  "reason_code": "3",   "amount": 30.00,  "note": "PCP copay per BSC plan"},
    {"id": uid("adj", "marcus-lbp-1"),        "claim_line_id": uid("cline", "marcus-lbp-1"),     "adjustment_group": "CO",  "reason_code": "45",  "amount": 56.60,  "note": "Charge exceeds UHC contracted rate"},
    {"id": uid("adj", "marcus-lbp-1-pr"),     "claim_line_id": uid("cline", "marcus-lbp-1"),     "adjustment_group": "PR",  "reason_code": "3",   "amount": 35.00,  "note": "UHC PCP copay"},
    {"id": uid("adj", "david-newpt-1"),       "claim_line_id": uid("cline", "david-newpt-1"),    "adjustment_group": "CO",  "reason_code": "45",  "amount": 56.50,  "note": "Charge exceeds Aetna fee schedule"},
    {"id": uid("adj", "david-newpt-1-pr"),    "claim_line_id": uid("cline", "david-newpt-1"),    "adjustment_group": "PR",  "reason_code": "1",   "amount": 50.00,  "note": "Patient deductible portion"},
    {"id": uid("adj", "carlos-ed-1-deny"),    "claim_line_id": uid("cline", "carlos-ed-em-1"),   "adjustment_group": "CO",  "reason_code": "97",  "amount": 565.00, "note": "Out-of-network ED claim - not a Kaiser facility"},
    {"id": uid("adj", "carlos-ed-2-deny"),    "claim_line_id": uid("cline", "carlos-ed-cta-2"),  "adjustment_group": "CO",  "reason_code": "197", "amount": 525.00, "note": "Service not authorized per Kaiser referral requirement"},
    {"id": uid("adj", "carlos-ed-3-deny"),    "claim_line_id": uid("cline", "carlos-ed-trop-3"), "adjustment_group": "CO",  "reason_code": "97",  "amount": 190.00, "note": "Out-of-network lab services"},
    {"id": uid("adj", "emily-ed-1-pr"),       "claim_line_id": uid("cline", "emily-ed-em-1"),    "adjustment_group": "PR",  "reason_code": "1",   "amount": 412.50, "note": "Applied to HDHP deductible (patient owes full allowed)"},
    {"id": uid("adj", "emily-ed-2-pr"),       "claim_line_id": uid("cline", "emily-ed-neb-2"),   "adjustment_group": "PR",  "reason_code": "1",   "amount": 132.00, "note": "Applied to HDHP deductible"},
    {"id": uid("adj", "dorothy-tha-1-co"),    "claim_line_id": uid("cline", "dorothy-tha-1"),    "adjustment_group": "CO",  "reason_code": "45",  "amount": 2700.00,"note": "Medicare allowed = MS-DRG 470 rate"},
    {"id": uid("adj", "dorothy-tha-1-pr"),    "claim_line_id": uid("cline", "dorothy-tha-1"),    "adjustment_group": "PR",  "reason_code": "2",   "amount": 1573.00,"note": "Part A inpatient deductible 2024"},
]

CLAIM_DENIALS = [
    {"id": uid("dn", "carlos-ed"),  "claim_id": CLM_CARLOS_ED, "denial_date": "2025-01-08", "carc_code": "97",  "rarc_code": "N130", "denial_reason": "The benefit for this service is included in the payment/allowance for another service that has been adjudicated. Out-of-network emergency services not eligible for OON benefit on this plan tier.","is_appealable": True, "appeal_deadline": "2025-04-08", "worked_by_provider_id": None, "status": "appealed"},
]

CLAIM_APPEALS = [
    {"id": uid("ap", "carlos-ed-l1"), "claim_denial_id": uid("dn", "carlos-ed"), "appeal_level": 1, "submitted_at": ts(2025,1,21,11,0,0), "submitted_by_provider_id": None, "narrative": "Patient presented to Springfield Regional ED with acute chest pain, palpitations, dyspnea. Symptoms required emergent evaluation under the Prudent Layperson standard. Kaiser member services agent (Mara Johnson, ref #KP-2024-INT-0918211) advised patient to seek nearest ED on 2024-12-28 at 21:08. Request reprocessing as in-network emergency.","outcome": "pending","decided_at": None, "recovered_amount": None},
]

PATIENT_STATEMENTS = [
    {"id": uid("ps", "margaret-2025-01"),"patient_id": P_MARGARET, "statement_date": "2025-02-01", "period_start": "2025-01-01", "period_end": "2025-01-31", "previous_balance": 0.00,    "charges_total": 285.00, "payments_total": 113.74, "adjustments_total": 142.82, "current_balance": 28.44,  "due_date": "2025-03-01", "status": "sent",    "delivery_method": "mail"},
    {"id": uid("ps", "robert-2024-12"),  "patient_id": P_ROBERT,   "statement_date": "2025-01-15", "period_start": "2024-11-01", "period_end": "2024-12-31", "previous_balance": 0.00,    "charges_total": 39060.00,"payments_total": 23354.50,"adjustments_total": 14047.00,"current_balance": 1658.50,"due_date": "2025-02-15", "status": "sent",    "delivery_method": "email"},
    {"id": uid("ps", "emily-2024-12"),   "patient_id": P_EMILY,    "statement_date": "2025-01-20", "period_start": "2024-12-01", "period_end": "2024-12-31", "previous_balance": 0.00,    "charges_total": 760.00, "payments_total": 0.00,   "adjustments_total": 0.00,    "current_balance": 1248.50,"due_date": "2025-02-20", "status": "sent",    "delivery_method": "email"},
    {"id": uid("ps", "dorothy-2024-12"), "patient_id": P_DOROTHY,  "statement_date": "2025-01-05", "period_start": "2024-11-01", "period_end": "2024-12-31", "previous_balance": 32.49,   "charges_total": 33160.00,"payments_total": 19976.96,"adjustments_total": 11643.03,"current_balance": 1605.49,"due_date": "2025-02-05", "status": "sent",    "delivery_method": "mail"},
    {"id": uid("ps", "carlos-2024-12"),  "patient_id": P_CARLOS,   "statement_date": "2025-01-10", "period_start": "2024-12-01", "period_end": "2024-12-31", "previous_balance": 0.00,    "charges_total": 1280.00,"payments_total": 0.00,   "adjustments_total": 0.00,    "current_balance": 4285.00,"due_date": "2025-02-10", "status": "appeal_in_process", "delivery_method": "mail"},
    {"id": uid("ps", "david-2024-12"),   "patient_id": P_DAVID,    "statement_date": "2025-01-15", "period_start": "2024-12-01", "period_end": "2024-12-31", "previous_balance": 0.00,    "charges_total": 425.00, "payments_total": 218.50, "adjustments_total": 156.50, "current_balance": 0.00,   "due_date": "2025-02-15", "status": "paid",    "delivery_method": "email"},
]

# ---------------------------------------------------------------------------
# 11. COMMUNICATIONS
# ---------------------------------------------------------------------------

PMT_MARGARET = uid("thr", "margaret-portal")
PMT_ROBERT   = uid("thr", "robert-portal")
PMT_LAKSHMI  = uid("thr", "lakshmi-portal")
PMT_DAVID    = uid("thr", "david-portal")
PMT_MARCUS   = uid("thr", "marcus-portal")
PMT_AISHA    = uid("thr", "aisha-portal")

PATIENT_MESSAGE_THREADS = [
    {"id": PMT_MARGARET,"patient_id": P_MARGARET,"subject": "Lab results from 1/13 visit",          "category": "lab_result",     "status": "resolved",   "priority": "normal","assigned_provider_id": PROV_MORGAN, "last_message_at": ts(2025,1,15,14,30,0)},
    {"id": PMT_ROBERT,  "patient_id": P_ROBERT,  "subject": "Question about ticagrelor side effects","category": "medication",    "status": "resolved",   "priority": "normal","assigned_provider_id": PROV_LISA,   "last_message_at": ts(2024,11,22,15,0,0)},
    {"id": PMT_LAKSHMI, "patient_id": P_LAKSHMI, "subject": "Refill albuterol inhaler",             "category": "refill_request", "status": "resolved",   "priority": "normal","assigned_provider_id": PROV_MORGAN, "last_message_at": ts(2025,1,18,11,30,0)},
    {"id": PMT_DAVID,   "patient_id": P_DAVID,   "subject": "Metformin upset stomach",              "category": "medication",     "status": "open",       "priority": "normal","assigned_provider_id": PROV_RAJESH, "last_message_at": ts(2025,1,12,9,45,0)},
    {"id": PMT_MARCUS,  "patient_id": P_MARCUS,  "subject": "Request physical therapy referral",    "category": "referral",       "status": "in_progress","priority": "normal","assigned_provider_id": PROV_MORGAN, "last_message_at": ts(2025,1,14,10,0,0)},
    {"id": PMT_AISHA,   "patient_id": P_AISHA,   "subject": "Insurance question - WIC eligibility", "category": "billing",        "status": "open",       "priority": "low",   "assigned_provider_id": None,        "last_message_at": ts(2025,1,20,13,15,0)},
]

PATIENT_MESSAGES = [
    {"id": uid("msg","margaret-1"),"thread_id": PMT_MARGARET,"channel": "portal","direction": "outgoing","sender_provider_id": PROV_MORGAN,"sender_patient_id": None,    "body": "Good morning Margaret — your A1c on 1/13 came back at 7.4%, slight improvement from 7.8% in November. eGFR steady at 48. Please continue current meds; we'll recheck in 3 months at your visit on 4/14. Reply with any questions.",                "sent_at": ts(2025,1,15,9,15,0),   "read_at": ts(2025,1,15,11,30,0),"attachment_uri": "/secure-portal/results/margaret-2025-01-13-cmp.pdf"},
    {"id": uid("msg","margaret-2"),"thread_id": PMT_MARGARET,"channel": "portal","direction": "incoming","sender_provider_id": None,        "sender_patient_id": P_MARGARET,"body": "Thank you Dr. Blackwell. Will I need to keep checking my sugars at home? My finger sticks have been 130-160 fasting.",                                                                                                                          "sent_at": ts(2025,1,15,12,45,0),  "read_at": ts(2025,1,15,14,0,0), "attachment_uri": None},
    {"id": uid("msg","margaret-3"),"thread_id": PMT_MARGARET,"channel": "portal","direction": "outgoing","sender_provider_id": PROV_MORGAN,"sender_patient_id": None,    "body": "Yes please continue 1-2x daily for now. Fasting 130-160 is OK but we want under 130. Your night-time metformin is doing its job — keep it up.",                                                                                            "sent_at": ts(2025,1,15,14,30,0),  "read_at": ts(2025,1,16,8,0,0),  "attachment_uri": None},
    {"id": uid("msg","robert-1"),  "thread_id": PMT_ROBERT,  "channel": "portal","direction": "incoming","sender_provider_id": None,        "sender_patient_id": P_ROBERT,"body": "Dr. Chen, I've noticed easier bruising and a small nosebleed last week since starting Brilinta. Is this normal? Should I be worried?",                                                                                                  "sent_at": ts(2024,11,20,16,20,0), "read_at": ts(2024,11,21,7,45,0),"attachment_uri": None},
    {"id": uid("msg","robert-2"),  "thread_id": PMT_ROBERT,  "channel": "portal","direction": "outgoing","sender_provider_id": PROV_LISA,   "sender_patient_id": None,    "body": "Hello Robert — minor bruising and brief epistaxis are expected on dual antiplatelet therapy (aspirin + ticagrelor). Continue both. Call us or come in if: bleeding > 20 min not stopping, black/tarry stools, severe headache, or any bleeding requiring medical attention. We'll discuss further at your 12/12 visit.","sent_at": ts(2024,11,21,9,30,0),  "read_at": ts(2024,11,21,12,0,0),"attachment_uri": None},
    {"id": uid("msg","robert-3"),  "thread_id": PMT_ROBERT,  "channel": "portal","direction": "incoming","sender_provider_id": None,        "sender_patient_id": P_ROBERT,"body": "Understood. Thank you.",                                                                                                                                                                                                                       "sent_at": ts(2024,11,22,15,0,0),  "read_at": ts(2024,11,22,16,30,0),"attachment_uri": None},
    {"id": uid("msg","lakshmi-1"), "thread_id": PMT_LAKSHMI, "channel": "portal","direction": "incoming","sender_provider_id": None,        "sender_patient_id": P_LAKSHMI,"body": "Need to refill my albuterol rescue inhaler — I'm down to last 30 puffs. Pollen season starting.",                                                                                                                                            "sent_at": ts(2025,1,17,8,0,0),    "read_at": ts(2025,1,17,9,30,0), "attachment_uri": None},
    {"id": uid("msg","lakshmi-2"), "thread_id": PMT_LAKSHMI, "channel": "portal","direction": "outgoing","sender_provider_id": PROV_MORGAN, "sender_patient_id": None,    "body": "Refill sent to CVS Springfield Main. 1 inhaler, 2 refills authorized. Should be ready in 2-3 hours. — Jenny (Dr. Blackwell's MA)",                                                                                                            "sent_at": ts(2025,1,17,11,15,0),  "read_at": ts(2025,1,17,17,0,0), "attachment_uri": None},
    {"id": uid("msg","lakshmi-3"), "thread_id": PMT_LAKSHMI, "channel": "portal","direction": "incoming","sender_provider_id": None,        "sender_patient_id": P_LAKSHMI,"body": "Picked up — thanks!",                                                                                                                                                                                                                          "sent_at": ts(2025,1,18,11,30,0),  "read_at": None,                  "attachment_uri": None},
    {"id": uid("msg","david-1"),   "thread_id": PMT_DAVID,   "channel": "portal","direction": "incoming","sender_provider_id": None,        "sender_patient_id": P_DAVID, "body": "Dr. Iyer, the metformin is upsetting my stomach pretty bad — nausea and diarrhea most mornings. I'm taking it with breakfast like you said. Should I stop?",                                                                                "sent_at": ts(2025,1,12,9,45,0),   "read_at": ts(2025,1,13,7,30,0), "attachment_uri": None},
    {"id": uid("msg","marcus-1"),  "thread_id": PMT_MARCUS,  "channel": "portal","direction": "incoming","sender_provider_id": None,        "sender_patient_id": P_MARCUS,"body": "Hi Dr. Blackwell — the back pain isn't getting better with stretches and meloxicam. Can I get a PT referral? My health insurance covers 20 visits a year.",                                                                              "sent_at": ts(2025,1,13,14,0,0),   "read_at": ts(2025,1,14,8,30,0), "attachment_uri": None},
    {"id": uid("msg","marcus-2"),  "thread_id": PMT_MARCUS,  "channel": "portal","direction": "outgoing","sender_provider_id": PROV_MORGAN, "sender_patient_id": None,    "body": "Of course — submitting referral to Springfield PT Associates today, Dr. Singh. Their office will call you within 2-3 business days. Continue meloxicam meanwhile, ice 20 min twice daily.",                                                "sent_at": ts(2025,1,14,10,0,0),   "read_at": ts(2025,1,14,11,30,0),"attachment_uri": None},
    {"id": uid("msg","aisha-1"),   "thread_id": PMT_AISHA,   "channel": "portal","direction": "incoming","sender_provider_id": None,        "sender_patient_id": P_AISHA, "body": "Hello — I was told I might qualify for WIC food assistance with my pregnancy. Does the clinic help with this paperwork or do I need to go somewhere else?",                                                                                  "sent_at": ts(2025,1,20,13,15,0),  "read_at": None,                  "attachment_uri": None},
]

CALL_LOGS = [
    {"id": uid("call","margaret-confirm"),"patient_id": P_MARGARET,"provider_id": None,         "direction": "outbound","phone_number": "+1-559-555-0182","started_at": ts(2025,1,11,14,30,0), "duration_seconds": 92,  "outcome": "completed", "reason": "appt_confirmation",  "summary": "Confirmed 1/13 appt. Patient asked about fasting requirements for labs. Confirmed CMP fasting recommended.","follow_up_required": False},
    {"id": uid("call","robert-postdc"),   "patient_id": P_ROBERT,  "provider_id": None,         "direction": "outbound","phone_number": "+1-650-555-0291","started_at": ts(2024,11,8,10,15,0),  "duration_seconds": 387, "outcome": "completed", "reason": "post_discharge",     "summary": "RN Tasha called for transitional care call. Pt taking all 7 d/c meds correctly. Has scale, no weight gain >2 lbs. F/u cards 12/12 confirmed. Brought up cost concern re: ticagrelor — patient enrolled in Brilinta savings card via Aetna.","follow_up_required": False},
    {"id": uid("call","carlos-followup"), "patient_id": P_CARLOS,  "provider_id": PROV_JOHANNA, "direction": "outbound","phone_number": "+1-559-555-0337","started_at": ts(2024,12,30,14,0,0),  "duration_seconds": 423, "outcome": "completed", "reason": "ed_followup",        "summary": "Patient anxious about ED workup. Reviewed normal troponin x2, normal CTA chest, normal EKG. Discussed health anxiety component. Encouraged f/u with PCP Dr. Singh (referral provided in ED). Patient verbalized understanding.","follow_up_required": True},
    {"id": uid("call","dorothy-daughter"),"patient_id": P_DOROTHY, "provider_id": None,         "direction": "inbound", "phone_number": "+1-916-555-2074","started_at": ts(2025,1,15,13,20,0),  "duration_seconds": 612, "outcome": "completed", "reason": "family_inquiry",     "summary": "Daughter (Sarah Kim, POA on file) asking about donepezil dose and PT schedule. Verified POA. Confirmed donepezil 10mg PO QHS. PT 3x/week Mon/Wed/Fri 0900 at SRMC Rehab.","follow_up_required": False},
    {"id": uid("call","james-noshow"),    "patient_id": P_JAMES,   "provider_id": None,         "direction": "outbound","phone_number": "+1-559-555-0418","started_at": ts(2025,1,20,15,45,0),  "duration_seconds": 0,   "outcome": "voicemail", "reason": "missed_appointment", "summary": "No answer. Left voicemail re: missed pulmonology f/u 1/20. Will retry tomorrow.","follow_up_required": True},
    {"id": uid("call","emily-erfu"),      "patient_id": P_EMILY,   "provider_id": None,         "direction": "outbound","phone_number": "+1-559-555-0822","started_at": ts(2024,12,17,11,30,0), "duration_seconds": 245, "outcome": "completed", "reason": "ed_followup",        "summary": "RN Tasha called. Pt reports back to baseline. No nighttime sx since d/c. Picking up budesonide-formoterol today. PCP appt scheduled 12/26.","follow_up_required": False},
]

INSURANCE_CORRESPONDENCES = [
    {"id": uid("ins-corr","robert-echo-auth"),    "payer_id": PAYER_AETNA,    "patient_id": P_ROBERT,   "claim_id": None,             "authorization_id": AUTH_ROBERT_ECHO,             "direction": "inbound", "correspondence_type": "auth_approval","channel": "edi_278","received_at": ts(2024,11,14,10,15,0),"sent_at": None,                 "subject": "Prior authorization approved: 93306 Echocardiogram",                  "body": "Prior auth 2024-11-12 for 93306 TTE w/ Doppler and color flow approved through 2024-12-31. Approved provider: SRMC Cardiology. Member: Robert W. Chen. Auth ref: AET-PA-2024-892014.","document_uri": "/secure/payer-corr/aet/AET-PA-2024-892014.pdf","requires_follow_up": False,"follow_up_by": None},
    {"id": uid("ins-corr","james-hh-denial"),     "payer_id": PAYER_MEDICARE, "patient_id": P_JAMES,    "claim_id": None,             "authorization_id": None,                         "direction": "inbound", "correspondence_type": "auth_denial",  "channel": "mail",   "received_at": ts(2025,1,16,14,0,0), "sent_at": None,                 "subject": "Home oxygen continuation - additional documentation required",        "body": "Continuation of home O2 LPM 2 requires recent ABG or SpO2 < 88% on RA. Please submit room air SpO2 from ED 1/4 visit and 6MWT if available. 30-day deadline to respond before discontinuation. Ref: MCARE-PA-2025-001873.","document_uri": "/secure/payer-corr/mcare/MCARE-PA-2025-001873.pdf","requires_follow_up": True,"follow_up_by": "2025-02-15"},
    {"id": uid("ins-corr","james-hh-response"),   "payer_id": PAYER_MEDICARE, "patient_id": P_JAMES,    "claim_id": None,             "authorization_id": None,                         "direction": "outbound","correspondence_type": "auth_appeal",  "channel": "fax",    "received_at": None,                 "sent_at": ts(2025,1,17,11,30,0),"subject": "Re: Home O2 continuation - documentation submitted",                "body": "Submitted: ABG 1/9 7.39/52/68/30 on 2L NC. SpO2 room air during 6MWT 86% at rest, dropped to 82% at 2 min. Continued O2 medically necessary per CMS NCD 240.2. Please reauthorize.","document_uri": "/secure/payer-corr/srmc-out/srmc-james-o2-appeal-2025-01-17.pdf","requires_follow_up": True,"follow_up_by": "2025-02-17"},
    {"id": uid("ins-corr","dorothy-tha-auth"),    "payer_id": PAYER_MEDICARE, "patient_id": P_DOROTHY,  "claim_id": None,             "authorization_id": AUTH_DOROTHY_THA,             "direction": "inbound", "correspondence_type": "auth_approval","channel": "edi_278","received_at": ts(2024,11,15,11,30,0),"sent_at": None,                 "subject": "Right total hip arthroplasty - 27130 - approved",                     "body": "Procedure authorized. MS-DRG 470 inpatient. Bundled payment includes 90-day post-op care. Ref MCARE-PA-2024-009124.","document_uri": "/secure/payer-corr/mcare/MCARE-PA-2024-009124.pdf","requires_follow_up": False,"follow_up_by": None},
    {"id": uid("ins-corr","carlos-deny"),         "payer_id": PAYER_KAISER,   "patient_id": P_CARLOS,   "claim_id": CLM_CARLOS_ED,    "authorization_id": None,                         "direction": "inbound", "correspondence_type": "claim_denial", "channel": "edi_835","received_at": ts(2025,1,8,9,45,0),  "sent_at": None,                 "subject": "Claim SRMC-CLM-2024-010237 denied - out of network",                  "body": "Services received at non-Kaiser facility 12/28/2024. Member may file appeal within 90 days. Member services: 1-800-464-4000. Denial ref: KP-DENIAL-2025-008217.","document_uri": "/secure/payer-corr/kp/KP-DENIAL-2025-008217.pdf","requires_follow_up": True,"follow_up_by": "2025-04-08"},
    {"id": uid("ins-corr","margaret-eligibility"),"payer_id": PAYER_MEDICAL,  "patient_id": P_MARGARET, "claim_id": None,             "authorization_id": None,                         "direction": "inbound", "correspondence_type": "eligibility",  "channel": "mail",   "received_at": ts(2024,12,18,13,0,0),"sent_at": None,                 "subject": "Annual Medi-Cal eligibility re-determination required",               "body": "Member Margaret Johnson eligibility expires 4/30/2025. Re-determination form sent. Return signed packet by 3/15/2025. Ref MCAL-ELIG-2024-882013.","document_uri": "/secure/payer-corr/mcal/MCAL-ELIG-2024-882013.pdf","requires_follow_up": True,"follow_up_by": "2025-03-15"},
]

INTER_PROVIDER_MESSAGES = [
    {"id": uid("ipm","margaret-nephro-cons"),"from_provider_id": PROV_MORGAN, "to_provider_id": PROV_KOFI,  "patient_id": P_MARGARET,"encounter_id": ENC_MARGARET_FU,  "message_type": "consult_request","subject": "Margaret Johnson - CKD3 referral",         "body": "Referring 68F with T2DM/HTN/CKD3 - eGFR trending 52→48→48 over 12 months. Stable on lisinopril 20mg, metformin reduced to 1000mg BID per CrCl. UACR 320 mg/g. Please evaluate for CKD optimization and possible SGLT2i addition. Thanks.","sent_at": ts(2025,1,13,14,30,0), "read_at": ts(2025,1,13,16,0,0), "acknowledged_at": ts(2025,1,14,8,0,0)},
    {"id": uid("ipm","robert-cards-cons"),   "from_provider_id": PROV_HASAN,  "to_provider_id": PROV_LISA,  "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_ED,    "message_type": "stat_consult",   "subject": "NSTEMI - urgent cards consult",            "body": "54M w/ 4hr substernal CP. Initial trop 0.42, repeat 1.85. EKG: ST depressions V4-V6. BP 152/96 HR 88. Loaded ASA 325 chewed, ticagrelor 180mg. On heparin gtt. Pls advise re: cath timing.","sent_at": ts(2024,11,4,19,15,0),  "read_at": ts(2024,11,4,19,18,0), "acknowledged_at": ts(2024,11,4,19,30,0)},
    {"id": uid("ipm","dorothy-ortho-cons"),  "from_provider_id": PROV_JOHANNA,"to_provider_id": PROV_DANIEL,"patient_id": P_DOROTHY, "encounter_id": ENC_DOROTHY_ED,   "message_type": "admit_request",  "subject": "R intertroch hip fx - admission",          "body": "81F with mech fall at home, R groin pain w/ inability to bear weight. XR confirms R intertrochanteric hip fx. Hx Alzheimer's, on donepezil. Daughter at bedside (POA). Pls admit ortho.","sent_at": ts(2024,11,18,17,45,0), "read_at": ts(2024,11,18,17,48,0),"acknowledged_at": ts(2024,11,18,18,0,0)},
    {"id": uid("ipm","james-pcp-discharge"), "from_provider_id": PROV_AMANDA, "to_provider_id": PROV_RAJESH,"patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,   "message_type": "discharge_summary","subject": "James O'Connor - discharge summary",        "body": "73M, GOLD-D COPD, d/c 1/9 after 5-day admission for severe exacerbation. Required NIV (BiPAP 14/6) x36hr. Treated w/ IV methylpred → PO prednisone taper, azithromycin x5d. Home O2 continued 2L NC. ABG at d/c: 7.39/52/68/30 on 2L. Please see in 5 days. F/u w/ me 1/20. Thanks Raj.","sent_at": ts(2025,1,9,15,0,0),  "read_at": ts(2025,1,9,18,30,0), "acknowledged_at": None},
    {"id": uid("ipm","emily-asthma-pcp"),    "from_provider_id": PROV_HASAN,  "to_provider_id": PROV_MORGAN,"patient_id": P_EMILY,   "encounter_id": ENC_EMILY_ED,     "message_type": "ed_followup",    "subject": "Emily Watson ED visit - asthma exacerbation","body": "24F college student, severe asthma exacerbation 12/15. Required 3x nebs in ED, IV methylpred 60mg, d/c on prednisone taper + step-up to budesonide-formoterol 160/4.5 BID. Triggers: dorm carpet, stress (finals). Pls reinforce action plan at f/u.","sent_at": ts(2024,12,16,2,0,0), "read_at": ts(2024,12,16,8,30,0), "acknowledged_at": None},
]

# ---------------------------------------------------------------------------
# 12. OPERATIONS — staff schedules, on-call, pharmacy inventory, equipment, tasks, audit
# ---------------------------------------------------------------------------

SHIFT_DAY     = uid("shift", "day")
SHIFT_NIGHT   = uid("shift", "night")
SHIFT_EVENING = uid("shift", "evening")
SHIFT_OFFICE  = uid("shift", "office")

def sched_id(prov_key: str, date_str: str) -> str:
    return uid("sched", f"{prov_key}-{date_str}")

STAFF_SCHEDULES = []
# Week of 2025-01-13 (Mon) to 2025-01-19 (Sun) — clinic + ED + ICU coverage
_week_clinic = [
    (PROV_MORGAN,  "morgan", DEPT_FAMILY, None, SHIFT_OFFICE, "physician", ["2025-01-13","2025-01-14","2025-01-15","2025-01-16","2025-01-17"], 8, 17),
    (PROV_RAJESH,  "rajesh", DEPT_IM,     None, SHIFT_OFFICE, "physician", ["2025-01-13","2025-01-14","2025-01-15","2025-01-16","2025-01-17"], 8, 17),
    (PROV_HARPER,  "harper", DEPT_IM,     None, SHIFT_OFFICE, "physician", ["2025-01-14","2025-01-16"], 8, 17),
    (PROV_LISA,    "lisa",   DEPT_CARDIO, None, SHIFT_OFFICE, "physician", ["2025-01-13","2025-01-15","2025-01-17"], 8, 17),
    (PROV_TYRELL,  "tyrell", DEPT_CARDIO, None, SHIFT_OFFICE, "physician", ["2025-01-14","2025-01-16"], 8, 17),
    (PROV_SARA,    "sara",   DEPT_OBGYN,  None, SHIFT_OFFICE, "physician", ["2025-01-13","2025-01-14","2025-01-15","2025-01-16"], 8, 17),
    (PROV_KOFI,    "kofi",   DEPT_MEDSURG,None, SHIFT_OFFICE, "physician", ["2025-01-14","2025-01-16"], 9, 17),
]
for prov, key, dept, unit, shift, role, dates, sh, eh in _week_clinic:
    for d in dates:
        y, m, dd = [int(x) for x in d.split("-")]
        STAFF_SCHEDULES.append({"id": sched_id(key, d),"provider_id": prov,"department_id": dept,"unit_id": unit,"shift_id": shift,"work_date": d,"scheduled_start": ts(y,m,dd,sh,0,0),"scheduled_end": ts(y,m,dd,eh,0,0),"role": role,"status": "completed" if d <= "2025-01-17" else "scheduled"})

# ED 24/7 coverage 2025-01-13..2025-01-17
_ed_coverage = [
    ("2025-01-13", PROV_HASAN,   "hasan",   SHIFT_DAY,     7, 19, "attending"),
    ("2025-01-13", PROV_JOHANNA, "johanna", SHIFT_NIGHT,  19, 7,  "attending"),
    ("2025-01-13", PROV_BRIAN_PA,"brian",   SHIFT_DAY,    11, 23, "midlevel"),
    ("2025-01-14", PROV_JOHANNA, "johanna", SHIFT_DAY,     7, 19, "attending"),
    ("2025-01-14", PROV_HASAN,   "hasan",   SHIFT_NIGHT,  19, 7,  "attending"),
    ("2025-01-15", PROV_HASAN,   "hasan",   SHIFT_DAY,     7, 19, "attending"),
    ("2025-01-15", PROV_JOHANNA, "johanna", SHIFT_NIGHT,  19, 7,  "attending"),
    ("2025-01-16", PROV_JOHANNA, "johanna", SHIFT_DAY,     7, 19, "attending"),
    ("2025-01-16", PROV_HASAN,   "hasan",   SHIFT_NIGHT,  19, 7,  "attending"),
    ("2025-01-17", PROV_HASAN,   "hasan",   SHIFT_DAY,     7, 19, "attending"),
    ("2025-01-17", PROV_JOHANNA, "johanna", SHIFT_NIGHT,  19, 7,  "attending"),
]
for d, prov, key, shift, sh, eh, role in _ed_coverage:
    y, m, dd = [int(x) for x in d.split("-")]
    end_y, end_m, end_dd, end_h = (y, m, dd+1, eh) if shift == SHIFT_NIGHT else (y, m, dd, eh)
    STAFF_SCHEDULES.append({"id": sched_id(f"{key}-ed-{shift[:6]}", d),"provider_id": prov,"department_id": DEPT_ED,"unit_id": UNIT_ED_ACUTE,"shift_id": shift,"work_date": d,"scheduled_start": ts(y,m,dd,sh,0,0),"scheduled_end": ts(end_y,end_m,end_dd,end_h,0,0),"role": role,"status": "completed"})

# Hospitalist coverage MICU + medsurg
_hospitalist = [
    ("2025-01-13", PROV_PRIYA, "priya", DEPT_MEDSURG, UNIT_MS_3W, SHIFT_DAY,    7, 19, "attending"),
    ("2025-01-13", PROV_NATE,  "nate",  DEPT_MEDSURG, UNIT_MS_3W, SHIFT_NIGHT, 19,  7, "attending"),
    ("2025-01-14", PROV_PRIYA, "priya", DEPT_MEDSURG, UNIT_MS_3W, SHIFT_DAY,    7, 19, "attending"),
    ("2025-01-15", PROV_PRIYA, "priya", DEPT_MEDSURG, UNIT_MS_3W, SHIFT_DAY,    7, 19, "attending"),
    ("2025-01-16", PROV_NATE,  "nate",  DEPT_MEDSURG, UNIT_MS_3W, SHIFT_DAY,    7, 19, "attending"),
    ("2025-01-17", PROV_NATE,  "nate",  DEPT_MEDSURG, UNIT_MS_3W, SHIFT_DAY,    7, 19, "attending"),
    ("2025-01-13", PROV_AMANDA,"amanda",DEPT_MEDSURG, UNIT_MICU,  SHIFT_DAY,    7, 19, "attending"),
    ("2025-01-14", PROV_AMANDA,"amanda",DEPT_MEDSURG, UNIT_MICU,  SHIFT_DAY,    7, 19, "attending"),
    ("2025-01-15", PROV_AMANDA,"amanda",DEPT_MEDSURG, UNIT_MICU,  SHIFT_DAY,    7, 19, "attending"),
]
for d, prov, key, dept, unit, shift, sh, eh, role in _hospitalist:
    y, m, dd = [int(x) for x in d.split("-")]
    end_y, end_m, end_dd, end_h = (y, m, dd+1, eh) if shift == SHIFT_NIGHT else (y, m, dd, eh)
    STAFF_SCHEDULES.append({"id": sched_id(f"{key}-{shift[:6]}", d),"provider_id": prov,"department_id": dept,"unit_id": unit,"shift_id": shift,"work_date": d,"scheduled_start": ts(y,m,dd,sh,0,0),"scheduled_end": ts(end_y,end_m,end_dd,end_h,0,0),"role": role,"status": "completed"})

ON_CALL_ASSIGNMENTS = [
    {"id": uid("oncall","cards-2025-w03"),  "provider_id": PROV_LISA,   "department_id": DEPT_CARDIO,  "specialty_id": uid("specialty","cardiology"),       "on_call_start": ts(2025,1,13,17,0,0), "on_call_end": ts(2025,1,20,7,0,0),  "pager_number": "+1-559-555-9201"},
    {"id": uid("oncall","cards-2025-w04"),  "provider_id": PROV_TYRELL, "department_id": DEPT_CARDIO,  "specialty_id": uid("specialty","cardiology"),       "on_call_start": ts(2025,1,20,7,0,0),  "on_call_end": ts(2025,1,27,7,0,0),  "pager_number": "+1-559-555-9201"},
    {"id": uid("oncall","ortho-2025-w03"),  "provider_id": PROV_DANIEL, "department_id": DEPT_ORTHO,   "specialty_id": uid("specialty","orthopedics"),      "on_call_start": ts(2025,1,13,17,0,0), "on_call_end": ts(2025,1,20,7,0,0),  "pager_number": "+1-559-555-9301"},
    {"id": uid("oncall","obgyn-2025-w03"),  "provider_id": PROV_SARA,   "department_id": DEPT_OBGYN,   "specialty_id": uid("specialty","ob-gyn"),           "on_call_start": ts(2025,1,13,17,0,0), "on_call_end": ts(2025,1,20,7,0,0),  "pager_number": "+1-559-555-9401"},
    {"id": uid("oncall","hospitalist-night-2025-01-13"),"provider_id": PROV_NATE, "department_id": DEPT_MEDSURG, "specialty_id": uid("specialty","internal-medicine"), "on_call_start": ts(2025,1,13,19,0,0),"on_call_end": ts(2025,1,14,7,0,0), "pager_number": "+1-559-555-9501"},
    {"id": uid("oncall","nephro-2025-w03"), "provider_id": PROV_KOFI,   "department_id": DEPT_MEDSURG, "specialty_id": uid("specialty","nephrology"),       "on_call_start": ts(2025,1,13,17,0,0), "on_call_end": ts(2025,1,20,7,0,0),  "pager_number": "+1-559-555-9601"},
]

PHARMACY_INVENTORY = [
    {"id": uid("pinv","metformin-1000"),   "medication_id": MED_METFORMIN_1000,"location_id": LOC_HOSPITAL,"lot_number": "GR240817A","expiration_date": "2026-08-31","quantity_on_hand": 8400, "reorder_level": 2000,"unit_cost": 0.04,  "last_restocked_at": ts(2025,1,3,9,0,0)},
    {"id": uid("pinv","lisinopril-20"),    "medication_id": MED_LISIN_20,      "location_id": LOC_HOSPITAL,"lot_number": "TV241015B","expiration_date": "2026-10-31","quantity_on_hand": 6200, "reorder_level": 1500,"unit_cost": 0.05,  "last_restocked_at": ts(2025,1,3,9,0,0)},
    {"id": uid("pinv","atorvastatin-80"),  "medication_id": MED_ATORV_80,      "location_id": LOC_HOSPITAL,"lot_number": "LU241201C","expiration_date": "2026-12-31","quantity_on_hand": 4100, "reorder_level": 1000,"unit_cost": 0.12,  "last_restocked_at": ts(2025,1,5,10,0,0)},
    {"id": uid("pinv","albuterol-mdi"),    "medication_id": MED_ALBUT_MDI,     "location_id": LOC_HOSPITAL,"lot_number": "TS241108D","expiration_date": "2026-04-30","quantity_on_hand": 320,  "reorder_level": 75,  "unit_cost": 18.50, "last_restocked_at": ts(2025,1,5,10,30,0)},
    {"id": uid("pinv","albuterol-neb"),    "medication_id": MED_ALBUT_NEB,     "location_id": LOC_HOSPITAL,"lot_number": "TS241128E","expiration_date": "2026-08-31","quantity_on_hand": 1850, "reorder_level": 500, "unit_cost": 0.74,  "last_restocked_at": ts(2025,1,8,11,0,0)},
    {"id": uid("pinv","aspirin-81"),       "medication_id": MED_ASA_81,        "location_id": LOC_HOSPITAL,"lot_number": "BA241105F","expiration_date": "2027-05-31","quantity_on_hand": 12400,"reorder_level": 3000,"unit_cost": 0.01,  "last_restocked_at": ts(2025,1,3,9,0,0)},
    {"id": uid("pinv","clopidogrel-75"),   "medication_id": MED_CLOPIDOGREL_75,"location_id": LOC_HOSPITAL,"lot_number": "TV241218G","expiration_date": "2026-11-30","quantity_on_hand": 2100, "reorder_level": 500, "unit_cost": 0.09,  "last_restocked_at": ts(2025,1,7,14,0,0)},
    {"id": uid("pinv","furosemide-40"),    "medication_id": MED_FUROSEMIDE_40, "location_id": LOC_HOSPITAL,"lot_number": "WC241020H","expiration_date": "2026-09-30","quantity_on_hand": 3400, "reorder_level": 800, "unit_cost": 0.06,  "last_restocked_at": ts(2025,1,3,9,30,0)},
    {"id": uid("pinv","morphine-iv-4"),    "medication_id": MED_MORPHINE_4,    "location_id": LOC_HOSPITAL,"lot_number": "HF241115J","expiration_date": "2026-06-30","quantity_on_hand": 180,  "reorder_level": 60,  "unit_cost": 1.85,  "last_restocked_at": ts(2025,1,2,8,0,0)},
    {"id": uid("pinv","oxycodone-5"),      "medication_id": MED_OXYCODONE_5,   "location_id": LOC_HOSPITAL,"lot_number": "PE241028K","expiration_date": "2027-01-31","quantity_on_hand": 940,  "reorder_level": 250, "unit_cost": 0.21,  "last_restocked_at": ts(2025,1,3,9,15,0)},
    {"id": uid("pinv","insulin-glargine"), "medication_id": MED_INSULIN_GLAR,  "location_id": LOC_HOSPITAL,"lot_number": "SA241001L","expiration_date": "2026-03-31","quantity_on_hand": 145,  "reorder_level": 40,  "unit_cost": 28.20, "last_restocked_at": ts(2025,1,3,10,0,0)},
    {"id": uid("pinv","amoxicillin-clinic"),"medication_id": MED_AMOX_500,     "location_id": LOC_CLINIC,  "lot_number": "SA241020M","expiration_date": "2026-10-31","quantity_on_hand": 720,  "reorder_level": 200, "unit_cost": 0.07,  "last_restocked_at": ts(2025,1,6,11,0,0)},
]

EQUIPMENT = [
    {"id": uid("eq","vent-1"),    "name": "Hamilton-G5 Ventilator #ICU-V01","equipment_type": "ventilator",      "manufacturer": "Hamilton Medical","model": "G5",        "serial_number": "G5-HM-218403", "location_id": LOC_HOSPITAL,"unit_id": UNIT_MICU,    "status": "in_use",    "last_maintenance_at": ts(2024,12,15,10,0,0),"next_maintenance_due": "2025-03-15"},
    {"id": uid("eq","vent-2"),    "name": "Hamilton-G5 Ventilator #ICU-V02","equipment_type": "ventilator",      "manufacturer": "Hamilton Medical","model": "G5",        "serial_number": "G5-HM-218404", "location_id": LOC_HOSPITAL,"unit_id": UNIT_MICU,    "status": "available", "last_maintenance_at": ts(2024,12,15,11,0,0),"next_maintenance_due": "2025-03-15"},
    {"id": uid("eq","monitor-1"), "name": "Philips IntelliVue MX800 #3W-M07","equipment_type": "patient_monitor","manufacturer": "Philips",         "model": "IntelliVue MX800","serial_number": "MX800-PH-541208","location_id": LOC_HOSPITAL,"unit_id": UNIT_MS_3W,   "status": "in_use",    "last_maintenance_at": ts(2024,11,20,14,0,0),"next_maintenance_due": "2025-05-20"},
    {"id": uid("eq","pump-1"),    "name": "Baxter Sigma Spectrum #BX-091",   "equipment_type": "infusion_pump",   "manufacturer": "Baxter",          "model": "Sigma Spectrum",  "serial_number": "SS-BX-784091",  "location_id": LOC_HOSPITAL,"unit_id": UNIT_MS_3W,   "status": "in_use",    "last_maintenance_at": ts(2024,10,5,9,0,0), "next_maintenance_due": "2025-04-05"},
    {"id": uid("eq","defib-1"),   "name": "Zoll R Series Defibrillator #ED-D02","equipment_type": "defibrillator","manufacturer": "Zoll Medical",    "model": "R Series",       "serial_number": "RS-ZL-309174",  "location_id": LOC_HOSPITAL,"unit_id": UNIT_ED_ACUTE,"status": "available", "last_maintenance_at": ts(2024,12,30,8,0,0),"next_maintenance_due": "2025-01-30"},
    {"id": uid("eq","ekg-1"),     "name": "GE MAC 5500 EKG Cart #ED-EKG01",  "equipment_type": "ekg_machine",     "manufacturer": "GE Healthcare",   "model": "MAC 5500",       "serial_number": "MAC-GE-682917", "location_id": LOC_HOSPITAL,"unit_id": UNIT_ED_ACUTE,"status": "available", "last_maintenance_at": ts(2024,11,2,10,0,0),"next_maintenance_due": "2025-05-02"},
    {"id": uid("eq","ct-1"),      "name": "Siemens SOMATOM Force CT",        "equipment_type": "ct_scanner",      "manufacturer": "Siemens Healthineers","model": "SOMATOM Force","serial_number": "SF-SI-104723",  "location_id": LOC_IMAGING, "unit_id": None,         "status": "available", "last_maintenance_at": ts(2024,12,10,7,0,0),"next_maintenance_due": "2025-06-10"},
    {"id": uid("eq","us-1"),      "name": "Philips EPIQ Elite Ultrasound #C-US02","equipment_type": "ultrasound",  "manufacturer": "Philips",        "model": "EPIQ Elite",     "serial_number": "EE-PH-892014",  "location_id": LOC_CLINIC,  "unit_id": None,         "status": "in_use",    "last_maintenance_at": ts(2024,11,12,13,0,0),"next_maintenance_due": "2025-05-12"},
]

TASKS = [
    {"id": uid("task","margaret-uacr"),   "assigned_provider_id": PROV_MORGAN, "created_by_provider_id": PROV_MORGAN, "patient_id": P_MARGARET,"encounter_id": ENC_MARGARET_FU,   "task_type": "lab_followup",    "title": "Repeat UACR in 3 months", "description": "Margaret Johnson - last UACR 320 mg/g 1/13. Repeat at 4/14 visit; if persistently >300 consider escalating ARB therapy.","priority": "normal","status": "pending",   "due_at": ts(2025,4,14,9,0,0),  "completed_at": None},
    {"id": uid("task","robert-lipid"),    "assigned_provider_id": PROV_LISA,   "created_by_provider_id": PROV_LISA,   "patient_id": P_ROBERT,  "encounter_id": ENC_ROBERT_CARDIO, "task_type": "lab_followup",    "title": "Recheck lipid panel 6wk", "description": "Robert Chen - on atorvastatin 80mg post-NSTEMI. Recheck LDL target <55 mg/dL on 1/24.","priority": "high",  "status": "pending",   "due_at": ts(2025,1,24,8,0,0),  "completed_at": None},
    {"id": uid("task","james-readmit"),   "assigned_provider_id": PROV_RAJESH, "created_by_provider_id": PROV_AMANDA, "patient_id": P_JAMES,   "encounter_id": ENC_JAMES_INPT,    "task_type": "transition_care", "title": "TCM call within 2 business days","description": "Reach by phone within 2 business days post-d/c 1/9. Review meds, oxygen, red-flag sx.","priority": "high",  "status": "completed", "due_at": ts(2025,1,13,17,0,0), "completed_at": ts(2025,1,13,11,30,0)},
    {"id": uid("task","dorothy-pt-followup"),"assigned_provider_id": PROV_DANIEL,"created_by_provider_id": PROV_DANIEL,"patient_id": P_DOROTHY,"encounter_id": ENC_DOROTHY_INPT,  "task_type": "referral_followup","title": "Confirm PT visits started","description": "Dorothy Kim THA — confirm home health PT 3x/week started by 11/26 per d/c plan.","priority": "high",  "status": "completed", "due_at": ts(2024,11,27,17,0,0),"completed_at": ts(2024,11,26,15,0,0)},
    {"id": uid("task","carlos-pcp"),      "assigned_provider_id": None,        "created_by_provider_id": PROV_JOHANNA, "patient_id": P_CARLOS, "encounter_id": ENC_CARLOS_ED,     "task_type": "referral",        "title": "Establish with Kaiser PCP (anxiety)","description": "ED reassurance and referral to Kaiser PCP Dr. Singh + Kaiser behavioral health for GAD f/u.","priority": "normal","status": "pending",   "due_at": ts(2025,1,15,17,0,0), "completed_at": None},
    {"id": uid("task","emily-asthma-plan"),"assigned_provider_id": PROV_MORGAN,"created_by_provider_id": PROV_HASAN,  "patient_id": P_EMILY,   "encounter_id": ENC_EMILY_ED,      "task_type": "care_plan",       "title": "Reinforce asthma action plan","description": "F/u 12/26 (post-ED). Step up to budesonide-formoterol BID. Review trigger avoidance, peak flow log.","priority": "normal","status": "completed", "due_at": ts(2024,12,26,17,0,0),"completed_at": ts(2024,12,26,11,0,0)},
    {"id": uid("task","aisha-wic-form"),  "assigned_provider_id": PROV_SARA,   "created_by_provider_id": PROV_SARA,    "patient_id": P_AISHA,   "encounter_id": ENC_AISHA_OB,      "task_type": "documentation",   "title": "Provide WIC Medical Referral Form","description": "Patient requesting WIC enrollment. Print/sign Medical Referral Form CDPH 4471 with current hemoglobin.","priority": "normal","status": "in_progress","due_at": ts(2025,1,23,17,0,0),"completed_at": None},
    {"id": uid("task","david-rx-fu"),     "assigned_provider_id": PROV_RAJESH, "created_by_provider_id": PROV_RAJESH,  "patient_id": P_DAVID,   "encounter_id": ENC_DAVID_NEWPT,   "task_type": "medication_followup","title": "Address metformin GI intolerance","description": "Pt reports GI side effects via portal 1/12. Consider metformin XR or dose reduction. Reach by phone.","priority": "normal","status": "pending",   "due_at": ts(2025,1,15,17,0,0), "completed_at": None},
    {"id": uid("task","margaret-flu"),    "assigned_provider_id": None,        "created_by_provider_id": PROV_MORGAN,  "patient_id": P_MARGARET,"encounter_id": ENC_MARGARET_FU,   "task_type": "preventive",      "title": "Schedule pneumococcal PCV20","description": "Eligible per ACIP. Schedule with MA next visit.","priority": "low",   "status": "pending",   "due_at": ts(2025,4,14,9,0,0),  "completed_at": None},
    {"id": uid("task","marcus-pt-coord"), "assigned_provider_id": None,        "created_by_provider_id": PROV_MORGAN,  "patient_id": P_MARCUS,  "encounter_id": ENC_MARCUS_LBP,    "task_type": "referral_followup","title": "Confirm PT scheduling completed","description": "Verify Springfield PT Associates contacted patient and first visit booked within 2 wk.","priority": "normal","status": "in_progress","due_at": ts(2025,1,21,17,0,0),"completed_at": None},
]

AUDIT_LOGS_SUMMARY = [
    {"id": uid("audit","margaret-access-1"),"actor_provider_id": PROV_MORGAN,"patient_id": P_MARGARET, "action": "read",  "resource_type": "encounter",         "resource_id": ENC_MARGARET_FU,    "occurred_at": ts(2025,1,13,9,42,0),  "success": True, "mongo_log_id": "audit-20250113-094200-a8c1"},
    {"id": uid("audit","margaret-result-update"),"actor_provider_id": PROV_MORGAN,"patient_id": P_MARGARET,"action": "update","resource_type": "lab_result",       "resource_id": uid("res","margaret-a1cv"),        "occurred_at": ts(2025,1,15,9,10,0), "success": True, "mongo_log_id": "audit-20250115-091012-b231"},
    {"id": uid("audit","robert-meds-access"),    "actor_provider_id": PROV_LISA,  "patient_id": P_ROBERT,  "action": "read",  "resource_type": "prescription",      "resource_id": None,                                "occurred_at": ts(2024,12,12,8,55,0),  "success": True, "mongo_log_id": "audit-20241212-085530-c918"},
    {"id": uid("audit","james-md-orders"),       "actor_provider_id": PROV_AMANDA,"patient_id": P_JAMES,   "action": "create","resource_type": "medication_order",  "resource_id": None,                                "occurred_at": ts(2025,1,4,7,15,0),    "success": True, "mongo_log_id": "audit-20250104-071512-d217"},
    {"id": uid("audit","dorothy-or-access"),     "actor_provider_id": PROV_DANIEL,"patient_id": P_DOROTHY, "action": "read",  "resource_type": "imaging_study",     "resource_id": None,                                "occurred_at": ts(2024,11,19,7,45,0),  "success": True, "mongo_log_id": "audit-20241119-074500-e042"},
    {"id": uid("audit","carlos-failed-attempt"), "actor_provider_id": None,       "patient_id": P_CARLOS,  "action": "read",  "resource_type": "encounter",         "resource_id": ENC_CARLOS_ED,      "occurred_at": ts(2025,1,5,2,17,0),   "success": False,"mongo_log_id": "audit-20250105-021745-f001"},
    {"id": uid("audit","aisha-print"),           "actor_provider_id": PROV_SARA,  "patient_id": P_AISHA,   "action": "export","resource_type": "encounter",         "resource_id": ENC_AISHA_OB,       "occurred_at": ts(2025,1,16,11,20,0), "success": True, "mongo_log_id": "audit-20250116-112048-a774"},
    {"id": uid("audit","emily-msg-send"),        "actor_provider_id": PROV_HASAN, "patient_id": P_EMILY,   "action": "create","resource_type": "inter_provider_message","resource_id": uid("ipm","emily-asthma-pcp"),  "occurred_at": ts(2024,12,16,2,0,30),  "success": True, "mongo_log_id": "audit-20241216-020030-b551"},
    {"id": uid("audit","david-rx-create"),       "actor_provider_id": PROV_RAJESH,"patient_id": P_DAVID,   "action": "create","resource_type": "prescription",      "resource_id": None,                                "occurred_at": ts(2024,12,19,14,30,0), "success": True, "mongo_log_id": "audit-20241219-143055-c194"},
    {"id": uid("audit","marcus-ref-create"),     "actor_provider_id": PROV_MORGAN,"patient_id": P_MARCUS,  "action": "create","resource_type": "referral",          "resource_id": None,                                "occurred_at": ts(2025,1,14,10,5,0),  "success": True, "mongo_log_id": "audit-20250114-100530-d873"},
]

# ---------------------------------------------------------------------------
# WRITER — emit one workbook per domain + one consolidated workbook
# ---------------------------------------------------------------------------

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DOMAIN_FILES = {
    "01_catalogs":       ["icd10_codes","cpt_codes","loinc_codes","snomed_codes","rxnorm_concepts"],
    "02_organization":   ["locations","specialties","shifts","appointment_types","payers","pharmacies","departments","units","rooms","beds","providers","provider_specialties","provider_licenses"],
    "03_patients":       ["patients","patient_identifiers","patient_addresses","patient_contacts","emergency_contacts","patient_consents"],
    "04_coverage":       ["insurance_plans","patient_coverages","authorizations"],
    "05_scheduling":     ["appointment_slots","referrals","appointments","appointment_status_history","waitlist_entries","appointment_reminders"],
    "06_encounters":     ["encounters","bed_assignments","encounter_diagnoses","encounter_procedures"],
    "07_clinical":       ["problem_list_entries","allergies","allergy_reactions","vital_signs","clinical_observations","care_plans","care_plan_goals"],
    "08_medications":    ["medications","prescriptions","medication_administrations","medication_reconciliations"],
    "09_labs_imaging":   ["lab_orders","lab_specimens","lab_results","imaging_orders","imaging_studies","imaging_reports"],
    "10_billing":        ["claims","claim_lines","charges","payments","adjustments","claim_denials","claim_appeals","patient_statements"],
    "11_communications": ["patient_message_threads","patient_messages","call_logs","insurance_correspondences","inter_provider_messages"],
    "12_operations":     ["staff_schedules","on_call_assignments","pharmacy_inventory","equipment","tasks","audit_logs_summary"],
}

DATA = {
    # 01 catalogs
    "icd10_codes":                ICD10_CODES,
    "cpt_codes":                  CPT_CODES,
    "loinc_codes":                LOINC_CODES,
    "snomed_codes":               SNOMED_CODES,
    "rxnorm_concepts":            RXNORM_CONCEPTS,
    # 02 organization
    "locations":                  LOCATIONS,
    "specialties":                SPECIALTIES,
    "shifts":                     SHIFTS,
    "appointment_types":          APPOINTMENT_TYPES,
    "payers":                     PAYERS,
    "pharmacies":                 PHARMACIES,
    "departments":                DEPARTMENTS,
    "units":                      UNITS,
    "rooms":                      ROOMS,
    "beds":                       BEDS,
    "providers":                  PROVIDERS,
    "provider_specialties":       PROVIDER_SPECIALTIES,
    "provider_licenses":          PROVIDER_LICENSES,
    # 03 patients
    "patients":                   PATIENTS,
    "patient_identifiers":        PATIENT_IDENTIFIERS,
    "patient_addresses":          PATIENT_ADDRESSES,
    "patient_contacts":           PATIENT_CONTACTS,
    "emergency_contacts":         EMERGENCY_CONTACTS,
    "patient_consents":           PATIENT_CONSENTS,
    # 04 coverage
    "insurance_plans":            INSURANCE_PLANS,
    "patient_coverages":          PATIENT_COVERAGES,
    "authorizations":             AUTHORIZATIONS,
    # 05 scheduling
    "appointment_slots":          APPOINTMENT_SLOTS,
    "referrals":                  REFERRALS,
    "appointments":               APPOINTMENTS,
    "appointment_status_history": APPOINTMENT_STATUS_HISTORY,
    "waitlist_entries":           WAITLIST_ENTRIES,
    "appointment_reminders":      APPOINTMENT_REMINDERS,
    # 06 encounters
    "encounters":                 ENCOUNTERS,
    "bed_assignments":            BED_ASSIGNMENTS,
    "encounter_diagnoses":        ENCOUNTER_DIAGNOSES,
    "encounter_procedures":       ENCOUNTER_PROCEDURES,
    # 07 clinical
    "problem_list_entries":       PROBLEM_LIST_ENTRIES,
    "allergies":                  ALLERGIES,
    "allergy_reactions":          ALLERGY_REACTIONS,
    "vital_signs":                VITAL_SIGNS,
    "clinical_observations":      CLINICAL_OBSERVATIONS,
    "care_plans":                 CARE_PLANS,
    "care_plan_goals":            CARE_PLAN_GOALS,
    # 08 medications
    "medications":                MEDICATIONS,
    "prescriptions":              PRESCRIPTIONS,
    "medication_administrations": MEDICATION_ADMINISTRATIONS,
    "medication_reconciliations": MEDICATION_RECONCILIATIONS,
    # 09 labs/imaging
    "lab_orders":                 LAB_ORDERS,
    "lab_specimens":              LAB_SPECIMENS,
    "lab_results":                LAB_RESULTS,
    "imaging_orders":             IMAGING_ORDERS,
    "imaging_studies":            IMAGING_STUDIES,
    "imaging_reports":            IMAGING_REPORTS,
    # 10 billing
    "claims":                     CLAIMS,
    "claim_lines":                CLAIM_LINES,
    "charges":                    CHARGES,
    "payments":                   PAYMENTS,
    "adjustments":                ADJUSTMENTS,
    "claim_denials":              CLAIM_DENIALS,
    "claim_appeals":              CLAIM_APPEALS,
    "patient_statements":         PATIENT_STATEMENTS,
    # 11 communications
    "patient_message_threads":    PATIENT_MESSAGE_THREADS,
    "patient_messages":           PATIENT_MESSAGES,
    "call_logs":                  CALL_LOGS,
    "insurance_correspondences":  INSURANCE_CORRESPONDENCES,
    "inter_provider_messages":    INTER_PROVIDER_MESSAGES,
    # 12 operations
    "staff_schedules":            STAFF_SCHEDULES,
    "on_call_assignments":        ON_CALL_ASSIGNMENTS,
    "pharmacy_inventory":         PHARMACY_INVENTORY,
    "equipment":                  EQUIPMENT,
    "tasks":                      TASKS,
    "audit_logs_summary":         AUDIT_LOGS_SUMMARY,
}

def _normalize(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, default=str)
    return value

def _ordered_columns(rows: list[dict]) -> list[str]:
    seen: dict[str, None] = {}
    for row in rows:
        for key in row.keys():
            seen.setdefault(key, None)
    return list(seen.keys())

def _write_sheet(ws, sheet_name: str, rows: list[dict]) -> None:
    ws.title = sheet_name[:31]
    if not rows:
        ws.append(["id"])
        return
    headers = _ordered_columns(rows)
    ws.append(headers)
    for row in rows:
        ws.append([_normalize(row.get(h)) for h in headers])
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(40, max(12, len(header) + 2))

def write_domain_workbooks(out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for domain, sheet_names in DOMAIN_FILES.items():
        wb = Workbook()
        wb.remove(wb.active)
        for sheet in sheet_names:
            rows = DATA.get(sheet, [])
            ws = wb.create_sheet(title=sheet[:31])
            _write_sheet(ws, sheet, rows)
        path = out_dir / f"{domain}.xlsx"
        wb.save(path)
        written.append(path)
    return written

def write_consolidated_workbook(path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for _, sheet_names in DOMAIN_FILES.items():
        for sheet in sheet_names:
            rows = DATA.get(sheet, [])
            ws = wb.create_sheet(title=sheet[:31])
            _write_sheet(ws, sheet, rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path

def write_summary(path: Path) -> Path:
    lines = ["# Hospital seed data summary\n"]
    grand_total = 0
    for domain, sheet_names in DOMAIN_FILES.items():
        domain_total = 0
        lines.append(f"\n## {domain}\n")
        lines.append("| Table | Rows |")
        lines.append("|---|---:|")
        for sheet in sheet_names:
            count = len(DATA.get(sheet, []))
            domain_total += count
            lines.append(f"| `{sheet}` | {count} |")
        lines.append(f"| **subtotal** | **{domain_total}** |")
        grand_total += domain_total
    lines.append(f"\n**Grand total rows: {grand_total}**\n")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path

def main() -> None:
    here = Path(__file__).resolve().parent
    by_domain = here / "by_domain"
    consolidated = here / "hospital_seed.xlsx"
    summary = here / "DATA_SUMMARY.md"
    files = write_domain_workbooks(by_domain)
    write_consolidated_workbook(consolidated)
    write_summary(summary)
    print(f"Wrote {len(files)} domain workbooks to {by_domain}")
    print(f"Wrote consolidated workbook to {consolidated}")
    print(f"Wrote summary to {summary}")
    print(f"Total tables: {sum(len(v) for v in DOMAIN_FILES.values())}")
    print(f"Total rows:   {sum(len(rows) for rows in DATA.values())}")

if __name__ == "__main__":
    main()
